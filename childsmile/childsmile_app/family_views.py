from .models import (
    Permissions,
    Role,
    Staff,
    SignedUp,
    General_Volunteer,
    Pending_Tutor,
    Tutors,
    Children,
    Tutorships,
    Matures,
    Feedback,
    Tutor_Feedback,
    General_V_Feedback,
    Tasks,
    Task_Types,
    PossibleMatches,  # Add this line
    InitialFamilyData,
    PrevTutorshipStatuses,
    SettlementsStreets
)
from .unused_views import (
    PermissionsViewSet,
    RoleViewSet,
    StaffViewSet,
    SignedUpViewSet,
    General_VolunteerViewSet,
    Pending_TutorViewSet,
    TutorsViewSet,
    ChildrenViewSet,
    TutorshipsViewSet,
    MaturesViewSet,
    FeedbackViewSet,
    Tutor_FeedbackViewSet,
    General_V_FeedbackViewSet,
    TaskViewSet,
    PossibleMatchesViewSet,
    VolunteerFeedbackReportView,
    TutorToFamilyAssignmentReportView,
    FamiliesWaitingForTutorsReportView,
    DepartedFamiliesReportView,
    NewFamiliesLastMonthReportView,
    PotentialTutorshipMatchReportView,
)
from django.db import DatabaseError, IntegrityError
from django.core.cache import cache
from rest_framework import viewsets, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action, api_view
from django.http import HttpResponse, JsonResponse, FileResponse
from django.views import View
from django.db import connection
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from django.utils.timezone import now
import datetime
import urllib3
from django.utils.timezone import make_aware
from geopy.exc import GeocoderTimedOut
from geopy.geocoders import Nominatim
import threading, time
from time import sleep
from math import sin, cos, sqrt, atan2, radians, ceil
import json
import os
import traceback
import re
import base64
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db.models import Count, F, Q , Prefetch
from difflib import get_close_matches
from .utils import *
from .audit_utils import log_api_action
from .logger import api_logger

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Load settlements and streets data from database for import validation
def get_settlements_dict():
    """
    Load settlements and streets from database table.
    Returns dict format: {city_name: [streets]}
    """
    try:
        settlements_dict = {}
        settlements = SettlementsStreets.objects.all().values('city_name', 'streets')
        for item in settlements:
            city_name = item['city_name']
            streets = item['streets'] if item['streets'] else []
            settlements_dict[city_name] = streets
        return settlements_dict
    except Exception as e:
        api_logger.warning(f"Failed to load settlements from database: {str(e)}")
        return {}

# Load settlements at startup for import validation (will be refreshed on each use)
SETTLEMENTS_N_STREETS = get_settlements_dict()


# Fallback special values dict for the 5 problematic rows that fail all other mechanisms
# Keys use * as wildcard (each part around * must appear in input)
# Values are the exact correct database values
FIVE_SPECIAL_VALUES = {
    # Cities
    'דגניה א"': 'דגניה א\'',  # Exact match with double quote
    'דגניה ב"': 'דגניה ב\'',  # Exact match with double quote
    'כרום': 'מג\'ד אל-כרום',  # If input contains "כרום", return full city name
    
    # Streets - patterns with * as wildcard (each part must be in input)
    'דולצ*אריה': 'דולצ\'ין אריה',  # If input has "דולצ" and "אריה", return this street
    'בוטינסקי': 'ז\'בוטינסקי',  # If input contains "בוטינסקי", return this street
    'פרופ*צ*חנובר*אהרון': 'פרופ׳ צ׳חנובר אהרון',  # If input has all 4 parts, return this street
}


def match_special_pattern(user_input, pattern_key):
    """
    Check if user_input matches a pattern_key from FIVE_SPECIAL_VALUES.
    Patterns use * as wildcard - all parts around * must appear in input (in order).
    
    Examples:
    - pattern: "דולצ*אריה" matches: "דולצין אריה", "דולצ׳ין אריה", etc.
    - pattern: "כרום" matches any input containing "כרום"
    - pattern: "פרופ*צ*חנובר*אהרון" matches input with all 4 parts in order
    
    Returns: True if pattern matches, False otherwise
    """
    if not user_input:
        return False
    
    if '*' not in pattern_key:
        # Simple substring match
        return pattern_key in user_input
    
    # Split by * to get parts that must appear in order
    parts = pattern_key.split('*')
    pos = 0
    
    for part in parts:
        if not part:  # Skip empty parts
            continue
        idx = user_input.find(part, pos)
        if idx == -1:
            return False
        pos = idx + len(part)
    
    return True


def get_special_value_if_matched(user_input):
    """
    Check if user_input matches any pattern in FIVE_SPECIAL_VALUES.
    If it matches, return the corresponding exact database value.
    Returns: (matched_value, was_matched) or (None, False)
    """
    if not user_input:
        return None, False
    
    for pattern_key, db_value in FIVE_SPECIAL_VALUES.items():
        if match_special_pattern(user_input, pattern_key):
            return db_value, True
    
    return None, False


def normalize_special_chars(text):
    """
    Normalize special characters for fuzzy matching.
    Removes/normalizes: apostrophes, commas, dashes, quotes, etc.
    Returns normalized text suitable for comparison.
    
    Examples:
    - "מג׳ד אל-כרום" → "מגד אל כרום"
    - "דולצ׳ין אריה" → "דולצין אריה"
    - "בן-גוריון" → "בן גוריון"
    """
    if not text:
        return ""
    
    # Replace various quote/apostrophe characters with nothing
    text = text.replace("׳", "")  # Hebrew geresh (׳)
    text = text.replace("'", "")  # Regular apostrophe (')
    text = text.replace("`", "")  # Backtick
    text = text.replace("´", "")  # Acute accent
    
    # Replace dashes/hyphens with space
    text = text.replace("-", " ")
    text = text.replace("–", " ")  # En dash
    text = text.replace("—", " ")  # Em dash
    
    # Replace commas with space
    text = text.replace(",", " ")
    
    # Normalize multiple spaces to single space
    text = " ".join(text.split())
    
    return text.strip()


def find_city_by_normalized_match(user_input, settlements_dict):
    """
    Find a city in settlements_dict by normalizing special characters.
    If exact match fails, try fuzzy match with normalized versions.
    
    Returns: (matched_city, was_corrected)
    - matched_city: The exact city name from DB (with correct punctuation)
    - was_corrected: Boolean indicating if correction was made
    """
    if not user_input or not settlements_dict:
        return None, False
    
    normalized_input = normalize_special_chars(user_input.strip())
    if not normalized_input:
        return None, False
    
    # Try exact match first (after normalization)
    for city in settlements_dict.keys():
        normalized_city = normalize_special_chars(city)
        if normalized_input == normalized_city:
            was_corrected = (user_input.strip() != city)
            return city, was_corrected
    
    # Try fuzzy matching with normalized versions
    normalized_cities = {normalize_special_chars(city): city for city in settlements_dict.keys()}
    matches = get_close_matches(normalized_input, normalized_cities.keys(), n=1, cutoff=0.75)
    
    if matches:
        matched_normalized = matches[0]
        matched_city = normalized_cities[matched_normalized]
        was_corrected = (user_input.strip() != matched_city)
        return matched_city, was_corrected
    
    return None, False


def find_street_by_normalized_match(user_input, streets_list):
    """
    Find a street in streets_list by normalizing special characters.
    Returns: (matched_street, was_corrected)
    - matched_street: The exact street name from DB (with correct punctuation)
    - was_corrected: Boolean indicating if correction was made
    """
    if not user_input or not streets_list:
        return None, False
    
    normalized_input = normalize_special_chars(user_input.strip())
    if not normalized_input:
        return None, False
    
    # Try exact match first (after normalization)
    for street in streets_list:
        normalized_street = normalize_special_chars(street)
        if normalized_input == normalized_street:
            was_corrected = (user_input.strip() != street)
            return street, was_corrected
    
    # Try fuzzy matching with normalized versions
    normalized_streets = {normalize_special_chars(street): street for street in streets_list}
    matches = get_close_matches(normalized_input, normalized_streets.keys(), n=1, cutoff=0.75)
    
    if matches:
        matched_normalized = matches[0]
        matched_street = normalized_streets[matched_normalized]
        was_corrected = (user_input.strip() != matched_street)
        return matched_street, was_corrected
    
    return None, False


def find_best_city_match(user_input):
    """
    Find the best matching city from the settlements database.
    Uses: exact mapping → normalized match (handles apostrophes/dashes) → fuzzy match
    Returns tuple: (matched_city, was_corrected)
    - matched_city: The best match from valid settlements (or original if no match)
    - was_corrected: Boolean indicating if correction was made
    
    This approach handles special characters (apostrophes, dashes, commas) in names.
    """
    if not user_input or not SETTLEMENTS_N_STREETS:
        return user_input, False
    
    # Normalize user input
    normalized_input = user_input.strip()
    if not normalized_input:
        return '', False
    
    # Reject "Israel" / "ישראל"
    if normalized_input.lower() in ['israel', 'ישראל']:
        return None, False  # Special marker for invalid
    
    # City mapping dictionary for known typos, variants, and prefix removals
    city_mapping = {
        # Direct mappings (exact replacement)
        'תל אביב': 'תל אביב - יפו',
        'מודיעין': 'מודיעין-מכבים-רעות',
        'מודעין': 'מודיעין-מכבים-רעות',  # Typo
        'פתח תקוה': 'פתח תקווה',  # Typo
        'קריית אתא': 'קרית אתא',  # Variant spelling
        'קריית נטפים': 'קרית נטפים',  # Variant spelling
        'יהוד מונוסון': 'יהוד-מונוסון',  # Space to dash
        
        # Remove prefixes (קיבוץ, מושב, יישוב)
        'קיבוץ חפץ חיים': 'חפץ חיים',
        'מושב בני ראם': 'בני ראם',
        'מושב חמד': 'חמד',
        'מושב פורת': 'פורת',
        'יד רמב״ם (מושב)': 'יד רמבם',
        'יישוב נופים': 'נופים',
        
        # Spelling variants
        'מגד אל כרום': 'מג\'ד אל-כרום',
        'מג׳ד אל כרום': 'מג\'ד אל-כרום',  # Excel geresh variant
        'מג׳ד אל-כרום': 'מג\'ד אל-כרום',  # Direct match with geresh
        
        # Comments extraction - take first part before text
        'גבעת שמואל אבל עושה שירות': 'גבעת שמואל',
        'מושה טפחות': 'טפחות',  # Typo: מושה should be מושב
        'עלי זהב לומד בשדרות': 'עלי זהב',
        'ירושלים- תא': 'ירושלים',  # Typo: תא is garbage
        'ראשל״צ': 'ראשון לציון',  # Abbreviation
        'הדר גנים': 'גנות הדר',  # Alternate name
        'רעננה(מגדל עוז)': 'רעננה',  # Remove address clarification
    }
    
    # **STEP 1: Check if exact match exists in hardcoded mapping dictionary FIRST (before normalization)**
    if normalized_input in city_mapping:
        mapped_city = city_mapping[normalized_input]
        return mapped_city, True
    
    # **STEP 2: Check for partial matches in mapping (useful for entries with comments)**
    # e.g., "גבעת שמואל אבל..." contains "גבעת שמואל"
    for key, value in city_mapping.items():
        if key in normalized_input:
            return value, True
    
    # **STEP 3: Try normalized matching (handles apostrophes, dashes, commas)**
    matched_city, was_corrected = find_city_by_normalized_match(normalized_input, SETTLEMENTS_N_STREETS)
    if matched_city:
        return matched_city, was_corrected
    
    # **STEP 4: Fallback - Check special values (for the 5 problematic rows that fail everything else)**
    special_value, was_matched = get_special_value_if_matched(normalized_input)
    if special_value:
        return special_value, True  # Mark as corrected since it's a fallback
    
    # Return original input (not in mapping, but might be valid)
    return normalized_input, False


def extract_street_and_apartment(user_input, city):
    """
    Extract street name and apartment number from combined input.
    Works like the manual modal: searches for valid street name in JSON,
    returns the matched street and remaining text as apartment number.
    Uses normalized matching to handle special characters.
    
    Returns tuple: (street_name, apartment_number, is_valid)
    - street_name: The street name found in JSON (or None if not found)
    - apartment_number: The remaining text (or empty string)
    - is_valid: True if street was found in JSON
    
    Examples:
    - "אין רחוב 1" → ("אין רחוב", "1", True)
    - "פטל 11" → ("פטל", "11", True)
    - "בנימין 276" → ("בנימין", "276", True)
    - "אבן גבירול 29" → ("אבן גבירול", "29", True)
    - "עיר תל אביב" → (None, "עיר תל אביב", False)
    """
    if not user_input or not city or not SETTLEMENTS_N_STREETS:
        return None, user_input or '', False
    
    # Normalize input
    normalized_input = user_input.strip()
    if not normalized_input:
        return None, '', False
    
    # Get streets for this city
    city_normalized = city.strip() if isinstance(city, str) else city
    streets = SETTLEMENTS_N_STREETS.get(city_normalized, [])
    if not streets:
        return None, normalized_input, False  # City not in JSON
    
    # **STEP 1: Try normalized matching (handles apostrophes, dashes, commas)**
    # This will match "דולצ׳ין אריה" even if user typed "דולצין אריה"
    matched_street, was_corrected = find_street_by_normalized_match(normalized_input, streets)
    if matched_street:
        # Extract apartment number (if any) from the input after the matched street
        apartment = normalized_input[len(normalize_special_chars(matched_street)):].strip()
        return matched_street, apartment, True
    
    # Try to find a street name that matches the beginning or is contained in the input
    # Sort by length (longest first) to match "אבן גבירול" before "אבן"
    sorted_streets = sorted(streets, key=len, reverse=True)
    
    for street in sorted_streets:
        street_stripped = street.strip()
        if not street_stripped:
            continue
        
        # Exact match
        if normalized_input == street_stripped:
            return street_stripped, '', True
        
        # Input starts with street name + space (e.g., "פטל 11" matches "פטל")
        if normalized_input.startswith(street_stripped + ' '):
            apartment = normalized_input[len(street_stripped):].strip()
            return street_stripped, apartment, True
        
        # Input starts with street name exactly (e.g., "פטל" with nothing after)
        if normalized_input.startswith(street_stripped):
            # Check if it's followed by space or digit or is exact match
            remainder = normalized_input[len(street_stripped):]
            if not remainder or remainder[0] in ' 0123456789/-':
                apartment = remainder.strip()
                return street_stripped, apartment, True
    
    # No street found in JSON
    return None, normalized_input, False


def find_best_street_match(user_input, city):
    """
    Find the best matching street for a given city.
    Uses: exact mapping → normalized match (handles apostrophes/dashes) → fuzzy match
    Returns tuple: (matched_street, was_corrected, is_placeholder)
    - matched_street: The best match from valid streets
    - was_corrected: Boolean indicating if correction was made
    - is_placeholder: True if we're using first street or placeholder
    
    This approach handles special characters (apostrophes, dashes, commas) in street names.
    """
    if not user_input or not city or not SETTLEMENTS_N_STREETS:
        return user_input, False, False
    
    # Keep original user input before normalization
    original_input = user_input.strip()
    normalized_input = original_input  # Will be re-normalized later if needed
    
    if not normalized_input:
        return None, False, False  # No street provided
    
    # Get streets for this city
    city_normalized = city.strip() if isinstance(city, str) else city
    streets = SETTLEMENTS_N_STREETS.get(city_normalized, [])
    if not streets:
        return normalized_input, False, False  # City not in JSON, return original street
    
    # **EXACT DIRECT MAPPINGS** - Excel geresh (׳) to DB apostrophe (')
    # Map EXACTLY to database values for the 5 problematic rows
    street_mapping = {
        'באר שבע': {},  # Removed - use normalized matching instead
        'בני ברק': {},  # Removed - use normalized matching instead
        'רחובות': {},  # Removed - use normalized matching instead
    }
    
    # **STEP 1: Check if exact match exists in city-specific mapping FIRST (before normalization)**
    if city_normalized in street_mapping:
        if original_input in street_mapping[city_normalized]:
            mapped_street = street_mapping[city_normalized][original_input]
            return mapped_street, True, False
    
    # Now normalize for further matching
    normalized_input = normalize_special_chars(original_input)
    
    # **STEP 2: Try normalized matching (handles apostrophes, dashes, commas)**
    matched_street, was_corrected = find_street_by_normalized_match(original_input, streets)
    if matched_street:
        return matched_street, was_corrected, False
    
    # **STEP 3: Check exact match**
    if normalized_input in streets:
        return normalized_input, False, False
    
    # **STEP 4: Try fuzzy matching**
    matches = get_close_matches(normalized_input, streets, n=1, cutoff=0.6)
    if matches:
        return matches[0], normalized_input != matches[0], False
    
    # **STEP 5: Fallback - Check special values (for the 5 problematic rows that fail everything else)**
    special_value, was_matched = get_special_value_if_matched(original_input)
    if special_value:
        return special_value, True, False  # Mark as corrected since it's a fallback
    
    # No match found
    return original_input, False, False


@conditional_csrf
@api_view(["GET"])
def get_complete_family_details(request):
    api_logger.info("get_complete_family_details called")
    """
    get all the data from children table after checking if the user has permission to view it.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='VIEW_FAMILY_DETAILS_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has VIEW permission on the "children" resource
    if not has_permission(request, "children", "VIEW"):
        log_api_action(
            request=request,
            action='VIEW_FAMILY_DETAILS_FAILED',
            success=False,
            error_message="You do not have permission to view this report",
            status_code=401
        )
        return JsonResponse(
            {"error": "You do not have permission to view this report."}, status=401
        )
    try:
        families = Children.objects.all()
        
        # MULTI-TUTOR SUPPORT: Prefetch all tutorships for efficiency
        # Only 'active' and 'pending_first_approval' exist (no second approval state)
        all_tutorships = Tutorships.objects.filter(
            tutorship_activation__in=['active', 'pending_first_approval']
        ).select_related('tutor__staff', 'tutor__id')
        
        # Build a dict of child_id -> list of tutors
        child_tutors_map = {}
        for tutorship in all_tutorships:
            child_id = tutorship.child_id
            if child_id not in child_tutors_map:
                child_tutors_map[child_id] = []
            if tutorship.tutor and tutorship.tutor.staff:
                child_tutors_map[child_id].append({
                    'tutor_id': tutorship.tutor.id_id,
                    'tutor_name': f"{tutorship.tutor.staff.first_name} {tutorship.tutor.staff.last_name}",
                    'tutor_phone': tutorship.tutor.id.phone if tutorship.tutor.id else None,  # Phone is on SignedUp, not Staff
                    'tutor_email': tutorship.tutor.staff.email,
                    'tutorship_id': tutorship.id,
                    'tutorship_status': tutorship.tutorship_activation,
                })
        
        families_data = [
            {
                "id": family.child_id,
                "first_name": family.childfirstname,
                "last_name": family.childsurname,
                "street_and_apartment_number": family.street_and_apartment_number,
                "city": family.city,
                "address": f"{family.street_and_apartment_number}, {family.city}",
                "registration_date": family.registrationdate.strftime("%d/%m/%Y"),
                "last_updated_date": family.lastupdateddate.strftime("%d/%m/%Y"),
                "gender": family.gender,
                "responsible_coordinator": get_staff_name_by_id(family.responsible_coordinator) or family.responsible_coordinator or "Unknown",
                "responsible_coordinator_id": family.responsible_coordinator if str(family.responsible_coordinator).isdigit() else "",
                "child_phone_number": family.child_phone_number if family.child_phone_number else None,
                "treating_hospital": family.treating_hospital,
                "date_of_birth": family.date_of_birth.strftime("%d/%m/%Y"),
                "medical_diagnosis": family.medical_diagnosis,
                "diagnosis_date": (
                    family.diagnosis_date.strftime("%d/%m/%Y")
                    if family.diagnosis_date
                    else None
                ),
                "marital_status": family.marital_status,
                "num_of_siblings": family.num_of_siblings if family.num_of_siblings is not None else 0,
                "details_for_tutoring": family.details_for_tutoring,
                "additional_info": family.additional_info,
                "tutoring_status": family.tutoring_status,
                "current_medical_state": family.current_medical_state,
                "when_completed_treatments": family.when_completed_treatments or None,
                "father_name": family.father_name if family.father_name else None,
                "father_phone": family.father_phone if family.father_phone else None,
                "mother_name": family.mother_name if family.mother_name else None,
                "mother_phone": family.mother_phone if family.mother_phone else None,
                "expected_end_treatment_by_protocol": family.expected_end_treatment_by_protocol or None,
                "has_completed_treatments": family.has_completed_treatments,
                "status": family.status,
                "age": family.age,
                "need_review": family.need_review,
                "is_in_frame": family.is_in_frame,
                "coordinator_comments": family.coordinator_comments,
                # MULTI-TUTOR SUPPORT: Add list of tutors
                "tutors": child_tutors_map.get(family.child_id, []),
                "tutors_count": len(child_tutors_map.get(family.child_id, [])),
            }
            for family in families
        ]

        # Fetch marital statuses only once
        marital_statuses_data = cache.get("marital_statuses_data")
        if not marital_statuses_data:
            # Get marital statuses from enum definition (not from actual data)
            marital_statuses = get_enum_values("marital_status")
            marital_statuses_data = [{"status": status} for status in marital_statuses if status]
            cache.set("marital_statuses_data", marital_statuses_data, timeout=300)

        # Fetch tutoring statuses only once
        tutoring_statuses_data = cache.get("tutoring_statuses_data")
        if not tutoring_statuses_data:
            # Get tutoring statuses from enum definition (not from actual data)
            tutoring_statuses = get_enum_values("tutoring_status")
            tutoring_statuses_data = [{"status": status} for status in tutoring_statuses if status]
            cache.set("tutoring_statuses_data", tutoring_statuses_data, timeout=300)

        # Fetch statuses only once
        statuses_data = cache.get("statuses_data")
        if not statuses_data:
            statuses = get_enum_values("status")
            statuses_data = [{"status": status} for status in statuses]
            cache.set("statuses_data", statuses_data, timeout=300)

        return JsonResponse(
            {
                "families": families_data,
                "marital_statuses": marital_statuses_data,
                "tutoring_statuses": tutoring_statuses_data,
                "statuses": statuses_data,
            },
            status=200,
        )
    except Exception as e:
        api_logger.error(f"An error occurred: {str(e)}")
        log_api_action(
            request=request,
            action='VIEW_FAMILY_DETAILS_FAILED',
            success=False,
            error_message=str(e),
            status_code=500
        )
        return JsonResponse({"error": str(e)}, status=500)

@conditional_csrf
@api_view(["POST"])
def create_family(request):
    api_logger.info("create_family called")
    """
    Create a new family in the children table after checking if the user has permission to create it.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='CREATE_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403,
            additional_data={'family_name': 'Unknown'}  # **ADD THIS LINE**
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has CREATE permission on the "children" resource
    if not has_permission(request, "children", "CREATE"):
        log_api_action(
            request=request,
            action='CREATE_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to create a family",
            status_code=401,
            additional_data={'family_name': 'Unknown'}  # **ADD THIS LINE**
        )
        return JsonResponse(
            {"error": "You do not have permission to create a family."}, status=401
        )

    try:
        # Extract data from the request
        data = request.data


        father_phone = data.get("father_phone")
        mother_phone = data.get("mother_phone")
        
        has_father = is_nonempty_phone(father_phone)
        has_mother = is_nonempty_phone(mother_phone)

        if not has_father and not has_mother:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="At least one parent phone number must be provided",
                status_code=400,
                additional_data={'family_name': f"{data.get('childfirstname', 'Unknown')} {data.get('childsurname', 'Unknown')}"}
            )
            return JsonResponse(
                {"error": "At least one parent phone number (father or mother) must be provided."},
                status=400,
            )

        # Validate required fields
        required_fields = [
            "child_id",
            "childfirstname",
            "childsurname",
            "gender",
            "city",
            "treating_hospital",
            "date_of_birth",
            "marital_status",
            "num_of_siblings",
            "details_for_tutoring",
            "marital_status",
            "tutoring_status",
            "street_and_apartment_number",
            "status",
        ]
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message=f"Missing required fields: {', '.join(missing_fields)}",
                status_code=400,
                additional_data={'family_name': f"{data.get('childfirstname', 'Unknown')} {data.get('childsurname', 'Unknown')}"}  # **ADD THIS LINE**
            )
            return JsonResponse(
                {"error": f"Missing required fields: {', '.join(missing_fields)}"},
                status=400,
            )

        # if one of the required fields is invalid, return an error response
        if not is_valid_date(data.get("date_of_birth")):
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="Invalid date of birth",
                status_code=400,
                additional_data={'family_name': f"{data.get('childfirstname', 'Unknown')} {data.get('childsurname', 'Unknown')}"}  # **ADD THIS LINE**
            )
            return JsonResponse(
                {"error": "Invalid date of birth"},
                status=400,
            )
        # Update the validation in create_family:
        if not is_valid_bigint_child_id(data.get("child_id")):
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="Invalid child_id - must be exactly 9 digits",
                status_code=400,
                additional_data={'family_name': f"{data.get('childfirstname', 'Unknown')} {data.get('childsurname', 'Unknown')}"}  # **ADD THIS LINE**
            )
            return JsonResponse(
                {"error": "Invalid child_id - must be exactly 9 digits"},
                status=400,
            )

        #if "למצוא_חונך" in data.get("tutoring_status", ""):
            # get the name of the responsible coordinator that has the role of 'Tutored Families Coordinator' else get the name of the responsible coordinator with the role of 'Family Coordinator' using not the user_id but getting it from the Staff table
                # Get the correct responsible coordinator based on tutoring status
        tutoring_status = data.get("tutoring_status", "")
        child_status = data.get("status", "")
        # Pass child_status to handle בריא/ז״ל children (they get "ללא")
        coordinator_staff_id = get_responsible_coordinator_for_family(tutoring_status, child_status=child_status)
        responsible_coordinator = coordinator_staff_id if coordinator_staff_id else user_id
        
        # Feature #3: Determine need_review for new child based on age and status
        # Set to False if: בריא/ז״ל status OR בוגר tutoring_status OR age >= 16
        date_of_birth = parse_date_field(data.get("date_of_birth"), "date_of_birth")
        need_review_value = True  # Default to True
        
        # Check age-based maturity (>= 16)
        if date_of_birth:
            from datetime import date
            today = date.today()
            age = (
                today.year
                - date_of_birth.year
                - (
                    (today.month, today.day)
                    < (date_of_birth.month, date_of_birth.day)
                )
            )
            if age >= 16:
                need_review_value = False
        
        # Check status and tutoring_status for other permanent conditions
        if child_status in ['בריא', 'ז״ל'] or tutoring_status == 'בוגר':
            need_review_value = False
        elif data.get("need_review") is not None:
            # Allow explicit override only if not in permanent condition
            need_review_value = data.get("need_review", True)

        # Create a new family record in the database
        family = Children.objects.create(
            child_id=data["child_id"],
            childfirstname=data["childfirstname"],
            childsurname=data["childsurname"],
            registrationdate=datetime.datetime.now(),
            lastupdateddate=datetime.datetime.now(),
            gender=True if data["gender"] == "נקבה" else False,
            responsible_coordinator=responsible_coordinator,
            city=data["city"],
            child_phone_number=data["child_phone_number"] if data.get("child_phone_number") else None,
            treating_hospital=data["treating_hospital"],
            date_of_birth=date_of_birth,
            medical_diagnosis=data.get("medical_diagnosis"),
            diagnosis_date=(
                data.get("diagnosis_date") if data.get("diagnosis_date") else None
            ),
            marital_status=data["marital_status"],
            num_of_siblings=data["num_of_siblings"],
            details_for_tutoring=(
                data["details_for_tutoring"]
                if data.get("details_for_tutoring")
                else "לא_רלוונטי"
            ),
            additional_info=data.get("additional_info"),
            tutoring_status=(
                data["tutoring_status"] if data.get("tutoring_status") else "לא_רלוונטי"
            ),
            current_medical_state=data.get("current_medical_state"),
            when_completed_treatments=(
                data.get("when_completed_treatments")
                if data.get("when_completed_treatments")
                else None
            ),
            father_name=data.get("father_name"),
            father_phone=data.get("father_phone"),
            mother_name=data.get("mother_name"),
            mother_phone=data.get("mother_phone"),
            street_and_apartment_number=data.get("street_and_apartment_number"),
            expected_end_treatment_by_protocol=(
                data.get("expected_end_treatment_by_protocol")
                if data.get("expected_end_treatment_by_protocol")
                else None
            ),
            has_completed_treatments=data.get("has_completed_treatments", False),
            status=(data["status"] if data.get("status") else "טיפולים"),
            is_in_frame=data.get("is_in_frame"),
            coordinator_comments=data.get("coordinator_comments"),
            # Feature #2 + #3: Set need_review based on age, status, and tutoring_status
            need_review=need_review_value,
        )

        # Log successful family creation
        log_api_action(
            request=request,
            action='CREATE_FAMILY_SUCCESS',
            affected_tables=['childsmile_app_children'],
            entity_type='Children',
            entity_ids=[family.child_id],
            success=True,
            additional_data={
                'family_name': f"{data['childfirstname']} {data['childsurname']}",
                'family_city': data['city'],
                'responsible_coordinator': get_staff_name_by_id(responsible_coordinator) or "Unknown"
            }
        )

        return JsonResponse(
            {"message": "Family created successfully", "ID": family.child_id},
            status=201,
        )
    except Exception as e:
        api_logger.error(f"An error occurred while creating a family: {str(e)}")
        log_api_action(
            request=request,
            action='CREATE_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'family_name': f"{data.get('childfirstname', 'Unknown')} {data.get('childsurname', 'Unknown')}" if 'data' in locals() else 'Unknown'}  # **ADD THIS LINE**
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["PUT"])
@transaction.atomic
def update_family(request, child_id):
    """
    Update an existing family in the children table and propagate changes to related tables.
    """
    api_logger.info(f"update_family called for child_id: {child_id}")
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403,
            additional_data={'family_name': 'Unknown'}  # **ADD THIS LINE**
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has UPDATE permission on the "children" resource
    if not has_permission(request, "children", "UPDATE"):
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to update this family",
            status_code=401,
            entity_type='Children',
            entity_ids=[child_id],
            additional_data={'family_name': 'Unknown'}  # **ADD THIS LINE**
        )
        return JsonResponse(
            {"error": "You do not have permission to update this family."}, status=401
        )

    try:
        # Fetch the existing family record
        try:
            family = Children.objects.get(child_id=child_id)
        except Children.DoesNotExist:
            log_api_action(
                request=request,
                action='UPDATE_FAMILY_FAILED',
                success=False,
                error_message="Family not found",
                status_code=404,
                entity_type='Children',
                entity_ids=[child_id],
                additional_data={'family_name': 'Unknown - Family not found'}
            )
            return JsonResponse({"error": "Family not found."}, status=404)

        # Extract data from the request
        data = request.data
        
        # Initialize field_changes early so it's available in all error handlers
        field_changes = []

        # Validate that the child_id in the request matches the existing child_id
        # UNLESS the user is intentionally changing the ID (will be handled by update_child_id endpoint)
        request_child_id = data.get("child_id")
        if request_child_id and str(request_child_id) != str(child_id):
            # Check if this is an intentional ID change
            # If the IDs are different and both are valid 9-digit numbers, allow it
            # (the actual ID change will be handled by the update_child_id endpoint after other fields are saved)
            request_child_id_str = str(request_child_id).strip()
            if not (request_child_id_str.isdigit() and len(request_child_id_str) == 9):
                log_api_action(
                    request=request,
                    action='UPDATE_FAMILY_FAILED',
                    success=False,
                    error_message="The child_id in the request does not match the existing child_id",
                    status_code=400,
                    entity_type='Children',
                    entity_ids=[child_id],
                    additional_data={
                        'family_name': f"{family.childfirstname} {family.childsurname}",
                        'attempted_changes': field_changes,
                        'changes_count': len(field_changes)
                    }
                )
                return JsonResponse(
                    {
                        "error": "The child_id in the request does not match the existing child_id."
                    },
                    status=400,
                )
            # If it's a valid ID format change, allow it to proceed (will be changed later via update_child_id endpoint)

        father_phone = data.get("father_phone")
        mother_phone = data.get("mother_phone")
        
        has_father = is_nonempty_phone(father_phone)
        has_mother = is_nonempty_phone(mother_phone)

        if not has_father and not has_mother:
            log_api_action(
                request=request,
                action='UPDATE_FAMILY_FAILED',
                success=False,
                error_message="At least one parent phone number must be provided",
                status_code=400,
                entity_type='Children',
                entity_ids=[child_id],
                additional_data={
                    'family_name': f"{family.childfirstname} {family.childsurname}",
                    'attempted_changes': [],
                    'changes_count': 0
                }
            )
            return JsonResponse(
                {"error": "At least one parent phone number (father or mother) must be provided."},
                status=400,
            )

        required_fields = [
            "child_id",
            "childfirstname",
            "childsurname",
            "gender",
            "city",
            "treating_hospital",
            "date_of_birth",
            "marital_status",
            "num_of_siblings",
            "details_for_tutoring",
            "marital_status",
            "tutoring_status",
            "street_and_apartment_number",
            "status",
        ]
        missing_fields = [field for field in required_fields if field not in data]
        if missing_fields:
            log_api_action(
                request=request,
                action='UPDATE_FAMILY_FAILED',
                success=False,
                error_message=f"Missing required fields: {', '.join(missing_fields)}",
                status_code=400,
                entity_type='Children',
                entity_ids=[child_id],
                additional_data={
                    'family_name': f"{family.childfirstname} {family.childsurname}",
                    'attempted_changes': field_changes,
                    'changes_count': len(field_changes)
                }
            )
            return JsonResponse(
                {"error": f"Missing required fields: {', '.join(missing_fields)}"},
                status=400,
            )

        # Store original values for audit
        original_name = f"{family.childfirstname} {family.childsurname}"
        
        # Track what fields are being changed - BEFORE updating
        # Check each field for changes and track them
        if family.childfirstname != data.get("childfirstname", family.childfirstname):
            field_changes.append(f"First Name: '{family.childfirstname}' → '{data.get('childfirstname')}'")
        if family.childsurname != data.get("childsurname", family.childsurname):
            field_changes.append(f"Last Name: '{family.childsurname}' → '{data.get('childsurname')}'")
        
        # Handle gender comparison
        new_gender = True if data.get("gender") == "נקבה" else False
        if family.gender != new_gender:
            old_gender = "נקבה" if family.gender else "זכר"
            new_gender_text = "נקבה" if new_gender else "זכר"
            field_changes.append(f"Gender: '{old_gender}' → '{new_gender_text}'")
            
        if family.city != data.get("city", family.city):
            field_changes.append(f"City: '{family.city}' → '{data.get('city')}'")
        if family.child_phone_number != data.get("child_phone_number", family.child_phone_number):
            field_changes.append(f"Phone: '{family.child_phone_number}' → '{data.get('child_phone_number')}'")
        if family.treating_hospital != data.get("treating_hospital", family.treating_hospital):
            field_changes.append(f"Hospital: '{family.treating_hospital}' → '{data.get('treating_hospital')}'")
        if family.medical_diagnosis != data.get("medical_diagnosis", family.medical_diagnosis):
            field_changes.append(f"Medical Diagnosis: '{family.medical_diagnosis}' → '{data.get('medical_diagnosis')}'")
        if family.marital_status != data.get("marital_status", family.marital_status):
            field_changes.append(f"Marital Status: '{family.marital_status}' → '{data.get('marital_status')}'")
        if family.num_of_siblings != data.get("num_of_siblings", family.num_of_siblings):
            field_changes.append(f"Siblings: '{family.num_of_siblings}' → '{data.get('num_of_siblings')}'")
        if family.details_for_tutoring != data.get("details_for_tutoring", family.details_for_tutoring):
            field_changes.append(f"Tutoring Details: '{family.details_for_tutoring}' → '{data.get('details_for_tutoring')}'")
        if family.additional_info != data.get("additional_info", family.additional_info):
            field_changes.append(f"Additional Info: '{family.additional_info}' → '{data.get('additional_info')}'")
        if family.tutoring_status != data.get("tutoring_status", family.tutoring_status):
            field_changes.append(f"Tutoring Status: '{family.tutoring_status}' → '{data.get('tutoring_status')}'")
        if family.current_medical_state != data.get("current_medical_state", family.current_medical_state):
            field_changes.append(f"Medical State: '{family.current_medical_state}' → '{data.get('current_medical_state')}'")
        if family.father_name != data.get("father_name", family.father_name):
            field_changes.append(f"Father Name: '{family.father_name}' → '{data.get('father_name')}'")
        if family.father_phone != data.get("father_phone", family.father_phone):
            field_changes.append(f"Father Phone: '{family.father_phone}' → '{data.get('father_phone')}'")
        if family.mother_name != data.get("mother_name", family.mother_name):
            field_changes.append(f"Mother Name: '{family.mother_name}' → '{data.get('mother_name')}'")
        if family.mother_phone != data.get("mother_phone", family.mother_phone):
            field_changes.append(f"Mother Phone: '{family.mother_phone}' → '{data.get('mother_phone')}'")
        if family.street_and_apartment_number != data.get("street_and_apartment_number", family.street_and_apartment_number):
            field_changes.append(f"Address: '{family.street_and_apartment_number}' → '{data.get('street_and_apartment_number')}'")
        if family.has_completed_treatments != data.get("has_completed_treatments", family.has_completed_treatments):
            field_changes.append(f"Completed Treatments: '{family.has_completed_treatments}' → '{data.get('has_completed_treatments')}'")
        if family.status != data.get("status", family.status):
            field_changes.append(f"Status: '{family.status}' → '{data.get('status')}'")
        if family.is_in_frame != data.get("is_in_frame", family.is_in_frame):
            field_changes.append(f"In Frame: '{family.is_in_frame}' → '{data.get('is_in_frame')}'")
        if family.coordinator_comments != data.get("coordinator_comments", family.coordinator_comments):
            field_changes.append(f"Coordinator Comments: '{family.coordinator_comments}' → '{data.get('coordinator_comments')}'")

        # Handle date fields separately (compare parsed dates)
        new_date_of_birth = parse_date_field(data.get("date_of_birth"), "date_of_birth")
        if family.date_of_birth != new_date_of_birth:
            field_changes.append(f"Birth Date: '{family.date_of_birth}' → '{new_date_of_birth}'")
        
        new_registration_date = parse_date_field(data.get("registration_date"), "registration_date")
        if family.registrationdate != new_registration_date:
            field_changes.append(f"Registration Date: '{family.registrationdate}' → '{new_registration_date}'")
            
        new_diagnosis_date = parse_date_field(data.get("diagnosis_date"), "diagnosis_date")
        if family.diagnosis_date != new_diagnosis_date:
            field_changes.append(f"Diagnosis Date: '{family.diagnosis_date}' → '{new_diagnosis_date}'")
            
        new_treatment_end = data.get("when_completed_treatments")
        if family.when_completed_treatments != new_treatment_end:
            field_changes.append(f"Treatment End: '{family.when_completed_treatments}' → '{new_treatment_end}'")
            
        new_expected_end = data.get("expected_end_treatment_by_protocol")
        if family.expected_end_treatment_by_protocol != new_expected_end:
            field_changes.append(f"Expected End: '{family.expected_end_treatment_by_protocol}' → '{new_expected_end}'")

        # **KEEP ALL EXISTING FIELD UPDATES AS THEY ARE** - Update fields in the Children table
        family.childfirstname = data.get("childfirstname", family.childfirstname)
        family.childsurname = data.get("childsurname", family.childsurname)
        family.gender = True if data.get("gender") == "נקבה" else False
        family.city = data.get("city", family.city)
        family.child_phone_number = data.get("child_phone_number", family.child_phone_number)
        family.treating_hospital = data.get("treating_hospital", family.treating_hospital)
        family.date_of_birth = parse_date_field(data.get("date_of_birth"), "date_of_birth")
        family.registrationdate = parse_date_field(data.get("registration_date"), "registration_date") or family.registrationdate
        family.medical_diagnosis = data.get("medical_diagnosis", family.medical_diagnosis)
        family.diagnosis_date = parse_date_field(data.get("diagnosis_date"), "diagnosis_date")
        family.marital_status = data.get("marital_status", family.marital_status)
        family.num_of_siblings = data.get("num_of_siblings", family.num_of_siblings)
        family.details_for_tutoring = data.get("details_for_tutoring", family.details_for_tutoring)
        family.additional_info = data.get("additional_info", family.additional_info)
        family.current_medical_state = data.get("current_medical_state", family.current_medical_state)
        family.when_completed_treatments = data.get("when_completed_treatments")
        family.father_name = data.get("father_name", family.father_name)
        family.father_phone = data.get("father_phone", family.father_phone)
        family.mother_name = data.get("mother_name", family.mother_name)
        family.mother_phone = data.get("mother_phone", family.mother_phone)
        family.street_and_apartment_number = data.get("street_and_apartment_number", family.street_and_apartment_number)
        family.expected_end_treatment_by_protocol = data.get("expected_end_treatment_by_protocol")
        family.has_completed_treatments = data.get("has_completed_treatments", family.has_completed_treatments)
        family.is_in_frame = data.get("is_in_frame", family.is_in_frame)
        family.coordinator_comments = data.get("coordinator_comments", family.coordinator_comments)
        
        # Handle tutoring_status change and auto-update coordinator
        old_tutoring_status = family.tutoring_status
        new_tutoring_status = data.get("tutoring_status", family.tutoring_status)
        family.tutoring_status = new_tutoring_status
        
        # Track old status for the coordinator logic
        old_status = family.status
        new_status = data.get("status", family.status)
        family.status = new_status
        
        # Auto-update responsible_coordinator if tutoring_status changed OR if status changed to/from בריא/ז״ל
        # Only auto-update if coordinator wasn't manually set in the request
        should_auto_update_coordinator = (
            (old_tutoring_status != new_tutoring_status) or 
            (old_status != new_status)
        ) and "responsible_coordinator" not in data
        
        if should_auto_update_coordinator:
            # Pass both tutoring_status and child status to determine coordinator
            new_coordinator_id = get_responsible_coordinator_for_family(new_tutoring_status, child_status=new_status)
            if new_coordinator_id:
                old_coordinator_name = get_staff_name_by_id(family.responsible_coordinator)
                family.responsible_coordinator = new_coordinator_id
                new_coordinator_name = get_staff_name_by_id(new_coordinator_id)
                if old_status != new_status:
                    field_changes.append(f"Responsible Coordinator (auto-updated): '{old_coordinator_name}' → '{new_coordinator_name}' (due to status change)")
                else:
                    field_changes.append(f"Responsible Coordinator (auto-updated): '{old_coordinator_name}' → '{new_coordinator_name}' (due to tutoring status change)")
        elif "responsible_coordinator" in data:
            # Allow manual override of responsible_coordinator
            new_coordinator = data.get("responsible_coordinator")
            if family.responsible_coordinator != new_coordinator:
                old_coordinator_name = get_staff_name_by_id(family.responsible_coordinator)
                family.responsible_coordinator = new_coordinator
                new_coordinator_name = get_staff_name_by_id(new_coordinator)
                field_changes.append(f"Responsible Coordinator (manual): '{old_coordinator_name}' → '{new_coordinator_name}'")
        
        # DECEASED/HEALTHY HANDLING: When child status changes to deceased or healthy, 
        # DELETE all tutorships (not make inactive - that's for inactive staff flow only)
        deceased_statuses = ['ז״ל']  # Only the correct enum value with Hebrew gershayim
        exit_statuses = deceased_statuses + ['בריא']  # Include healthy
        tutors_freed = []
        if new_status in exit_statuses and old_status not in exit_statuses:
            # Child is deceased or healthy - DELETE tutorships and free tutors
            active_tutorships = Tutorships.objects.filter(
                child_id=child_id,
                tutorship_activation__in=['active', 'pending_first_approval']
            )
            for tutorship in active_tutorships:
                tutor = tutorship.tutor
                if tutor:
                    # Mark tutor as available
                    tutor.tutorship_status = 'אין_חניך'
                    tutor.save()
                    tutors_freed.append({
                        'tutor_id': tutor.id_id,
                        'tutor_name': f"{tutor.staff.first_name} {tutor.staff.last_name}" if tutor.staff else 'Unknown'
                    })
                # DELETE the tutorship (not inactive - that's for staff deactivation only)
                tutorship.delete()
            
            # Also clean up PrevTutorshipStatuses for this child
            PrevTutorshipStatuses.objects.filter(child_id=child_id).delete()
            
            if tutors_freed:
                status_reason = "deceased" if new_status in deceased_statuses else "healthy"
                field_changes.append(f"Tutors freed due to {status_reason} status: {[t['tutor_name'] for t in tutors_freed]}")
                api_logger.info(f"Child {child_id} status changed to {status_reason}. Freed {len(tutors_freed)} tutors: {tutors_freed}")
        
        # Feature #2 + #3: Handle need_review for בריא/ז״ל children AND age >= 16 maturity
        # CRITICAL RULE: Once need_review=False due to age (>= 16) or status (בריא/ז״ל/בוגר), 
        # it MUST NOT revert back to True. These are permanent lifecycle events.
        
        old_need_review = family.need_review
        
        # Check if child is mature (age >= 16) - if so, need_review MUST be False
        from .utils import check_and_handle_age_maturity
        maturity_check = check_and_handle_age_maturity(family)
        is_mature = maturity_check.get('mature', False)
        
        # Check if tutoring_status is בוגר (mature) - if so, need_review MUST be False
        is_mature_tutoring = new_tutoring_status == 'בוגר'
        
        # Determine if need_review MUST be False (permanent, cannot be overridden)
        must_be_false = is_mature or is_mature_tutoring or new_status in exit_statuses
        
        if must_be_false:
            # These are permanent conditions - need_review MUST be False and cannot be changed
            family.need_review = False
        elif old_status in exit_statuses and new_status not in exit_statuses:
            # Child status changed FROM בריא/ז״ל to something else
            # Only reset to True if was True before and there's no other permanent condition
            if old_need_review and not is_mature and not is_mature_tutoring:
                family.need_review = True
        
        # Allow manual override ONLY if none of the permanent conditions apply
        if "need_review" in data and not must_be_false:
            family.need_review = data.get("need_review", True)
        elif "need_review" in data and must_be_false:
            # Request tries to override permanent condition - log warning and ignore
            api_logger.warning(f"Attempt to override permanent need_review=False for child {child_id} (age: {maturity_check.get('age', '?')}, mature_tutoring: {is_mature_tutoring})")
        
        # If need_review changed to False, delete all review tasks for this child
        if old_need_review and not family.need_review:
            review_task_type = Task_Types.objects.filter(task_type='שיחת ביקורת').first()
            if review_task_type:
                deleted_tasks = Tasks.objects.filter(
                    related_child_id=child_id,
                    task_type=review_task_type
                ).delete()
                if deleted_tasks[0] > 0:
                    reason = ""
                    if is_mature:
                        reason = f" (age >= 16)"
                    elif is_mature_tutoring:
                        reason = f" (tutoring_status: בוגר)"
                    elif new_status in exit_statuses:
                        reason = f" (status: {new_status})"
                    field_changes.append(f"Deleted {deleted_tasks[0]} review tasks (need_review set to False{reason})")
                    api_logger.info(f"Deleted {deleted_tasks[0]} review tasks for child {child_id}{reason}")
        
        # Track the need_review change for audit log
        if old_need_review != family.need_review:
            field_changes.append(f"Need Review: {old_need_review} → {family.need_review}")

        
        # Save the updated family record
        try:
            family.save()
            api_logger.debug(f"Family with child_id {child_id} saved successfully.")
        except DatabaseError as db_error:
            api_logger.error(f"Database error while saving family: {str(db_error)}")
            log_api_action(
                request=request,
                action='UPDATE_FAMILY_FAILED',
                success=False,
                error_message=f"Database error: {str(db_error)}",
                status_code=500,
                entity_type='Children',
                entity_ids=[child_id],
                additional_data={
                    'family_name': f"{family.childfirstname} {family.childsurname}",
                    'attempted_changes': field_changes,  # **ADD THIS**
                    'changes_count': len(field_changes)  # **ADD THIS**
                }
            )
            return JsonResponse(
                {"error": f"Database error: {str(db_error)}"}, status=500
            )

        # Propagate changes to related tables
        Tasks.objects.filter(related_child_id=child_id).update(
            updated_at=datetime.datetime.now(),
        )

        # Update tutor's tutee_wellness and relationship_status if tutorship exists
        # MULTI-TUTOR SUPPORT: Update ALL tutors for this child
        # Only 'active' and 'pending_first_approval' exist (no second approval state)
        tutorships = Tutorships.objects.filter(
            child_id=child_id,
            tutorship_activation__in=['active', 'pending_first_approval']
        )
        for tutorship in tutorships:
            if tutorship.tutor_id:
                Tutors.objects.filter(id_id=tutorship.tutor_id).update(
                    tutee_wellness=family.current_medical_state,
                    relationship_status=family.marital_status
                )

        # Log successful family update
        affected_tables_list = ['childsmile_app_children', 'childsmile_app_tasks', 'childsmile_app_tutors']
        if tutors_freed:
            affected_tables_list.append('childsmile_app_tutorships')
        
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_SUCCESS',
            affected_tables=affected_tables_list,
            entity_type='Children',
            entity_ids=[family.child_id],
            success=True,
            additional_data={
                'updated_family_name': f"{family.childfirstname} {family.childsurname}",
                'original_family_name': original_name,
                'family_city': family.city,
                'family_status': family.status,
                'field_changes': field_changes,  # **ADD THIS**
                'changes_count': len(field_changes),  # **ADD THIS**
                'tutors_freed': tutors_freed if tutors_freed else None
            }
        )

        api_logger.debug(f"Family with child_id {child_id} updated successfully.")

        return JsonResponse(
            {
                "message": "Family updated successfully",
                "ID": family.child_id,
            },
            status=200,
        )
    except Exception as e:
        api_logger.error(f"An error occurred while updating the family: {str(e)}")
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='Children',
            entity_ids=[child_id],
            additional_data={
                'family_name': f"{family.childfirstname} {family.childsurname}" if 'family' in locals() else 'Unknown',
                'attempted_changes': field_changes if 'field_changes' in locals() else [],  # **ADD THIS**
                'changes_count': len(field_changes) if 'field_changes' in locals() else 0  # **ADD THIS**
            }
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["DELETE"])
def delete_family(request, child_id):
    api_logger.info(f"delete_family called for child_id: {child_id}")
    """
    Delete a family from the children table after checking if the user has permission to delete it.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='DELETE_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has DELETE permission on the "children" resource
    if not has_permission(request, "children", "DELETE"):
        log_api_action(
            request=request,
            action='DELETE_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to delete this family",
            status_code=401,
            entity_type='Children',
            entity_ids=[child_id]
        )
        return JsonResponse(
            {"error": "You do not have permission to delete this family."}, status=401
        )

    try:
        # Fetch the existing family record
        try:
            family = Children.objects.get(child_id=child_id)
            family_name = f"{family.childfirstname} {family.childsurname}"
        except Children.DoesNotExist:
            log_api_action(
                request=request,
                action='DELETE_FAMILY_FAILED',
                success=False,
                error_message="Family not found",
                status_code=404,
                entity_type='Children',
                entity_ids=[child_id],
                additional_data={'family_name': 'Unknown - Family not found'}  # **ADD THIS**
            )
            return JsonResponse({"error": "Family not found."}, status=404)

        # Delete the family record
        family.delete()

        # delete related records in childsmile_app_tasks
        Tasks.objects.filter(related_child_id=child_id).delete()
        api_logger.debug(f"Related tasks for child_id {child_id} deleted.")

        # delete related records in childsmile_app_tutorships
        Tutorships.objects.filter(child_id=child_id).delete()
        api_logger.debug(f"Related tutorship records for child_id {child_id} deleted.")

        # Log successful family deletion
        log_api_action(
            request=request,
            action='DELETE_FAMILY_SUCCESS',
            affected_tables=['childsmile_app_children', 'childsmile_app_tasks', 'childsmile_app_tutorships'],
            entity_type='Children',
            entity_ids=[child_id],
            success=True,
            additional_data={
                'deleted_family_name': family_name,
                'family_first_name': family.childfirstname,  # **ADD THIS**
                'family_last_name': family.childsurname,  # **ADD THIS**
                'family_city': family.city,  # **ADD THIS**
                'family_status': family.status,  # **ADD THIS**
                'family_phone': family.child_phone_number,  # **ADD THIS**
                'family_hospital': family.treating_hospital,  # **ADD THIS**
                'family_marital_status': family.marital_status,  # **ADD THIS**
                'family_tutoring_status': family.tutoring_status,  # **ADD THIS**
                'date_of_birth': family.date_of_birth.strftime("%d/%m/%Y") if family.date_of_birth else None,  # **ADD THIS**
                'medical_diagnosis': family.medical_diagnosis,  # **ADD THIS**
                'current_medical_state': family.current_medical_state  # **ADD THIS**
            }
        )
        return JsonResponse(
            {"message": "Family deleted successfully", "ID": child_id},
            status=200,
        )

    except Exception as e:
        api_logger.error(f"An error occurred while deleting the family: {str(e)}")
        log_api_action(
            request=request,
            action='DELETE_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='Children',
            entity_ids=[child_id],
            additional_data={
                'deleted_family_name': family_name if 'family' in locals() else 'Unknown',  # **ADD THIS**
                'family_first_name': family.childfirstname if 'family' in locals() else 'Unknown',  # **ADD THIS**
                'family_last_name': family.childsurname if 'family' in locals() else 'Unknown',  # **ADD THIS**
                'family_city': family.city if 'family' in locals() else 'Unknown',  # **ADD THIS**
                'family_status': family.status if 'family' in locals() else 'Unknown'  # **ADD THIS**
            }
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["GET"])
def get_initial_family_data(request):
    api_logger.info("get_initial_family_data called")
    """
    Retrieve all initial family data from the InitialFamilyData model.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='VIEW_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has VIEW permission on the "initial_family_data" resource
    if not has_initial_family_data_permission(request, "view"):
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='VIEW_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to view this data",
            status_code=401
        )
        return JsonResponse(
            {"error": "You do not have permission to view this data."}, status=401
        )
    try:
        initial_family_data = InitialFamilyData.objects.all()
        data = [
            {
                "initial_family_data_id": item.initial_family_data_id,
                "names": item.names,
                "phones": item.phones,
                "other_information": item.other_information,
                "created_at": item.created_at,
                "updated_at": item.updated_at,
                "family_added": item.family_added,
            }
            for item in initial_family_data
        ]
        return JsonResponse({"initial_family_data": data}, status=200)
    except Exception as e:
        api_logger.error(
            f"An error occurred while retrieving initial family data: {str(e)}"
        )
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='VIEW_INITIAL_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500
        )
        return JsonResponse({"error": str(e)}, status=500)


""" create a new view for the InitialFamilyData model that will create a new row"""


@conditional_csrf
@api_view(["POST"])
def create_initial_family_data(request):
    api_logger.info("create_initial_family_data called")
    """
    Create a new initial family data record in the InitialFamilyData model.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='CREATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has CREATE permission on the "initial_family_data" resource
    if not has_initial_family_data_permission(request, "create"):
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='CREATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to create initial family data",
            status_code=401
        )
        return JsonResponse(
            {"error": "You do not have permission to create initial family data."},
            status=401,
        )

    try:
        data = request.data

        # Validate required fields
        required_fields = ["names", "phones"]
        missing_fields = [
            field for field in required_fields if not data.get(field, "").strip()
        ]
        if missing_fields:
            # ❌ MISSING AUDIT LOG
            log_api_action(
                request=request,
                action='CREATE_INITIAL_FAMILY_FAILED',
                success=False,
                error_message=f"Missing or empty required fields: {', '.join(missing_fields)}",
                status_code=400
            )
            return JsonResponse(
                {
                    "error": f"Missing or empty required fields: {', '.join(missing_fields)}"
                },
                status=400,
            )
        
        # Create a new initial family data record in the database
        initial_family_data = InitialFamilyData.objects.create(
            names=data["names"],
            phones=data["phones"],
            other_information=(
                data.get("other_information") if data.get("other_information") else None
            ),
            created_at=make_aware(datetime.datetime.now()),
            updated_at=make_aware(datetime.datetime.now()),
            family_added=False,  # Default to False
        )

        api_logger.debug(
            f"Initial family data created successfully with ID {initial_family_data.initial_family_data_id}"
        )
        
        # ❌ MISSING SUCCESS AUDIT LOG
        log_api_action(
            request=request,
            action='CREATE_INITIAL_FAMILY_SUCCESS',
            affected_tables=['childsmile_app_initialfamilydata'],
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data.initial_family_data_id],
            success=True,
            additional_data={
                'family_names': data["names"],
                'family_phones': data["phones"]
            }
        )
        
        return JsonResponse(
            {
                "message": "Initial family data created successfully",
                "initial_family_data_id": initial_family_data.initial_family_data_id,
            },
            status=201,
        )
    except Exception as e:
        api_logger.error(f"An error occurred while creating initial family data: {str(e)}")
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='CREATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500
        )
        return JsonResponse({"error": str(e)}, status=500)


""" create a new view for the InitialFamilyData model that will update an existing row by id"""


@conditional_csrf
@api_view(["PUT"])
def update_initial_family_data(request, initial_family_data_id):
    api_logger.info(f"update_initial_family_data called for initial_family_data_id: {initial_family_data_id}")
    """
    Update an existing initial family data record in the InitialFamilyData model.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Check if the user has UPDATE permission on the "initial_family_data" resource
    if not has_initial_family_data_permission(request, "update"):
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to update initial family data",
            status_code=401,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {"error": "You do not have permission to update initial family data."},
            status=401,
        )

    try:
        initial_family_data = InitialFamilyData.objects.get(
            initial_family_data_id=initial_family_data_id
        )
    except InitialFamilyData.DoesNotExist:
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Initial family data not found",
            status_code=404,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse({"error": "Initial family data not found."}, status=404)

    data = request.data
    required_fields = ["names", "phones"]
    missing_fields = [
        field for field in required_fields if not data.get(field, "").strip()
    ]
    if missing_fields:
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message=f"Missing or empty required fields: {', '.join(missing_fields)}",
            status_code=400,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {"error": f"Missing or empty required fields: {', '.join(missing_fields)}"},
            status=400,
        )

    # Update the fields if they are provided in the request
    initial_family_data.names = data.get("names", initial_family_data.names)
    initial_family_data.phones = data.get("phones", initial_family_data.phones)
    initial_family_data.other_information = data.get(
        "other_information", initial_family_data.other_information
    )
    initial_family_data.updated_at = make_aware(datetime.datetime.now())
    initial_family_data.family_added = data.get(
        "family_added", initial_family_data.family_added
    )

    # Save the updated record
    try:
        initial_family_data.save()
        
        # ❌ MISSING SUCCESS AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_SUCCESS',
            affected_tables=['childsmile_app_initialfamilydata'],
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            success=True,
            additional_data={
                'updated_family_names': initial_family_data.names,
                'family_added_status': initial_family_data.family_added
            }
        )
        
        return JsonResponse(
            {"message": "Initial family data updated successfully"}, status=200
        )
    except Exception as e:
        api_logger.error(f"An error occurred while updating initial family data: {str(e)}")
        # ❌ MISSING AUDIT LOG
        log_api_action(
            request=request,
            action='UPDATE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["PUT"])
def mark_initial_family_complete(request, initial_family_data_id):
    api_logger.info(f"mark_initial_family_complete called for initial_family_data_id: {initial_family_data_id}")
    """
    Update an existing initial family data record in the InitialFamilyData model.
    """
    user_id = request.session.get("user_id")
   
    if not user_id:
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Fetch the user and their roles
    user = Staff.objects.get(staff_id=user_id)
    user_roles = set(user.roles.values_list("role_name", flat=True))
    allowed_roles = {"System Administrator", "Technical Coordinator"}
    if not user_roles.intersection(allowed_roles):
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message="You do not have permission to mark initial family data as complete",
            status_code=401,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {
                "error": "You do not have permission to mark initial family data as complete."
            },
            status=401,
        )

    # Check if the user has UPDATE permission on the "initial_family_data" resource
    if not has_initial_family_data_permission(request, "update"):
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message="You do not have permission to update initial family data",
            status_code=401,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {"error": "You do not have permission to update initial family data."},
            status=401,
        )

    try:
        initial_family_data = InitialFamilyData.objects.get(
            initial_family_data_id=initial_family_data_id
        )
    except InitialFamilyData.DoesNotExist:
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message="Initial family data not found",
            status_code=404,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            additional_data={
                'family_names': 'Unknown - Not Found',  # **ADD THIS**
                'family_phones': 'Unknown - Not Found'  # **ADD THIS**
            }
        )
        return JsonResponse({"error": "Initial family data not found."}, status=404)

    data = request.data

    # Update the fields if they are provided in the request
    initial_family_data.updated_at = make_aware(datetime.datetime.now())
    initial_family_data.family_added = data.get(
        "family_added", initial_family_data.family_added
    )

    if initial_family_data.family_added:
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message="Initial family data is already marked as complete",
            status_code=400,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            additional_data={
                'family_names': initial_family_data.names,  # **ADD THIS**
                'family_phones': initial_family_data.phones  # **ADD THIS**
            }
        )
        return JsonResponse(
            {"error": "Initial family data is already marked as complete."},
            status=400,
        )
    
    # Save the updated record
    try:
        initial_family_data.save()

        related_tasks = Tasks.objects.filter(
            initial_family_data_id_fk=initial_family_data_id
        )
        if related_tasks.exists():
            for task in related_tasks:
                task.status = "הושלמה"
                task.save()
                api_logger.debug(f"Task {task.task_id} marked as completed.")

        # Log successful mark as complete
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_SUCCESS',
            affected_tables=['childsmile_app_initialfamilydata', 'childsmile_app_tasks'],
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            success=True,
            additional_data={
                'family_names': initial_family_data.names,  # **ADD THIS**
                'family_phones': initial_family_data.phones,  # **ADD THIS**
                'family_added_status': initial_family_data.family_added
            }
        )

        return JsonResponse(
            {"message": "Initial family data successfully marked as complete"},
            status=200,
        )
    except Exception as e:
        api_logger.error(
            f"An error occurred while marking initial family data complete: {str(e)}"
        )
        log_api_action(
            request=request,
            action='MARK_FAMILY_ADDED_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            additional_data={
                'family_names': initial_family_data.names if 'initial_family_data' in locals() else 'Unknown - Error',  # **ADD THIS**
                'family_phones': initial_family_data.phones if 'initial_family_data' in locals() else 'Unknown - Error'  # **ADD THIS**
            }
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["DELETE"])
def delete_initial_family_data(request, initial_family_data_id):
    api_logger.info(f"delete_initial_family_data called for initial_family_data_id: {initial_family_data_id}")
    """
    Delete an existing initial family data record in the InitialFamilyData model.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )

    # Fetch the user and their roles
    user = Staff.objects.get(staff_id=user_id)
    user_roles = set(user.roles.values_list("role_name", flat=True))
    allowed_roles = {"System Administrator", "Technical Coordinator"}
    if not user_roles.intersection(allowed_roles):
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to delete initial family data",
            status_code=401,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {"error": "You do not have permission to delete initial family data."},
            status=401,
        )

    # Check if the user has DELETE permission on the "initial_family_data" resource
    if not has_initial_family_data_permission(request, "delete"):
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to delete initial family data",
            status_code=401,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id]
        )
        return JsonResponse(
            {"error": "You do not have permission to delete initial family data."},
            status=401,
        )

    try:
        initial_family_data = InitialFamilyData.objects.get(
            initial_family_data_id=initial_family_data_id
        )
        
        # Store names and phones for audit BEFORE deletion
        family_names = initial_family_data.names
        family_phones = initial_family_data.phones

        related_tasks = Tasks.objects.filter(
            initial_family_data_id_fk=initial_family_data_id
        )
        if related_tasks.exists():
            for task in related_tasks:
                task.delete()
                api_logger.debug(
                    f"Task {task.task_id} deleted due to initial family data deletion."
                )

        # Delete the initial family data record
        initial_family_data.delete()

        # Log successful deletion
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_SUCCESS',
            affected_tables=['childsmile_app_initialfamilydata', 'childsmile_app_tasks'],
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            success=True,
            additional_data={
                'deleted_family_names': family_names,  # **ADD THIS**
                'deleted_family_phones': family_phones  # **ADD THIS**
            }
        )

        return JsonResponse(
            {"message": "Initial family data deleted successfully"}, status=200
        )
    except InitialFamilyData.DoesNotExist:
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message="Initial family data not found",
            status_code=404,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            additional_data={
                'deleted_family_names': 'Unknown - Not Found',  # **ADD THIS**
                'deleted_family_phones': 'Unknown - Not Found'  # **ADD THIS**
            }
        )
        return JsonResponse({"error": "Initial family data not found."}, status=404)
    except Exception as e:
        api_logger.error(f"An error occurred while deleting initial family data: {str(e)}")
        log_api_action(
            request=request,
            action='DELETE_INITIAL_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='InitialFamilyData',
            entity_ids=[initial_family_data_id],
            additional_data={
                'deleted_family_names': 'Unknown - Error',  # **ADD THIS**
                'deleted_family_phones': 'Unknown - Error'  # **ADD THIS**
            }
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["PUT"])
def update_child_id(request, old_id):
    """
    Update a child's ID (Israeli ID) across all related tables.
    This performs a cascading update of the ID in Children table and all FK references.
    """
    api_logger.info(f"update_child_id called for old_id: {old_id}")
    
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403,
            entity_type='Children',
            entity_ids=[old_id]
        )
        return JsonResponse({"detail": "Authentication credentials were not provided."}, status=403)
    
    # Check permission for children table updates
    if not has_permission(request, "children", "UPDATE"):
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message="You do not have permission to update child IDs",
            status_code=401,
            entity_type='Children',
            entity_ids=[old_id]
        )
        return JsonResponse({"error": "You do not have permission to update child IDs."}, status=401)

    data = request.data
    new_id = data.get("new_id")
    
    if not new_id:
        return JsonResponse({"error": "New ID is required."}, status=400)
    
    # Validate new ID format (9 digits for Israeli ID)
    new_id_str = str(new_id).strip()
    if not new_id_str.isdigit() or len(new_id_str) != 9:
        return JsonResponse({"error": "Invalid ID format. Israeli ID must be exactly 9 digits."}, status=400)
    
    new_id = int(new_id_str)
    
    # Check if new ID already exists
    if Children.objects.filter(child_id=new_id).exists():
        return JsonResponse({"error": "This child ID already exists in the system."}, status=400)
    
    try:
        child = Children.objects.get(child_id=old_id)
    except Children.DoesNotExist:
        return JsonResponse({"error": "Child/Family not found."}, status=404)
    
    try:
        with transaction.atomic():
            # Store original data for audit log
            original_data = {
                'child_name': f"{child.childfirstname} {child.childsurname}",
                'city': child.city,
                'status': child.status
            }
            
            affected_tables = ['childsmile_app_children']
            
            # IMPORTANT: Update tables in correct order - child tables before parent tables
            # to avoid FK constraint violations
            
            # 1. Update PrevTutorshipStatuses FK (references Children via child_id)
            prev_status_updated = PrevTutorshipStatuses.objects.filter(child_id=old_id).update(child_id=new_id)
            if prev_status_updated:
                affected_tables.append('childsmile_app_prevtutorshipstatuses')
                api_logger.debug(f"Updated {prev_status_updated} PrevTutorshipStatuses records")
            
            # 2. Update Tutorships FK (references Children via child)
            tutorship_updated = Tutorships.objects.filter(child_id=old_id).update(child_id=new_id)
            if tutorship_updated:
                affected_tables.append('childsmile_app_tutorships')
                api_logger.debug(f"Updated {tutorship_updated} Tutorships records")
            
            # 3. Update Tasks related_child FK (references Children)
            tasks_updated = Tasks.objects.filter(related_child_id=old_id).update(related_child_id=new_id)
            if tasks_updated:
                affected_tables.append('childsmile_app_tasks')
                api_logger.debug(f"Updated {tasks_updated} Tasks records")
            
            # 4. Finally update Children record - change the primary key
            # Delete the old record and create a new one with new ID
            # Store all fields from old record
            child_data = {
                'child_id': new_id,
                'childfirstname': child.childfirstname,
                'childsurname': child.childsurname,
                'registrationdate': child.registrationdate,
                'lastupdateddate': datetime.datetime.now().date(),
                'gender': child.gender,
                'responsible_coordinator': child.responsible_coordinator,
                'city': child.city,
                'child_phone_number': child.child_phone_number,
                'treating_hospital': child.treating_hospital,
                'date_of_birth': child.date_of_birth,
                'medical_diagnosis': child.medical_diagnosis,
                'diagnosis_date': child.diagnosis_date,
                'marital_status': child.marital_status,
                'num_of_siblings': child.num_of_siblings,
                'details_for_tutoring': child.details_for_tutoring,
                'additional_info': child.additional_info,
                'tutoring_status': child.tutoring_status,
                'current_medical_state': child.current_medical_state,
                'when_completed_treatments': child.when_completed_treatments,
                'father_name': child.father_name,
                'father_phone': child.father_phone,
                'mother_name': child.mother_name,
                'mother_phone': child.mother_phone,
                'street_and_apartment_number': child.street_and_apartment_number,
                'expected_end_treatment_by_protocol': child.expected_end_treatment_by_protocol,
                'has_completed_treatments': child.has_completed_treatments,
                'status': child.status,
                'last_review_talk_conducted': child.last_review_talk_conducted,
            }
            
            # Delete the old record
            child.delete()
            
            # Create new record with new ID
            new_child = Children.objects.create(**child_data)
            api_logger.debug(f"Created new Children record with new_id: {new_id}")
            
            # Log successful ID update
            log_api_action(
                request=request,
                action='UPDATE_FAMILY_SUCCESS',
                affected_tables=affected_tables,
                entity_type='Children',
                entity_ids=[old_id, new_id],
                success=True,
                additional_data={
                    'update_type': 'CHILD_ID_CHANGE',
                    'old_id': old_id,
                    'new_id': new_id,
                    'child_name': original_data['child_name'],
                    'child_city': original_data['city'],
                    'child_status': original_data['status'],
                    'tutorships_updated': tutorship_updated,
                    'tasks_updated': tasks_updated,
                    'prev_statuses_updated': prev_status_updated
                }
            )
            
            return JsonResponse({
                "message": "Child ID updated successfully across all related tables.",
                "old_id": old_id,
                "new_id": new_id,
                "affected_tables": affected_tables,
                "summary": {
                    "tutorships_updated": tutorship_updated,
                    "tasks_updated": tasks_updated,
                    "prev_statuses_updated": prev_status_updated
                }
            }, status=200)
            
    except Exception as e:
        api_logger.error(f"Error updating child ID: {str(e)}\n{traceback.format_exc()}")
        log_api_action(
            request=request,
            action='UPDATE_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='Children',
            entity_ids=[old_id],
            additional_data={'attempted_new_id': new_id}
        )
        return JsonResponse({"error": f"Error updating child ID: {str(e)}"}, status=500)

# ==================== BULK IMPORT FAMILIES ====================


@conditional_csrf
@api_view(['POST'])
def import_families_endpoint(request):
    
    try:
        # Check feature flag - inline (no helper method)
        if not (os.environ.get("FAMILIES_IMPORT_ENABLED", "False").lower() == "true"):
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="Bulk import feature is not enabled",
                status_code=403,
                additional_data={'reason': 'Feature disabled'}
            )
            return JsonResponse(
                {'error': 'Bulk import feature is not enabled'},
                status=403
            )
        
        # Check authentication (403 - not authenticated)
        user_id = request.session.get("user_id")
        if not user_id:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="Authentication credentials were not provided",
                status_code=403,
                additional_data={'reason': 'Not authenticated'}
            )
            return JsonResponse(
                {'error': 'Authentication credentials were not provided'},
                status=403
            )
        
        try:
            user = Staff.objects.get(staff_id=user_id)
        except Staff.DoesNotExist:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="User not found",
                status_code=403,
                additional_data={'reason': 'User does not exist'}
            )
            return JsonResponse(
                {'error': 'User not found'},
                status=403
            )
        
        # Check permission (401 - authenticated but no permission)
        has_family_create = has_permission(request, "children", "CREATE")
        
        if not has_family_create:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_FAILED',
                success=False,
                error_message="You do not have permission to import families",
                status_code=401,
                additional_data={'reason': 'Permission denied', 'user_email': user.email}
            )
            api_logger.critical(f"Unauthorized family import attempt by {user.email} - lacks CREATE permission")
            return JsonResponse(
                {'error': 'You do not have permission to import families'},
                status=401
            )
        
        api_logger.info(f"Import permission granted for user: {user.email}")
        
        # Refresh settlements data from database
        global SETTLEMENTS_N_STREETS
        SETTLEMENTS_N_STREETS = get_settlements_dict()
        api_logger.info(f"Loaded {len(SETTLEMENTS_N_STREETS)} settlements from database")
        
        # Get uploaded file
        if 'file' not in request.FILES:
            return JsonResponse(
                {'error': 'No file provided'},
                status=400
            )
        
        file = request.FILES['file']
        
        # Validate file extension
        if not file.name.endswith('.xlsx'):
            return JsonResponse(
                {'error': 'File must be .xlsx format'},
                status=400
            )
        
        dry_run = request.data.get('dry_run', 'false').lower() == 'true'
        
        # Read Excel file
        try:
            df = pd.read_excel(file, dtype=str)
        except Exception as e:
            api_logger.error(f"Failed to read Excel file: {str(e)}")
            return JsonResponse(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=400
            )
        
        if df.empty:
            return JsonResponse(
                {'error': 'Excel file is empty'},
                status=400
            )
        
        # Log actual column names for debugging
        api_logger.info(f"Excel columns found: {list(df.columns)}")
        
        # Create flexible column mapping - try to find columns by name
        # This handles both "שם מלא" and separate "שם פרטי"/"שם משפחה" columns
        columns = {col.strip(): col for col in df.columns}  # Normalized column names
        
        # Find column names (case-insensitive and trim whitespace)
        def find_column(potential_names):
            """Find which column exists from a list of potential names"""
            for name in potential_names:
                normalized = name.strip()
                for col, orig_col in columns.items():
                    if col == normalized:
                        return orig_col
            return None
        def find_column_contains(substring):
            for col in df.columns:
                if substring in col:
                    return col
            return None

        # Map all possible column names
        col_first_name = find_column(['שם פרטי']) or 'שם פרטי'
        col_last_name = find_column(['שם משפחה']) or 'שם משפחה'
        col_full_name = find_column(['שם מלא של הילד/ה', 'שם מלא']) or None
        col_child_id = find_column(['תעודת זהות ילד/ה', 'תעודת זהות', 'ID']) or 'תעודת זהות ילד/ה'
        col_status = find_column(['סטטוס']) or 'סטטוס'
        col_marital_status = find_column(['סטטוס זוגי']) or 'סטטוס זוגי'
        col_tutoring_status = find_column(['מצב חונכות', 'סטטוס חונכות']) or 'מצב חונכות'
        col_gender = find_column(['מין']) or 'מין'
        col_city = find_column(['עיר']) or 'עיר'
        col_phone = find_column(['מספר טלפון של הילד/ה', 'מספר טלפון']) or 'מספר טלפון של הילד/ה'
        col_hospital = find_column(['בית חולים מטפל', 'בית חולים']) or 'בית חולים מטפל'
        col_diagnosis = find_column(['אבחנה רפואית', 'אבחנה']) or 'אבחנה רפואית'
        col_diagnosis_date = find_column(['תאריך אבחון', 'תאריך אבחנה']) or 'תאריך אבחון'
        col_num_of_siblings = find_column_contains("כמה אחים") or 'כמה אחים יש?'
        col_registration_date = find_column(['תאריך רישום']) or 'תאריך רישום'
        # NEW: Find TWO separate columns for street and apartment
        col_street = find_column(['רחוב']) or 'רחוב'
        col_apartment = find_column(['מס׳ דירה', 'מס דירה']) or 'מס׳ דירה'
        col_medical_state = find_column(['מצב רפואי עדכני', 'מצב רפואי']) or 'מצב רפואי עדכני'
        col_completed_treatments = find_column(['האם סיים/ה טיפולים?', 'סיום טיפולים']) or 'האם סיים/ה טיפולים?'
        col_treatment_end_protocol = find_column(['צפי סיום על פי פרוטוקול?', 'תאריך סיום צפוי']) or 'צפי סיום על פי פרוטוקול?'
        col_when_completed_treatments = find_column(['מתי סיים/ה טיפולים?', 'תאריך סיום טיפולים']) or 'מתי סיים/ה טיפולים?'
        col_father_name = find_column(['שם האב']) or 'שם האב'
        col_father_phone = find_column_contains("טלפון של האב") or 'מס טלפון של האב'
        col_mother_name = find_column(['שם האם']) or 'שם האם'
        col_mother_phone = find_column_contains("טלפון של האם") or 'מס טלפון של האם'
        col_tutoring_details = find_column(['פרטים לצורך חונכות', 'פרטים לחונכות']) or 'פרטים לצורך חונכות (לצורך ליה ונעם)'
        col_additional_info = find_column(['האם יש משהו ספציפי שתרצו לבקש/לדעת?', 'מידע נוסף', 'הערות']) or 'האם יש משהו ספציפי שתרצו לבקש/לדעת?'
        col_is_in_frame = find_column(['האם נמצא במסגרת?']) or 'האם נמצא במסגרת?'
        col_coordinator_comments = find_column(['הערות רכז', 'הערות רכז/מנהל']) or 'הערות רכז'
        
        api_logger.info(f"Column mapping: first_name={col_first_name}, last_name={col_last_name}, full_name={col_full_name}")
        
        # Get valid enum values for validation during dry-run
        valid_marital_statuses = set(get_enum_values("marital_status"))
        valid_tutoring_statuses = set(get_enum_values("tutoring_status"))
        valid_statuses = set(get_enum_values("status"))
        
        # Process records
        total_records = len(df)
        success_count = 0
        error_count = 0
        skipped_count = 0
        invalid_enum_count = 0
        
        results = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            
            # Extract first and last names from separate columns (preferred) or full name
            first_name_raw = row.get(col_first_name, '')
            last_name_raw = row.get(col_last_name, '')
            
            child_first_name = '' if (first_name_raw is None or pd.isna(first_name_raw) or str(first_name_raw).lower() == 'nan') else str(first_name_raw).strip()
            child_last_name = '' if (last_name_raw is None or pd.isna(last_name_raw) or str(last_name_raw).lower() == 'nan') else str(last_name_raw).strip()
            
            # Fallback to full_name column if separate names are missing
            if (not child_first_name or not child_last_name) and col_full_name:
                full_name_raw = row.get(col_full_name, '')
                full_name = '' if (full_name_raw is None or pd.isna(full_name_raw) or str(full_name_raw).lower() == 'nan') else str(full_name_raw).strip()
                
                if full_name:
                    name_parts = full_name.split()
                    if not child_last_name:
                        child_last_name = name_parts[-1]  # Last part is surname
                    if not child_first_name:
                        child_first_name = ' '.join(name_parts[:-1]) if len(name_parts) > 1 else name_parts[0]
            
            # Parse status early for surname logic
            status_raw = row.get(col_status, '')
            status_val = '' if (status_raw is None or pd.isna(status_raw) or str(status_raw).lower() == 'nan') else str(status_raw).strip()
            
            # Handle missing surname: if no surname and status is בריא/ז״ל, use status as surname
            # Otherwise use "XXX" as placeholder and create update task
            if not child_last_name:
                if status_val in ['בריא', 'ז״ל']:
                    # Use status as surname for healthy/deceased children
                    child_last_name = status_val
                    needs_surname_task = False
                else:
                    # Use XXX as placeholder
                    child_last_name = 'XXX'
                    needs_surname_task = True
            else:
                needs_surname_task = False
            
            result = {
                'row_num': row_num,
                'child_first_name': child_first_name,
                'child_last_name': child_last_name,
                'child_id': '',
                'status': '',
                'details': '',
                'enum_warnings': []
            }
            
            try:
                # Parse child ID and pad with leading zeros if needed
                child_id_raw = row.get(col_child_id, '')
                child_id = '' if (child_id_raw is None or pd.isna(child_id_raw) or str(child_id_raw).lower() == 'nan') else str(child_id_raw).strip()
                
                # Pad with leading zeros to make it 9 digits
                if child_id:
                    try:
                        child_id_int_temp = int(child_id)
                        child_id = str(child_id_int_temp).zfill(9)  # Pad to 9 digits with leading zeros
                    except ValueError:
                        child_id = ''
                
                result['child_id'] = child_id
                
                # Validate required fields
                if not child_first_name or not child_last_name:
                    result['status'] = 'Error'
                    result['details'] = 'חסר שם פרטי או שם משפחה של ילד'
                    error_count += 1
                    results.append(result)
                    continue
                
                if not child_id:
                    result['status'] = 'Error'
                    result['details'] = 'חסרה תעודת זהות של ילד'
                    error_count += 1
                    results.append(result)
                    continue
                
                try:
                    child_id_int = int(child_id)
                except ValueError:
                    result['status'] = 'Error'
                    result['details'] = f'תעודת זהות לא מספרית: {child_id}'
                    error_count += 1
                    results.append(result)
                    continue
                
                # Check if child already exists
                if Children.objects.filter(child_id=child_id_int).exists():
                    result['status'] = 'Skipped'
                    result['details'] = f'תעודת זהות כבר קיימת במערכת'
                    skipped_count += 1
                    results.append(result)
                    continue
                
                # Parse birth_date FIRST (needed for both dry-run age check and live import)
                date_val = row.get('תאריך לידה', '')
                birth_date = None
                if date_val and not pd.isna(date_val) and str(date_val).lower() != 'nan':
                    try:
                        from datetime import datetime as dt
                        # If already a datetime object (from pandas), use it directly
                        if isinstance(date_val, dt):
                            birth_date = date_val.date()
                        else:
                            date_str = str(date_val).strip()
                            # Try multiple date formats: ISO, EU (dd/mm/yyyy), US (mm/dd/yyyy), dotted, dashed
                            date_formats = [
                                '%Y-%m-%d %H:%M:%S',  # ISO with time
                                '%Y-%m-%d',           # ISO
                                '%d/%m/%Y',           # EU: dd/mm/yyyy
                                '%m/%d/%Y',           # US: mm/dd/yyyy
                                '%d-%m-%Y',           # dd-mm-yyyy
                                '%m-%d-%Y',           # mm-dd-yyyy
                                '%d.%m.%Y',           # dd.mm.yyyy
                                '%m.%d.%Y',           # mm.dd.yyyy
                                '%d/%m/%y',           # dd/mm/yy (2-digit year)
                                '%m/%d/%y',           # mm/dd/yy
                                '%d-%m-%y',           # dd-mm-yy
                                '%m-%d-%y',           # mm-dd-yy
                                '%d.%m.%y',           # dd.mm.yy
                                '%m.%d.%y',           # mm.dd.yy
                                '%Y/%m/%d',           # yyyy/mm/dd
                                '%Y.%m.%d',           # yyyy.mm.dd
                            ]
                            for fmt in date_formats:
                                try:
                                    birth_date = dt.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    except Exception:
                        birth_date = None
                
                # === CREATE RECORD ===
                # Parse optional fields
                city_raw = row.get(col_city, '')
                city_original = '' if (city_raw is None or pd.isna(city_raw) or str(city_raw).lower() == 'nan') else str(city_raw).strip()
                
                # Remove commas and extra spaces from city
                city_original = city_original.replace(',', '').strip() if city_original else ''
                
                # Try EXACT matching FIRST (for Excel data manually copied from system)
                # Only use smart matching functions if exact match fails
                if not city_original:
                    result['status'] = 'Error'
                    result['details'] = 'שגיאה: עיר חסרה (עמודה K) - חובה מלא'
                    result['enum_warnings'].append('עיר חסרה')
                    error_count += 1
                    results.append(result)
                    continue
                
                # **STEP 1: Try EXACT match first** (for manually entered system values)
                city_found = None
                for json_city_key in SETTLEMENTS_N_STREETS.keys():
                    if json_city_key.strip() == city_original:
                        city_found = json_city_key
                        break
                
                if city_found:
                    city = city_found  # Use exact match
                else:
                    # **STEP 2: Fallback to smart matching** (find_best_city_match with special fallback values)
                    city, city_was_corrected = find_best_city_match(city_original)
                    
                    if city is None:
                        result['status'] = 'Error'
                        result['details'] = f'שגיאה: עיר "{city_original}" לא קיימת במערכת (עמודה K)'
                        result['enum_warnings'].append(f'עיר "{city_original}" לא קיימת במערכת')
                        error_count += 1
                        results.append(result)
                        continue
                    
                    # If city was corrected, note it
                    if city_was_corrected and city != city_original:
                        result['enum_warnings'].append(f'עיר תוקנה: "{city_original}" → "{city}"')
                
                # Parse street and apartment from Excel - now TWO separate columns
                street_raw = row.get(col_street, '')
                street_original = '' if (street_raw is None or pd.isna(street_raw) or str(street_raw).lower() == 'nan') else str(street_raw).strip()
                street_original = street_original.replace(',', '').strip() if street_original else ''
                
                apartment_raw = row.get(col_apartment, '')
                apartment_original = '' if (apartment_raw is None or pd.isna(apartment_raw) or str(apartment_raw).lower() == 'nan') else str(apartment_raw).strip()
                apartment_original = apartment_original.replace(',', '').strip() if apartment_original else ''
                
                # Validate required address fields: street is REQUIRED
                if not street_original:
                    result['status'] = 'Error'
                    result['details'] = 'שגיאה: רחוב חסר (עמודה J) - חובה מלא'
                    result['enum_warnings'].append('רחוב חסר')
                    error_count += 1
                    results.append(result)
                    continue
                
                # Get the city's streets from JSON
                city_streets = SETTLEMENTS_N_STREETS.get(city, [])
                
                # **STEP 1: Try EXACT match first** (for manually entered system values)
                matched_street = None
                if street_original in city_streets:
                    matched_street = street_original  # Use exact match
                else:
                    # **STEP 2: Fallback to smart matching** (find_best_street_match with special fallback values)
                    matched_street, street_was_corrected, is_placeholder = find_best_street_match(street_original, city)
                    
                    # If street is not found and no correction was made, it's an error
                    if matched_street == street_original and not street_was_corrected:
                        # Check if it exists in the city's streets as fallback
                        if street_original not in city_streets:
                            result['status'] = 'Error'
                            result['details'] = f'שגיאה: רחוב "{street_original}" לא קיים ברשימת הרחובות של "{city}" (עמודה J)'
                            result['enum_warnings'].append(f'רחוב "{street_original}" לא קיים ברשימת הרחובות של "{city}"')
                            error_count += 1
                            results.append(result)
                            continue
                    
                    # If street was corrected, note it
                    if street_was_corrected and matched_street != street_original:
                        result['enum_warnings'].append(f'רחוב תוקן: "{street_original}" → "{matched_street}"')
                
                # Concatenate street and apartment for storage
                street_and_apartment = f"{matched_street} {apartment_original}".strip() if apartment_original else matched_street
                
                # Parse ALL fields (for both dry-run and real import)
                phone_raw = row.get(col_phone, '')
                phone = '' if (phone_raw is None or pd.isna(phone_raw) or str(phone_raw).lower() == 'nan') else str(phone_raw).strip()
                # Remove commas from phone
                phone = phone.replace(',', '').strip() if phone else ''
                # Format phone: remove dashes/spaces, pad to 10 digits with trailing zeros if needed
                if phone:
                    phone_normalized = phone.replace('-', '').replace(' ', '').strip()
                    # Remove any non-digit characters
                    phone_normalized = ''.join(c for c in phone_normalized if c.isdigit())
                    # Pad to 10 digits with trailing zeros if fewer than 10
                    if phone_normalized:
                        phone_normalized = phone_normalized.ljust(10, '0')  # Pad with trailing zeros
                    if phone_normalized and len(phone_normalized) == 10:
                        phone = f"{phone_normalized[:3]}-{phone_normalized[3:]}"  # Format as XXX-XXXXXXX
                    else:
                        phone = ''  # Invalid phone
                else:
                    phone = ''
                
                hospital_raw = row.get(col_hospital, '')
                hospital = '' if (hospital_raw is None or pd.isna(hospital_raw) or str(hospital_raw).lower() == 'nan') else str(hospital_raw).strip()
                # Remove commas and trim spaces
                hospital = hospital.replace(',', '').strip() if hospital else ''
                
                diagnosis_raw = row.get(col_diagnosis, '')
                diagnosis = '' if (diagnosis_raw is None or pd.isna(diagnosis_raw) or str(diagnosis_raw).lower() == 'nan') else str(diagnosis_raw).strip()
                # Remove commas and trim spaces
                diagnosis = diagnosis.replace(',', '').strip() if diagnosis else ''
                
                # Parse medical state from Excel
                medical_state_raw = row.get(col_medical_state, '')
                medical_state = '' if (medical_state_raw is None or pd.isna(medical_state_raw) or str(medical_state_raw).lower() == 'nan') else str(medical_state_raw).strip()
                
                # Parse father details from Excel
                father_name_raw = row.get(col_father_name, '')
                father_name = '' if (father_name_raw is None or pd.isna(father_name_raw) or str(father_name_raw).lower() == 'nan') else str(father_name_raw).strip()
                
                father_phone_raw = row.get(col_father_phone, '')
                father_phone = '' if (father_phone_raw is None or pd.isna(father_phone_raw) or str(father_phone_raw).lower() == 'nan') else str(father_phone_raw).strip()
                # Format father phone similar to child phone
                if father_phone:
                    father_phone_normalized = father_phone.replace('-', '').replace(' ', '').strip()
                    father_phone_normalized = ''.join(c for c in father_phone_normalized if c.isdigit())
                    if father_phone_normalized:
                        father_phone_normalized = father_phone_normalized.ljust(10, '0')
                    if father_phone_normalized and len(father_phone_normalized) == 10:
                        father_phone = f"{father_phone_normalized[:3]}-{father_phone_normalized[3:]}"
                    else:
                        father_phone = ''
                
                # Parse mother details from Excel
                mother_name_raw = row.get(col_mother_name, '')
                mother_name = '' if (mother_name_raw is None or pd.isna(mother_name_raw) or str(mother_name_raw).lower() == 'nan') else str(mother_name_raw).strip()
                
                mother_phone_raw = row.get(col_mother_phone, '')
                mother_phone = '' if (mother_phone_raw is None or pd.isna(mother_phone_raw) or str(mother_phone_raw).lower() == 'nan') else str(mother_phone_raw).strip()
                # Format mother phone similar to child phone
                if mother_phone:
                    mother_phone_normalized = mother_phone.replace('-', '').replace(' ', '').strip()
                    mother_phone_normalized = ''.join(c for c in mother_phone_normalized if c.isdigit())
                    if mother_phone_normalized:
                        mother_phone_normalized = mother_phone_normalized.ljust(10, '0')
                    if mother_phone_normalized and len(mother_phone_normalized) == 10:
                        mother_phone = f"{mother_phone_normalized[:3]}-{mother_phone_normalized[3:]}"
                    else:
                        mother_phone = ''
                
                # Parse tutoring details from Excel
                tutoring_details_raw = row.get(col_tutoring_details, '')
                tutoring_details = '' if (tutoring_details_raw is None or pd.isna(tutoring_details_raw) or str(tutoring_details_raw).lower() == 'nan') else str(tutoring_details_raw).strip()
                
                # Parse additional info from Excel
                additional_info_raw = row.get(col_additional_info, '')
                additional_info = '' if (additional_info_raw is None or pd.isna(additional_info_raw) or str(additional_info_raw).lower() == 'nan') else str(additional_info_raw).strip()

                # Parse is_in_frame from Excel (frame status)
                is_in_frame_raw = row.get(col_is_in_frame, '')
                is_in_frame = '' if (is_in_frame_raw is None or pd.isna(is_in_frame_raw) or str(is_in_frame_raw).lower() == 'nan') else str(is_in_frame_raw).strip()

                # Parse coordinator_comments from Excel
                coordinator_comments_raw = row.get(col_coordinator_comments, '')
                coordinator_comments = '' if (coordinator_comments_raw is None or pd.isna(coordinator_comments_raw) or str(coordinator_comments_raw).lower() == 'nan') else str(coordinator_comments_raw).strip()

                
                # Parse diagnosis_date from Excel - ONLY if valid format found, else stay NULL
                diagnosis_date = None
                diagnosis_date_raw = row.get(col_diagnosis_date, '')
                if diagnosis_date_raw and not pd.isna(diagnosis_date_raw) and str(diagnosis_date_raw).lower() != 'nan':
                    try:
                        from datetime import datetime as dt
                        # If already a datetime object (from pandas), use it directly
                        if isinstance(diagnosis_date_raw, dt):
                            diagnosis_date = diagnosis_date_raw.date()
                        else:
                            date_str = str(diagnosis_date_raw).strip()
                            # Try multiple date formats
                            date_formats = [
                                '%Y-%m-%d %H:%M:%S',  # ISO with time
                                '%Y-%m-%d',           # ISO
                                '%d/%m/%Y',           # EU: dd/mm/yyyy
                                '%m/%d/%Y',           # US: mm/dd/yyyy
                                '%d-%m-%Y',           # dd-mm-yyyy
                                '%m-%d-%Y',           # mm-dd-yyyy
                                '%d.%m.%Y',           # dd.mm.yyyy
                                '%m.%d.%Y',           # mm.dd.yyyy
                                '%d/%m/%y',           # dd/mm/yy (2-digit year)
                                '%m/%d/%y',           # mm/dd/yy
                                '%d-%m-%y',           # dd-mm-yy
                                '%m-%d-%y',           # mm-dd-yy
                                '%d.%m.%y',           # dd.mm.yy
                                '%m.%d.%y',           # mm.dd.yy
                                '%Y/%m/%d',           # yyyy/mm/dd
                                '%Y.%m.%d',           # yyyy.mm.dd
                            ]
                            for fmt in date_formats:
                                try:
                                    diagnosis_date = dt.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    except Exception:
                        # If any error occurs during parsing, leave as None
                        diagnosis_date = None
                
                # Parse registration_date from Excel - use as is, else default to today
                registration_date = datetime.datetime.now().date()  # Default to today
                registration_date_raw = row.get(col_registration_date, '')
                if registration_date_raw and not pd.isna(registration_date_raw) and str(registration_date_raw).lower() != 'nan':
                    try:
                        from datetime import datetime as dt
                        if isinstance(registration_date_raw, dt):
                            registration_date = registration_date_raw.date()
                        else:
                            date_str = str(registration_date_raw).strip()
                            date_formats = [
                                '%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y',
                                '%d-%m-%Y', '%m-%d-%Y', '%d.%m.%Y', '%m.%d.%Y',
                                '%d/%m/%y', '%m/%d/%y', '%d-%m-%y', '%m-%d-%y',
                                '%d.%m.%y', '%m.%d.%y', '%Y/%m/%d', '%Y.%m.%d'
                            ]
                            for fmt in date_formats:
                                try:
                                    registration_date = dt.strptime(date_str, fmt).date()
                                    break
                                except ValueError:
                                    continue
                    except Exception:
                        registration_date = datetime.datetime.now().date()
                
                # Parse when_completed_treatments from Excel
                when_completed_treatments = None
                when_completed_treatments_raw = row.get(col_when_completed_treatments, '')
                if when_completed_treatments_raw and not pd.isna(when_completed_treatments_raw) and str(when_completed_treatments_raw).lower() != 'nan':
                    # Just accept the raw value as-is - this is a freeform text field
                    when_completed_treatments = str(when_completed_treatments_raw).strip()
                
                # Parse expected_end_treatment_by_protocol from Excel
                expected_end_treatment_by_protocol = None
                treatment_end_protocol_raw = row.get(col_treatment_end_protocol, '')
                if treatment_end_protocol_raw and not pd.isna(treatment_end_protocol_raw) and str(treatment_end_protocol_raw).lower() != 'nan':
                    # Just accept the raw value as-is - this is a freeform text field
                    expected_end_treatment_by_protocol = str(treatment_end_protocol_raw).strip()
                
                # Parse completed_treatments boolean from Excel
                has_completed_treatments_raw = row.get(col_completed_treatments, '')
                has_completed_treatments = None  # Can be True, False, or None
                if has_completed_treatments_raw and not pd.isna(has_completed_treatments_raw) and str(has_completed_treatments_raw).lower() != 'nan':
                    has_completed_str = str(has_completed_treatments_raw).lower().strip()
                    if has_completed_str in ['true', 'כן', 'yes', '1', 'true']:
                        has_completed_treatments = True
                    elif has_completed_str in ['false', 'לא', 'no', '0', 'false']:
                        has_completed_treatments = False
                
                # Parse marital_status from Excel if provided
                marital_status_raw = row.get(col_marital_status, '')
                marital_status = '' if (marital_status_raw is None or pd.isna(marital_status_raw) or str(marital_status_raw).lower() == 'nan') else str(marital_status_raw).strip()
                
                # Parse tutoring_status from Excel, default to 'למצוא_חונך' if not provided
                tutoring_status_raw = row.get(col_tutoring_status, '')
                tutoring_status_val = '' if (tutoring_status_raw is None or pd.isna(tutoring_status_raw) or str(tutoring_status_raw).lower() == 'nan') else str(tutoring_status_raw).strip()
                # Normalize: convert spaces to underscores for tutoring_status only
                tutoring_status_normalized = tutoring_status_val.replace(' ', '_') if tutoring_status_val else ''
                final_tutoring_status = tutoring_status_normalized if tutoring_status_normalized else 'למצוא_חונך'
                
                # Feature #3: Check סוג column for maturity (בוגר/ת = mature)
                # If סוג contains "בוגר", set need_review=False on import
                sug_raw = row.get('סוג', '')
                sug_val = '' if (sug_raw is None or pd.isna(sug_raw) or str(sug_raw).lower() == 'nan') else str(sug_raw).strip()
                dont_need_review_since_mature = 'בוגר' in sug_val  # Matches "בוגר/ת" and "בוגר"
                
                # ALSO check age >= 16 for maturity (independent of סוג column)
                # If child is 16+, also set need_review=False
                is_mature_by_age = False
                if birth_date:
                    from datetime import date as date_class
                    today = date_class.today()
                    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
                    is_mature_by_age = age >= 16
                
                # Parse status from Excel - REQUIRED field
                status_raw = row.get(col_status, '')
                status_val = '' if (status_raw is None or pd.isna(status_raw) or str(status_raw).lower() == 'nan') else str(status_raw).strip()
                
                # Validate that status is provided
                if not status_val:
                    result['status'] = 'Error'
                    result['details'] = f'סטטוס חובה - חסר או ריק'
                    error_count += 1
                    results.append(result)
                    continue
                
                # Validate that status is a valid enum value
                if status_val not in valid_statuses:
                    result['status'] = 'Error'
                    result['details'] = f'סטטוס "{status_val}" לא תקין. ערכים חוקיים: {", ".join(valid_statuses)}'
                    invalid_enum_count += 1
                    results.append(result)
                    continue
                
                final_status = status_val
                
                # Parse gender from Excel - female=True, male=False
                gender_val = row.get(col_gender, '')
                if isinstance(gender_val, bool):
                    gender = gender_val
                elif isinstance(gender_val, str):
                    gender = gender_val.lower() in ['true', 'נקבה', 'female', 'f', '1']
                else:
                    gender = False
                
                # birth_date already parsed earlier (needed for both dry-run and live import)
                
                # Parse שיחת ביקורת (review talk) - maps to last_review_talk_conducted
                # If value is "לא צריך" (string), set need_review=False; otherwise parse as date
                review_talk_raw = row.get('שיחת ביקורת', '')
                review_talk_value = '' if (review_talk_raw is None or pd.isna(review_talk_raw) or str(review_talk_raw).lower() == 'nan') else str(review_talk_raw).strip()
                last_review_talk_conducted = None
                need_review_from_review_talk = None  # Will be set if "לא צריך" is found
                
                if review_talk_value:
                    if review_talk_value == 'לא צריך':
                        # If marked as "לא צריך", set need_review to False
                        need_review_from_review_talk = False
                    else:
                        # Try to parse as date - handle multiple date formats including Hebrew months
                        try:
                            from datetime import datetime as dt
                            import re
                            
                            # If already a datetime object (from pandas), use it directly
                            if isinstance(review_talk_raw, dt):
                                last_review_talk_conducted = review_talk_raw.date()
                            else:
                                date_str = review_talk_value
                                
                                # Hebrew month names mapping
                                hebrew_months = {
                                    'ינואר': '01', 'יוני': '06', 'יולי': '07', 'אוגוסט': '08',
                                    'ספטמבר': '09', 'אוקטובר': '10', 'נובמבר': '11', 'דצמבר': '12',
                                    'ספטמבר': '09', 'אוקטובר': '10', 'נובמבר': '11', 'דצמבר': '12',
                                    'פברואר': '02', 'מרץ': '03', 'אפריל': '04', 'מאי': '05',
                                    'יוני': '06', 'יולי': '07', 'אגוסטוס': '08',
                                    # Short forms
                                    'ינ': '01', 'פב': '02', 'מר': '03', 'אפ': '04', 'מא': '05',
                                    'יו': '06', 'יול': '07', 'אוג': '08', 'ספ': '09', 'אוק': '10',
                                    'נו': '11', 'דצ': '12',
                                }
                                
                                # Try to parse Hebrew date format: "D Month" or "D Month YYYY"
                                for heb_month, month_num in hebrew_months.items():
                                    if heb_month in date_str:
                                        # Extract day and year if present
                                        day_match = re.search(r'(\d{1,2})\s+' + heb_month, date_str)
                                        year_match = re.search(heb_month + r'.*?(\d{4})', date_str)
                                        
                                        if day_match:
                                            day = int(day_match.group(1))
                                            year = int(year_match.group(1)) if year_match else datetime.now().year
                                            try:
                                                last_review_talk_conducted = dt(year, int(month_num), day).date()
                                                break
                                            except ValueError:
                                                continue
                                
                                # If not parsed yet, try standard formats
                                if not last_review_talk_conducted:
                                    date_formats = [
                                        '%Y-%m-%d %H:%M:%S',  # ISO with time
                                        '%Y-%m-%d',           # ISO
                                        '%d/%m/%Y',           # EU: dd/mm/yyyy
                                        '%m/%d/%Y',           # US: mm/dd/yyyy
                                        '%d-%m-%Y',           # dd-mm-yyyy
                                        '%m-%d-%Y',           # mm-dd-yyyy
                                        '%d.%m.%Y',           # dd.mm.yyyy
                                        '%m.%d.%Y',           # mm.dd.yyyy
                                        '%d/%m/%y',           # dd/mm/yy (2-digit year)
                                        '%m/%d/%y',           # mm/dd/yy
                                        '%d-%m-%y',           # dd-mm-yy
                                        '%m-%d-%y',           # mm-dd-yy
                                        '%d.%m.%y',           # dd.mm.yy
                                        '%m.%d.%y',           # mm.dd.yy
                                        '%Y/%m/%d',           # yyyy/mm/dd
                                        '%Y.%m.%d',           # yyyy.mm.dd
                                        '%Y',                 # Year only
                                    ]
                                    for fmt in date_formats:
                                        try:
                                            last_review_talk_conducted = dt.strptime(date_str, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                        except Exception:
                            last_review_talk_conducted = None
                
                # Parse num_of_siblings from Excel - default to 0 if missing or invalid
                num_of_siblings = 0
                num_siblings_raw = row.get(col_num_of_siblings, '')
                if num_siblings_raw and not pd.isna(num_siblings_raw) and str(num_siblings_raw).lower() != 'nan':
                    try:
                        num_of_siblings = int(str(num_siblings_raw).strip())
                        # Validate reasonable range (0-20)
                        if num_of_siblings < 0 or num_of_siblings > 20:
                            num_of_siblings = 0
                    except (ValueError, TypeError):
                        num_of_siblings = 0
                
                # (Address validation already done earlier - applies to both dry-run and wet-run)
                
                # Determine responsible_coordinator based on medical status and tutoring status
                # Pass child_status (final_status) to handle בריא/ז״ל children (they get "ללא")
                responsible_coordinator = get_responsible_coordinator_for_family(final_tutoring_status, child_status=final_status)
                
                # === DUPLICATE DETECTION ===
                # Check for duplicates: firstname + surname + city + any matching phone field
                # This prevents accidentally creating duplicate records
                duplicate_query = Children.objects.filter(
                    childfirstname=child_first_name,
                    childsurname=child_last_name,
                    city=city
                )
                
                is_duplicate = False
                if duplicate_query.exists():
                    # Found same name + city, now check if any phone matches
                    for existing_child in duplicate_query:
                        # Check if any of the phone fields match
                        phones_match = False
                        if phone and existing_child.child_phone_number and phone == existing_child.child_phone_number:
                            phones_match = True
                        if father_phone and existing_child.father_phone and father_phone == existing_child.father_phone:
                            phones_match = True
                        if mother_phone and existing_child.mother_phone and mother_phone == existing_child.mother_phone:
                            phones_match = True
                        
                        if phones_match:
                            is_duplicate = True
                            break
                
                if is_duplicate:
                    result['status'] = 'Skipped'
                    result['details'] = f'דילוג: דובליקט חשוד - {child_first_name} {child_last_name} ב{city} קיים כבר במערכת עם טלפון תואם'
                    result['enum_warnings'].append('דובליקט חשוד - בדוק יד')
                    skipped_count += 1
                    results.append(result)
                    continue
                
                # ONLY DO DATABASE WRITES IF NOT DRY RUN
                if not dry_run:
                    with transaction.atomic():
                        child = Children.objects.create(
                            child_id=child_id_int,
                            childfirstname=child_first_name,
                            childsurname=child_last_name,
                            city=city,
                            child_phone_number=phone if phone else None,
                            treating_hospital=hospital,
                            medical_diagnosis=diagnosis,
                            diagnosis_date=diagnosis_date,
                            date_of_birth=birth_date,
                            marital_status=marital_status,
                            num_of_siblings=num_of_siblings if num_of_siblings is not None else 0,
                            registrationdate=registration_date,
                            lastupdateddate=datetime.datetime.now().date(),
                            status=final_status,
                            tutoring_status=final_tutoring_status,
                            responsible_coordinator=responsible_coordinator,
                            gender=gender,
                            last_review_talk_conducted=last_review_talk_conducted,
                            street_and_apartment_number=street_and_apartment,
                            current_medical_state=medical_state,
                            when_completed_treatments=when_completed_treatments,
                            has_completed_treatments=has_completed_treatments,
                            expected_end_treatment_by_protocol=expected_end_treatment_by_protocol,
                            father_name=father_name if father_name else None,
                            father_phone=father_phone if father_phone else None,
                            mother_name=mother_name if mother_name else None,
                            mother_phone=mother_phone if mother_phone else None,
                            details_for_tutoring=tutoring_details if tutoring_details else None,
                            additional_info=additional_info if additional_info else None,
                            is_in_frame=is_in_frame if is_in_frame else None,
                            coordinator_comments=coordinator_comments if coordinator_comments else None,
                            # Feature #2 + #3: Auto-set need_review based on:
                            # - Set to False if: בריא/ז״ל status OR "לא צריך" in review_talk column OR בוגר (סוג column) OR age >= 16
                            # - Otherwise default to True
                            need_review=False if (final_status in ['בריא', 'ז״ל'] or need_review_from_review_talk is False or dont_need_review_since_mature or is_mature_by_age) else True
                        )
                        
                        # If surname is "XXX" (missing), create עדכון משפחה task for Families Coordinators
                        if needs_surname_task:
                            try:
                                # Get all Families Coordinators to verify missing surnames
                                coordinators = Staff.objects.filter(
                                    roles__role_name='Families Coordinator'
                                ).distinct()
                                
                                if coordinators.exists():
                                    task_desc = f"עדכון שם משפחה חסר לילד/ה {child_first_name} (כרגע: {child_last_name}) - תעודת זהות: {child_id}"
                                    for coordinator in coordinators:
                                        Tasks.objects.create(
                                            task_type='עדכון משפחה',
                                            description=task_desc,
                                            assigned_to=coordinator,
                                            assigned_by=user,
                                            date_assigned=datetime.datetime.now(),
                                            status='חדש'
                                        )
                            except Exception as e:
                                # Log task creation failure but don't fail the import
                                api_logger.warning(f"Failed to create surname update task for child {child_id}: {str(e)}")
                
                # RECORD RESULT (for both dry-run and real import)
                result['status'] = 'OK'
                result['details'] = f'נוצרה בהצלחה' if not dry_run else 'בדיקה בלבד: משפחה תקינה'
                success_count += 1
                results.append(result)
                
            except IntegrityError as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאת מסד נתונים: {str(e)[:50]}'
                error_count += 1
                results.append(result)
            except Exception as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאה כללית: {str(e)[:50]}'
                error_count += 1
                results.append(result)
        
        # Create result Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "תוצאות ייבוא"
        
        headers = ['שורה', 'שם פרטי ילד', 'שם משפחה ילד', 'תעודת זהות', 'סטטוס ייבוא', 'פרטים', 'אזהרות שדות']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.get('row_num', ''))
            ws.cell(row=row_idx, column=2, value=result['child_first_name'])
            ws.cell(row=row_idx, column=3, value=result['child_last_name'])
            ws.cell(row=row_idx, column=4, value=result.get('child_id', ''))
            
            status_cell = ws.cell(row=row_idx, column=5, value=result['status'])
            if result['status'] == 'OK':
                status_cell.fill = ok_fill
            elif result['status'] == 'Error':
                status_cell.fill = error_fill
            elif result['status'] == 'Skipped' or result['status'] == 'Warning':
                status_cell.fill = warning_fill
            
            ws.cell(row=row_idx, column=6, value=result.get('details', ''))
            
            # Add enum warnings if any
            if result.get('enum_warnings'):
                warnings_text = ' | '.join(result['enum_warnings'])
                ws.cell(row=row_idx, column=7, value=warnings_text)
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 60
        ws.column_dimensions['G'].width = 80
        ws.sheet_view.rightToLeft = True
        
        result_excel = BytesIO()
        wb.save(result_excel)
        result_excel.seek(0)
        
        # Log the import action - Log ALL imports (dry-run or live) as long as there are records processed
        if total_records > 0:
            log_api_action(
                request=request,
                action='CREATE_FAMILY_SUCCESS',
                affected_tables=['childsmile_app_children'],
                entity_type='Bulk Import Families',
                success=True,
                additional_data={
                    'total_records': total_records,
                    'success_count': success_count,
                    'error_count': error_count,
                    'skipped_count': skipped_count,
                    'dry_run': dry_run,
                    'is_bulk_import': True
                }
            )
        
        # Return results
        response_data = {
            'total': total_records,
            'success': success_count,
            'skipped': skipped_count,
            'error': error_count,
            'dry_run': dry_run,
            'message': f'✅ ייבוא הושלם: {success_count} מתוך {total_records} המשפחות יובאו בהצלחה' if success_count > 0 else '❌ לא הצליח לייבא משפחות',
            'has_errors': error_count > 0 or skipped_count > 0,
            'result_file_available': True
        }
        
        # If dry_run, return Excel file for preview
        if dry_run:
            response = FileResponse(result_excel, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            from datetime import datetime as dt
            response['Content-Disposition'] = f'attachment; filename="import_preview_families_{dt.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
        
        # For regular import, return JSON with Excel file bytes embedded as base64
        result_excel.seek(0)
        excel_base64 = base64.b64encode(result_excel.read()).decode('utf-8')
        response_data['result_file'] = excel_base64
        from datetime import datetime as dt
        response_data['result_filename'] = f"import_results_families_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return JsonResponse(response_data, status=200)
    
    except Exception as e:
        api_logger.error(f"Import families endpoint error: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='CREATE_FAMILY_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'error_type': type(e).__name__}
        )
        return JsonResponse(
            {'error': f'Unexpected error: {str(e)}'},
            status=500
        )


@conditional_csrf
@api_view(['GET'])
def get_settlements_data(request):
    """
    GET /api/settlements/
    Returns all settlements and their streets from database.
    Replaces the 65MB JSON file load.
    Used by frontend to populate city/street modals.
    """
    user_id = request.session.get("user_id")
    if not user_id:
        return JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )
    
    try:        
        # Get all settlements from database
        settlements = SettlementsStreets.objects.all().values('city_name', 'streets')
        
        # Convert to JSON format matching the old JSON structure
        settlements_dict = {}
        for settlement in settlements:
            city_name = settlement['city_name']
            streets = settlement['streets'] if settlement['streets'] else []
            settlements_dict[city_name] = streets
        
        return JsonResponse(settlements_dict, status=200, safe=False)
    
    except Exception as e:
        api_logger.error(f"Get settlements error: {str(e)}")
        return JsonResponse(
            {'error': f'Error fetching settlements: {str(e)}'},
            status=500
        )