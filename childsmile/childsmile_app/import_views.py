"""
Import Views
============
Handles bulk import of volunteers and families from Excel files.
Only available when BLOCK_ACCESS_AFTER_APPROVAL feature flag is enabled.
"""

import os
import pandas as pd
from datetime import datetime
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.core.files.uploadedfile import UploadedFile
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.http import FileResponse

from .models import (
    SignedUp, Staff, Role, General_Volunteer, Pending_Tutor, Tutors
)
from .utils import conditional_csrf
from .audit_utils import log_api_action
from .logger import api_logger


# ==================== FEATURE FLAG CHECK ====================
def is_import_enabled():
    """Check if bulk import feature is enabled via environment variable."""
    return os.environ.get("BLOCK_ACCESS_AFTER_APPROVAL", "False").lower() == "true"


def check_import_permission(request):
    """Verify user has permission and feature flag is enabled."""
    if not is_import_enabled():
        raise PermissionError("Bulk import feature is not enabled")
    
    # TODO: Add role-based permission check if needed
    # For now, require admin/staff with import permissions
    if not hasattr(request.user, 'staff_profile'):
        raise PermissionError("User must be staff member")
    
    api_logger.info(f"Import permission granted for user: {request.user}")


# ==================== HELPER FUNCTIONS ====================
def clean_email(email_val):
    """Clean email: strip whitespace and newlines."""
    if not email_val or pd.isna(email_val):
        return None
    email = str(email_val).strip().replace('\n', '').replace('\r', '')
    if not email or email.lower() == 'nan':
        return None
    return email


def parse_gender(gender_val):
    """Parse gender value to boolean (True = Female, False = Male)."""
    if isinstance(gender_val, bool):
        return gender_val
    if isinstance(gender_val, str):
        gender_lower = gender_val.lower().strip()
        if gender_lower in ['true', 'נקבה', 'female', 'f', '1']:
            return True
        return False
    return False


def parse_want_tutor(val):
    """Parse want_tutor value to boolean."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        val_lower = val.lower().strip()
        if val_lower in ['true', 'כן', 'yes', '1']:
            return True
    return False


def parse_status(status_val):
    """Parse status value from Excel - keep original value."""
    if not status_val or pd.isna(status_val) or str(status_val).lower() == 'nan':
        return None
    return str(status_val).strip()


def parse_birth_date(date_val):
    """Parse birth date from Excel. Returns date object or None."""
    if not date_val or pd.isna(date_val) or str(date_val).lower() == 'nan':
        return None
    try:
        if isinstance(date_val, datetime):
            return date_val.date()
        date_str = str(date_val).strip()
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y']:
            try:
                parsed_date = datetime.strptime(date_str, fmt).date()
                today = datetime.now().date()
                age = today.year - parsed_date.year
                if (today.month, today.day) < (parsed_date.month, parsed_date.day):
                    age -= 1
                if 13 <= age <= 120:
                    return parsed_date
                continue
            except ValueError:
                continue
        return None
    except Exception:
        return None


def calculate_age_from_birth_date(birth_date):
    """Calculate age from birth date."""
    if not birth_date:
        return 0
    today = datetime.now().date()
    age = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    return max(0, age)


def get_clean_string(val):
    """Get a clean string value, or empty string if None/NaN."""
    if val is None or pd.isna(val) or str(val).lower() == 'nan':
        return ''
    return str(val).strip()


def clean_city(city_val):
    """Clean and normalize city value."""
    city_mapping = {
        'תל אביב': 'תל אביב - יפו',
        'מודיעין': 'מודיעין-מכבים-רעות',
        'מודעין': 'מודיעין-מכבים-רעות',
        'פתח תקוה': 'פתח תקווה',
        'קריית אתא': 'קרית אתא',
        'קריית נטפים': 'קרית נטפים',
        'יהוד מונוסון': 'יהוד-מונוסון',
        'קיבוץ חפץ חיים': 'חפץ חיים',
        'מושב בני ראם': 'בני ראם',
        'מושב חמד': 'חמד',
        'מושב פורת': 'פורת',
        'יד רמב״ם (מושב)': 'יד רמבם',
        'יישוב נופים': 'נופים',
        'מגד אל כרום': 'מג\'ד אל-כרום',
        'גבעת שמואל אבל עושה שירות': 'גבעת שמואל',
        'מושה טפחות': 'טפחות',
        'עלי זהב לומד בשדרות': 'עלי זהב',
        'ירושלים- תא': 'ירושלים',
        'ראשל״צ': 'ראשון לציון',
        'הדר גנים': 'גנות הדר',
        'רעננה(מגדל עוז)': 'רעננה',
    }
    
    city = get_clean_string(city_val)
    if not city:
        return ''
    
    city = city.replace('\n', ' ').replace('\r', '').strip()
    
    for separator in ['/', ',']:
        if separator in city:
            city = city.split(separator)[0].strip()
    
    if city in city_mapping:
        return city_mapping[city]
    
    for key, value in city_mapping.items():
        if key in city:
            return value
    
    return city


def create_result_excel(results):
    """Create an Excel file with import results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "תוצאות ייבוא"
    
    headers = ['שורה', 'שם פרטי', 'שם משפחה', 'מייל', 'סטטוס ייבוא', 'סוג רשומה', 'פרטים']
    
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result.get('row_num', ''))
        ws.cell(row=row_idx, column=2, value=result['first_name'])
        ws.cell(row=row_idx, column=3, value=result['surname'])
        ws.cell(row=row_idx, column=4, value=result.get('email', ''))
        
        status_cell = ws.cell(row=row_idx, column=5, value=result['status'])
        if result['status'] == 'OK':
            status_cell.fill = ok_fill
        elif result['status'] == 'Error':
            status_cell.fill = error_fill
        elif result['status'] == 'Warning':
            status_cell.fill = warning_fill
        
        ws.cell(row=row_idx, column=6, value=result.get('record_type', ''))
        ws.cell(row=row_idx, column=7, value=result.get('details', ''))
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60
    ws.sheet_view.rightToLeft = True
    
    # Return as bytes instead of saving to disk
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ==================== IMPORT ENDPOINTS ====================

@conditional_csrf
@api_view(['POST'])
@permission_classes([IsAuthenticated])
def import_volunteers_endpoint(request):
    """
    POST /api/import/volunteers/
    
    Accepts an Excel file and imports volunteers/tutors into the system.
    
    Request:
        - file: Excel file (.xlsx)
        - dry_run: boolean (optional, default=False) - validate without creating records
    
    Returns:
        - JSON response with summary and download link to results file
        - or Excel file if dry_run=True with preview
    """
    try:
        # Check feature flag
        check_import_permission(request)
        
        # Get uploaded file
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file: UploadedFile = request.FILES['file']
        
        # Validate file extension
        if not file.name.endswith('.xlsx'):
            return Response(
                {'error': 'File must be .xlsx format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        dry_run = request.data.get('dry_run', 'false').lower() == 'true'
        
        # Read Excel file
        try:
            df = pd.read_excel(file, dtype=str)
        except Exception as e:
            api_logger.error(f"Failed to read Excel file: {str(e)}")
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if df.empty:
            return Response(
                {'error': 'Excel file is empty'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get required roles
        try:
            general_volunteer_role = Role.objects.get(role_name="General Volunteer")
        except Role.DoesNotExist:
            return Response(
                {'error': 'Role "General Volunteer" not found in database'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        try:
            tutor_role = Role.objects.get(role_name="Tutor")
        except Role.DoesNotExist:
            return Response(
                {'error': 'Role "Tutor" not found in database'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        
        # Process records
        total_records = len(df)
        success_count = 0
        error_count = 0
        skipped_count = 0
        general_volunteer_count = 0
        tutor_with_tutee_count = 0
        tutor_no_tutee_count = 0
        pending_tutor_count = 0
        
        results = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            first_name = get_clean_string(row.get('שם פרטי', ''))
            surname = get_clean_string(row.get('שם משפחה', ''))
            
            result = {
                'row_num': row_num,
                'first_name': first_name,
                'surname': surname,
                'email': '',
                'status': '',
                'record_type': '',
                'details': ''
            }
            
            try:
                # Parse data
                id_number = get_clean_string(row.get('תעודת זהות', ''))
                phone = get_clean_string(row.get('מספר טלפון', ''))
                email = clean_email(row.get('מייל'))
                city = clean_city(row.get('עיר מגורים', ''))
                birth_date = parse_birth_date(row.get('תאריך לידה', ''))
                calculated_age = calculate_age_from_birth_date(birth_date)
                gender = parse_gender(row.get('מין', ''))
                want_tutor = parse_want_tutor(row.get('סוג התנדבות', ''))
                status_val = parse_status(row.get('סטטוס', ''))
                volunteer_comment = get_clean_string(row.get('הערות המתנדב', ''))
                coordinator_comment = get_clean_string(row.get('הערות הרכז', ''))
                
                result['email'] = email or ''
                
                # Build comments
                signedup_comment_parts = []
                tutor_preferences = None
                
                if want_tutor:
                    if volunteer_comment:
                        tutor_preferences = volunteer_comment
                    if coordinator_comment:
                        signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
                else:
                    if volunteer_comment:
                        signedup_comment_parts.append(f"הערות מתנדב: {volunteer_comment}")
                    if coordinator_comment:
                        signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
                
                signedup_comment = ' | '.join(signedup_comment_parts) if signedup_comment_parts else None
                
                # Validate
                if not first_name or not surname:
                    result['status'] = 'Error'
                    result['details'] = 'חסר שם פרטי או שם משפחה'
                    error_count += 1
                    results.append(result)
                    continue
                
                if not id_number:
                    result['status'] = 'Error'
                    result['details'] = 'חסרה תעודת זהות'
                    error_count += 1
                    results.append(result)
                    continue
                
                try:
                    id_int = int(id_number)
                except ValueError:
                    result['status'] = 'Error'
                    result['details'] = f'תעודת זהות לא מספרית: {id_number}'
                    error_count += 1
                    results.append(result)
                    continue
                
                # Check duplicates
                if SignedUp.objects.filter(id=id_int).exists():
                    result['status'] = 'Skipped'
                    result['details'] = f'ת.ז. כבר קיימת במערכת'
                    skipped_count += 1
                    results.append(result)
                    continue
                
                if email and Staff.objects.filter(email=email).exists():
                    result['status'] = 'Skipped'
                    result['details'] = f'מייל כבר קיים במערכת'
                    skipped_count += 1
                    results.append(result)
                    continue
                
                # Determine record type
                if want_tutor:
                    if status_val == "יש חניך":
                        record_type = "חונך - יש חניך"
                    elif status_val == "אין חניך":
                        record_type = "חונך - אין חניך"
                    else:
                        record_type = f"חונך ממתין - {status_val or 'ללא סטטוס'}"
                else:
                    record_type = "מתנדב כללי"
                
                result['record_type'] = record_type
                
                if dry_run:
                    result['status'] = 'OK'
                    result['details'] = f'בדיקה בלבד: {record_type}'
                    success_count += 1
                    results.append(result)
                    continue
                
                # === CREATE RECORDS ===
                with transaction.atomic():
                    # 1. Create SignedUp
                    signedup = SignedUp.objects.create(
                        id=id_int,
                        first_name=first_name,
                        surname=surname,
                        age=calculated_age,
                        birth_date=birth_date,
                        gender=gender,
                        phone=phone,
                        city=city,
                        comment=signedup_comment,
                        email=email,
                        want_tutor=want_tutor,
                    )
                    
                    # 2. Create Staff with unique username
                    username = f"{first_name}_{surname}"
                    index = 1
                    original_username = username
                    while Staff.objects.filter(username=username).exists():
                        username = f"{original_username}_{index}"
                        index += 1
                    
                    staff = Staff.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=surname,
                        created_at=now(),
                        registration_approved=True,
                        is_active=True,
                        deactivation_reason="suspended"
                    )
                    
                    # 3. Create role-specific record
                    if want_tutor:
                        staff.roles.add(tutor_role)
                        
                        if status_val == "יש חניך":
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="יש_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            tutor_with_tutee_count += 1
                        elif status_val == "אין חניך":
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="אין_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            tutor_no_tutee_count += 1
                        else:
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="אין_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            Pending_Tutor.objects.create(
                                id_id=signedup.id,
                                pending_status="ממתין",
                            )
                            pending_tutor_count += 1
                    else:
                        staff.roles.add(general_volunteer_role)
                        General_Volunteer.objects.create(
                            id_id=signedup.id,
                            staff_id=staff.staff_id,
                            signupdate=now().date(),
                            comments="",
                        )
                        general_volunteer_count += 1
                
                result['status'] = 'OK'
                result['details'] = f'נוצר בהצלחה: {record_type}'
                success_count += 1
                results.append(result)
                
            except IntegrityError as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאת מסד נתונים: {str(e)}'
                error_count += 1
                results.append(result)
            except Exception as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאה כללית: {str(e)}'
                error_count += 1
                results.append(result)
        
        # Create result file
        result_excel = create_result_excel(results)
        
        # Log action
        log_api_action(
            request.user,
            'IMPORT_VOLUNTEERS',
            'ייבוא מתנדבים',
            {'total': total_records, 'success': success_count, 'errors': error_count, 'dry_run': dry_run}
        )
        
        # Return results
        response_data = {
            'total': total_records,
            'success': success_count,
            'skipped': skipped_count,
            'error': error_count,
            'dry_run': dry_run,
            'breakdown': {
                'general_volunteer': general_volunteer_count,
                'tutor_with_tutee': tutor_with_tutee_count,
                'tutor_no_tutee': tutor_no_tutee_count,
                'pending_tutor': pending_tutor_count,
            }
        }
        
        # If dry_run, return Excel file for preview
        if dry_run:
            response = FileResponse(result_excel, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="import_preview_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
        
        # Otherwise return JSON summary
        return Response(response_data, status=status.HTTP_200_OK)
    
    except PermissionError as e:
        api_logger.warning(f"Import permission denied: {str(e)}")
        return Response(
            {'error': str(e)},
            status=status.HTTP_403_FORBIDDEN
        )
    except Exception as e:
        api_logger.error(f"Import endpoint error: {str(e)}", exc_info=True)
        return Response(
            {'error': f'Unexpected error: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

