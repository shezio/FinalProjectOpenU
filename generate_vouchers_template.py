# -*- coding: utf-8 -*-
"""
Generate the Hebrew Excel template Liam fills in for a voucher distribution
(חלוקת תלושים). The filled workbook is later handed back and turned into SQL
INSERTs for:
    * childsmile_app_voucherdistribution  (גיליון "חלוקות")
    * childsmile_app_voucherrecipient      (גיליון "מקבלים")

Every column header states the DATA TYPE + whether it's NULLABLE, colour-coded:
    orange  = חובה            (NOT NULL – must be filled)
    yellow  = חובה מותנה       (nullable in the DB, but required by the questionnaire)
    blue    = אופציונלי         (NULL – may be left empty)
    green   = ברירת מחדל / אוטומטי (has a DB default / auto-filled)

Run:  python3 generate_vouchers_template.py
Out:  vouchers_import_template_he.xlsx
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_FILE = "vouchers_import_template_he.xlsx"

# ── enum value lists (must match models.py / voucher_views.py EXACTLY) ────────
VOUCHER_TYPES = ["רמי לוי", "תו פלוס - קרפור", "אחר"]
QUESTIONNAIRE_TYPES = ["עמותה", "כללי", "ללא", "עמותה וכללי"]
TREATMENT_STATUSES = ["טיפולים", "מעקבים", "אחזקה", "ז״ל", "בריא", "עזב"]
DELIVERED_STATUSES = ["כן", "איסוף עצמי", "לא"]
YES_NO = ["כן", "לא"]

# ── colours per nullability "level" ──────────────────────────────────────────
FILL = {
    "required":    PatternFill("solid", fgColor="F4B183"),  # orange
    "conditional": PatternFill("solid", fgColor="FFE699"),  # yellow
    "optional":    PatternFill("solid", fgColor="BDD7EE"),  # blue
    "default":     PatternFill("solid", fgColor="C6E0B4"),  # green
}
HEADER_FONT = Font(name="Arial", bold=True, size=11, color="1F1F1F")
THIN = Side(style="thin", color="9E9E9E")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _header_text(label, type_he, null_he):
    """The visible header cell: label on line 1, (type · nullability) on line 2."""
    return f"{label}\n({type_he} · {null_he})"


def _build_sheet(wb, title, columns):
    """columns: list of dicts with keys
       label, type_he, null_he, level, db, comment, width, values(optional)."""
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True

    for idx, col in enumerate(columns, start=1):
        letter = get_column_letter(idx)
        cell = ws.cell(row=1, column=idx, value=_header_text(col["label"], col["type_he"], col["null_he"]))
        cell.font = HEADER_FONT
        cell.fill = FILL[col["level"]]
        cell.alignment = WRAP_CENTER
        cell.border = BORDER
        cell.comment = Comment(col["comment"], "ChildSmile")
        cell.comment.width = 320
        cell.comment.height = 160
        ws.column_dimensions[letter].width = col["width"]

        if col.get("values"):
            options = ",".join(col["values"])
            dv = DataValidation(
                type="list",
                formula1=f'"{options}"',
                allow_blank=True,
                showErrorMessage=True,
            )
            dv.error = "יש לבחור ערך מהרשימה בלבד."
            dv.errorTitle = "ערך לא תקין"
            dv.prompt = "בחר/י מהרשימה: " + " / ".join(col["values"])
            dv.promptTitle = col["label"]
            ws.add_data_validation(dv)
            dv.add(f"{letter}2:{letter}2000")

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 46
    return ws


# ── גיליון 1: חלוקות (VoucherDistribution) ───────────────────────────────────
DISTRIBUTION_COLUMNS = [
    dict(label="שם החלוקה", type_he="טקסט", null_he="חובה", level="required",
         db="name — VARCHAR(255) NOT NULL", width=28,
         comment='name — VARCHAR(255), NOT NULL (חובה).\nשם ייחודי לחלוקה, למשל: "חלוקת תלושי פסח 2026".\nמשמש כמפתח לקישור השורות בגיליון "מקבלים".'),
    dict(label="סוג תו", type_he="רשימה", null_he="חובה", level="required",
         db="voucher_type — VARCHAR(30) NOT NULL", width=20, values=VOUCHER_TYPES,
         comment="voucher_type — VARCHAR(30), NOT NULL (חובה).\nערך אחד בלבד מתוך: רמי לוי / תו פלוס - קרפור / אחר."),
    dict(label="סכום התחלתי (₪)", type_he="מספר עשרוני", null_he="חובה", level="required",
         db="initial_amount — NUMERIC(12,2) NOT NULL", width=18,
         comment="initial_amount — NUMERIC(12,2), NOT NULL (חובה).\nהתקציב הכולל של החלוקה, למשל 15000 או 15000.50.\nללא סימן ₪ ופסיקים — ספרות ונקודה עשרונית בלבד."),
    dict(label="תאריך התחלה", type_he="תאריך", null_he="אופציונלי", level="optional",
         db="start_date — DATE NULL", width=16,
         comment="start_date — DATE, NULL (אופציונלי).\nפורמט מומלץ: YYYY-MM-DD (למשל 2026-03-25)."),
    dict(label="תאריך סיום", type_he="תאריך", null_he="אופציונלי", level="optional",
         db="end_date — DATE NULL", width=16,
         comment="end_date — DATE, NULL (אופציונלי).\nפורמט מומלץ: YYYY-MM-DD."),
    dict(label="הושלם (דרכנו)", type_he="כן/לא", null_he="ברירת מחדל: לא", level="default",
         db="is_completed — BOOLEAN DEFAULT FALSE", width=16, values=YES_NO,
         comment="is_completed — BOOLEAN, NOT NULL DEFAULT FALSE.\nאם ריק — ייחשב 'לא'. 'כן' = החלוקה הסתיימה."),
    dict(label="סוג שאלון", type_he="רשימה", null_he="ברירת מחדל: ללא", level="default",
         db="questionnaire_type — VARCHAR(10) DEFAULT 'ללא'", width=18, values=QUESTIONNAIRE_TYPES,
         comment="questionnaire_type — VARCHAR(10), NOT NULL DEFAULT 'ללא'.\nעמותה / כללי / ללא / עמותה וכללי.\nקובע אילו שדות חובה בגיליון 'מקבלים'."),
    dict(label="הערות", type_he="טקסט חופשי", null_he="אופציונלי", level="optional",
         db="notes — TEXT NULL", width=32,
         comment="notes — TEXT, NULL (אופציונלי). עד 4000 תווים."),
]

# ── גיליון 2: מקבלים (VoucherRecipient) ──────────────────────────────────────
RECIPIENT_COLUMNS = [
    dict(label="שם החלוקה", type_he="טקסט", null_he="חובה", level="required",
         db="distribution (FK) — INTEGER NOT NULL", width=26,
         comment='קישור לחלוקה (distribution_id, FK, NOT NULL).\nחייב להתאים בדיוק לערך "שם החלוקה" מגיליון "חלוקות".'),
    dict(label="שם מלא", type_he="טקסט", null_he="חובה", level="required",
         db="full_name — VARCHAR(255) NOT NULL", width=24,
         comment="full_name — VARCHAR(255), NOT NULL (חובה).\nשם ההורה / הפונה."),
    dict(label="טלפון", type_he="טקסט", null_he="אופציונלי (חובה בשאלון)", level="conditional",
         db="phone — VARCHAR(20) NULL", width=16,
         comment="phone — VARCHAR(20), NULL בבסיס הנתונים, אך חובה בשאלון הציבורי.\n10 ספרות שמתחילות ב-0 (למשל 0541234567). ספרות בלבד."),
    dict(label='ת"ז הורה', type_he="טקסט/מספר", null_he="אופציונלי", level="optional",
         db="parent_id_number — VARCHAR(20) NULL", width=16,
         comment='parent_id_number — VARCHAR(20), NULL (אופציונלי).\n5-9 ספרות. ללא בדיקת ספרת ביקורת.'),
    dict(label="שם הילד/ה", type_he="טקסט", null_he="אופציונלי (חובה בעמותה)", level="conditional",
         db="child_name — VARCHAR(255) NULL", width=22,
         comment="child_name — VARCHAR(255), NULL בבסיס הנתונים.\nחובה עבור משפחות עמותה (questionnaire_type=עמותה)."),
    dict(label='ת"ז הילד/ה', type_he="טקסט/מספר", null_he="אופציונלי (חובה בעמותה)", level="conditional",
         db="child_id_number — VARCHAR(20) NULL", width=16,
         comment='child_id_number — VARCHAR(20), NULL בבסיס הנתונים.\nחובה עבור משפחות עמותה. 5-9 ספרות.\nמשמש לקישור אוטומטי למשפחה קיימת (ת"ז אמיתית).'),
    dict(label="מצב טיפול", type_he="רשימה", null_he="אופציונלי", level="optional",
         db="child_treatment_status — VARCHAR(50) NULL", width=16, values=TREATMENT_STATUSES,
         comment="child_treatment_status — VARCHAR(50), NULL (אופציונלי).\nערך אחד מתוך: טיפולים / מעקבים / אחזקה / ז״ל / בריא / עזב.\nרלוונטי בעיקר למשפחות עמותה."),
    dict(label="מספר ילדים בבית", type_he="מספר שלם", null_he="אופציונלי", level="optional",
         db="num_children_at_home — INTEGER NULL", width=15,
         comment="num_children_at_home — INTEGER, NULL (אופציונלי).\nמספר שלם בטווח 0-30."),
    dict(label="עיר", type_he="טקסט", null_he="אופציונלי", level="optional",
         db="city — VARCHAR(255) NULL", width=16,
         comment="city — VARCHAR(255), NULL (אופציונלי)."),
    dict(label="כתובת", type_he="טקסט", null_he="אופציונלי", level="optional",
         db="street_address — VARCHAR(255) NULL", width=24,
         comment="street_address — VARCHAR(255), NULL (אופציונלי).\nרחוב + מספר + קומה + דירה."),
    dict(label="תיאור המקרה", type_he="טקסט חופשי", null_he="אופציונלי", level="optional",
         db="case_description — TEXT NULL", width=30,
         comment="case_description — TEXT, NULL (אופציונלי). עד 4000 תווים."),
    dict(label="גורם מפנה", type_he="טקסט", null_he="אופציונלי (חובה בכללי)", level="conditional",
         db="referral_source — VARCHAR(255) NULL", width=20,
         comment="referral_source — VARCHAR(255), NULL בבסיס הנתונים.\nחובה עבור פניות כלליות (questionnaire_type=כללי)."),
    dict(label="חותמת זמן", type_he="תאריך ושעה", null_he="אופציונלי", level="optional",
         db="submitted_at — TIMESTAMP NULL", width=18,
         comment="submitted_at — TIMESTAMP, NULL (אופציונלי).\nמועד מילוי הטופס המקורי. פורמט: YYYY-MM-DD HH:MM.\nהשאר/י ריק עבור שורות שהוזנו ידנית."),
    dict(label="סכום מאושר (₪)", type_he="מספר עשרוני", null_he="אופציונלי", level="optional",
         db="approved_amount — NUMERIC(10,2) NULL", width=16,
         comment="approved_amount — NUMERIC(10,2), NULL (אופציונלי).\nסכום התלוש שאושר לנתמך. ספרות ונקודה בלבד.\nסך כל הסכומים המאושרים בחלוקה לא יעלה על 'סכום התחלתי'."),
    dict(label="מוכן", type_he="כן/לא", null_he="ברירת מחדל: לא", level="default",
         db="ready — BOOLEAN DEFAULT FALSE", width=12, values=YES_NO,
         comment="ready — BOOLEAN, NOT NULL DEFAULT FALSE.\nאם ריק — ייחשב 'לא'."),
    dict(label="מתנדב", type_he="טקסט", null_he="אופציונלי", level="optional",
         db="assigned_volunteer — VARCHAR(255) NULL", width=18,
         comment="assigned_volunteer — VARCHAR(255), NULL (אופציונלי).\nשם המתנדב המשייך/מוסר. טקסט חופשי."),
    dict(label="נמסר", type_he="רשימה", null_he="אופציונלי", level="optional",
         db="delivered — VARCHAR(20) NULL", width=14, values=DELIVERED_STATUSES,
         comment="delivered — VARCHAR(20), NULL (אופציונלי).\nערך אחד מתוך: כן / איסוף עצמי / לא."),
    dict(label="תאריך מסירה", type_he="תאריך", null_he="אופציונלי (חובה כאשר נמסר=כן/איסוף עצמי)", level="conditional",
         db="delivered_date — DATE NULL", width=16,
         comment="delivered_date — DATE, NULL בבסיס הנתונים.\nמועד המסירה בפועל. פורמט: YYYY-MM-DD.\nחובה כאשר 'נמסר' = כן או איסוף עצמי (במערכת ברירת המחדל היא היום)."),
    dict(label="הערות", type_he="טקסט חופשי", null_he="אופציונלי", level="optional",
         db="notes — TEXT NULL", width=26,
         comment="notes — TEXT, NULL (אופציונלי). עד 4000 תווים. למשל 'סל מזון'."),
]


def _build_instructions(wb):
    ws = wb.create_sheet(title="הוראות")
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 95

    title_font = Font(name="Arial", bold=True, size=16, color="1F4E79")
    h_font = Font(name="Arial", bold=True, size=12, color="1F4E79")
    body = Font(name="Arial", size=11)
    right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    def put(row, text, font, fill=None, height=None, col="B", span=True):
        c = ws.cell(row=row, column=2 if col == "B" else 3, value=text)
        c.font = font
        c.alignment = right
        if fill:
            c.fill = fill
        if span:
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        if height:
            ws.row_dimensions[row].height = height

    put(2, "תבנית הזנת נתונים — חלוקת תלושים (חיוך של ילד)", title_font, height=26)
    put(4, "כיצד למלא", h_font)
    put(5, "1. גיליון «חלוקות» — הזינו שורה אחת לכל חלוקה (בדרך כלל שורה אחת).", body, height=20)
    put(6, "2. גיליון «מקבלים» — שורה לכל משפחה/נתמך. עמודת «שם החלוקה» חייבת להתאים בדיוק לשם מגיליון «חלוקות».", body, height=20)
    put(7, "3. אין למחוק, לשנות או להוסיף עמודות ואין לשנות את שורת הכותרת. מתחילים למלא משורה 2.", body, height=20)
    put(8, "4. בעמודות עם רשימה נפתחת — בחרו ערך מהרשימה בלבד (אחרת ההזנה תיחסם).", body, height=20)
    put(9, "5. תאריכים בפורמט YYYY-MM-DD, סכומים בספרות ונקודה עשרונית בלבד (ללא ₪ וללא פסיקים).", body, height=20)
    put(10, "6. כל כותרת עמודה מציינת את סוג הנתון והאם ניתן להשאיר ריק. פרטים מלאים בהערה שעל הכותרת (המשולש האדום).", body, height=20)

    put(12, "מקרא צבעי הכותרות", h_font)
    legend = [
        ("חובה", "required", "שדה חובה (NOT NULL) — אסור להשאיר ריק."),
        ("חובה מותנה", "conditional", "ניתן להשאיר ריק בבסיס הנתונים, אך חובה לפי סוג השאלון/הטופס (עמותה / כללי / שאלון ציבורי)."),
        ("אופציונלי", "optional", "ניתן להשאיר ריק (NULL)."),
        ("ברירת מחדל / אוטומטי", "default", "אם יישאר ריק — ייכנס ערך ברירת מחדל (למשל 'לא')."),
    ]
    r = 13
    for name, level, desc in legend:
        cA = ws.cell(row=r, column=2, value=name)
        cA.fill = FILL[level]
        cA.font = Font(name="Arial", bold=True, size=11)
        cA.alignment = right
        cA.border = BORDER
        cB = ws.cell(row=r, column=3, value=desc)
        cB.font = body
        cB.alignment = right
        r += 1

    put(r + 1, "ערכים מותרים ברשימות הנפתחות", h_font)
    r += 2
    enum_rows = [
        ("סוג תו", " / ".join(VOUCHER_TYPES)),
        ("סוג שאלון", " / ".join(QUESTIONNAIRE_TYPES)),
        ("מצב טיפול", " / ".join(TREATMENT_STATUSES)),
        ("נמסר", " / ".join(DELIVERED_STATUSES)),
        ("הושלם / מוכן", " / ".join(YES_NO)),
    ]
    for name, vals in enum_rows:
        cA = ws.cell(row=r, column=2, value=name)
        cA.font = Font(name="Arial", bold=True, size=11)
        cA.alignment = right
        cB = ws.cell(row=r, column=3, value=vals)
        cB.font = body
        cB.alignment = right
        r += 1

    put(r + 1, "שדות שאין למלא (מנוהלים אוטומטית): מזהה, תאריך יצירה/עדכון, מקושר למשפחה — נוצרים אוטומטית במערכת.",
        Font(name="Arial", italic=True, size=10, color="808080"), height=30)

    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet

    _build_instructions(wb)
    _build_sheet(wb, "חלוקות", DISTRIBUTION_COLUMNS)
    _build_sheet(wb, "מקבלים", RECIPIENT_COLUMNS)

    wb.active = 0  # open on the instructions sheet
    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE} with sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
