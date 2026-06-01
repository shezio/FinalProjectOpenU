"""
Management command: import_refunds_2026
=======================================
Imports historical 2026 expense refund data from an Excel spreadsheet.
Receipt files are NOT required (file_url left null).

Usage:
    python manage.py import_refunds_2026 --file /path/to/refunds_2026.xlsx [--dry-run]

Expected Excel columns (Hebrew headers):
    שם מלא           → staff_full_name (used to look up Staff by first_name + last_name)
    אימייל           → staff email (used as primary lookup key)
    תאריך הוצאה      → expense_date (dd/mm/yyyy or yyyy-mm-dd)
    סכום מבוקש       → requested_amount
    סכום שאושר       → approved_amount (can be empty)
    תיאור            → description
    הערת מתנדב       → volunteer_comment (optional)
    הערת מנהל        → admin_comment (optional)
    אושר על ידי      → approved_by (one of: נעם/טל/נבו/אורי/ליאם, optional)
    סטטוס            → status (Hebrew: ממתין/אושר/אושר חלקית/שולם/בוטל/נדחה)
    אמצעי תשלום      → refund_method (optional)
    טלפון            → phone_number (optional)

Rows with missing mandatory fields are SKIPPED and logged.
Duplicate detection: skips rows where (staff + expense_date + requested_amount) already exists.
"""

import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from django.db import transaction

from childsmile_app.models import Staff, ExpenseRefund
from childsmile_app.logger import api_logger


COLUMN_MAP = {
    'שם מלא': 'staff_full_name',
    'אימייל': 'email',
    'תאריך הוצאה': 'expense_date',
    'סכום מבוקש': 'requested_amount',
    'סכום שאושר': 'approved_amount',
    'תיאור': 'description',
    'הערת מתנדב': 'volunteer_comment',
    'הערת מנהל': 'admin_comment',
    'אושר על ידי': 'approved_by',
    'סטטוס': 'status',
    'אמצעי תשלום': 'refund_method',
    'טלפון': 'phone_number',
}

VALID_STATUSES = ['ממתין', 'אושר', 'אושר חלקית', 'שולם', 'בוטל/נדחה']
VALID_METHODS = ['ביט', 'פייבוקס', 'העברה בנקאית', 'אשראי', 'מזומן']
VALID_APPROVED_BY = ['נעם', 'טל', 'נבו', 'אורי', 'ליאם']


def _parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    val = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        return Decimal(str(val).replace(',', '.').strip())
    except InvalidOperation:
        return None


class Command(BaseCommand):
    help = 'Import historical 2026 expense refund data from an Excel file (no receipt files required)'

    def add_arguments(self, parser):
        parser.add_argument('--file', type=str, required=True,
                            help='Absolute path to the Excel (.xlsx) file')
        parser.add_argument('--dry-run', action='store_true', default=False,
                            help='Validate and report without writing to the database')

    def handle(self, *args, **options):
        filepath = options['file']
        dry_run = options['dry_run']

        self.stdout.write(f"{'[DRY RUN] ' if dry_run else ''}Starting refund import from: {filepath}")

        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR('openpyxl is not installed. Run: pip install openpyxl'))
            return

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
        except FileNotFoundError:
            self.stderr.write(self.style.ERROR(f'File not found: {filepath}'))
            return
        except Exception as e:
            self.stderr.write(self.style.ERROR(f'Failed to open Excel file: {e}'))
            return

        # Read header row
        headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        col_index = {COLUMN_MAP[h]: idx for idx, h in enumerate(headers) if h in COLUMN_MAP}

        required_cols = {'email', 'expense_date', 'requested_amount', 'description'}
        missing_cols = required_cols - set(col_index.keys())
        if missing_cols:
            self.stderr.write(self.style.ERROR(f'Missing required columns: {missing_cols}'))
            return

        created = skipped = errors = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def get(col_name):
                idx = col_index.get(col_name)
                return row[idx] if idx is not None and idx < len(row) else None

            email = str(get('email') or '').strip().lower()
            expense_date = _parse_date(get('expense_date'))
            requested_amount = _parse_decimal(get('requested_amount'))
            description = str(get('description') or '').strip()
            staff_full_name = str(get('staff_full_name') or '').strip()

            # Skip blank rows
            if not any([email, expense_date, requested_amount]):
                continue

            # Validate mandatory fields
            if not email or not expense_date or not requested_amount or not description:
                self.stdout.write(self.style.WARNING(
                    f"Row {row_num}: Missing mandatory field(s) — skipped. "
                    f"email={email!r} date={expense_date} amount={requested_amount}"
                ))
                skipped += 1
                continue

            # Look up Staff
            try:
                staff_obj = Staff.objects.get(email__iexact=email)
            except Staff.DoesNotExist:
                self.stdout.write(self.style.WARNING(
                    f"Row {row_num}: No staff found for email {email!r} — skipped."
                ))
                skipped += 1
                continue
            except Staff.MultipleObjectsReturned:
                staff_obj = Staff.objects.filter(email__iexact=email).first()

            # Duplicate check
            if ExpenseRefund.objects.filter(
                staff=staff_obj,
                expense_date=expense_date,
                requested_amount=requested_amount,
                is_deleted=False,
            ).exists():
                self.stdout.write(
                    f"Row {row_num}: Duplicate detected for {email} on {expense_date} — skipped."
                )
                skipped += 1
                continue

            # Optional fields
            approved_amount = _parse_decimal(get('approved_amount'))
            volunteer_comment = str(get('volunteer_comment') or '').strip() or None
            admin_comment = str(get('admin_comment') or '').strip() or None
            status_val = str(get('status') or 'ממתין').strip()
            if status_val not in VALID_STATUSES:
                status_val = 'ממתין'
            approved_by = str(get('approved_by') or '').strip() or None
            if approved_by and approved_by not in VALID_APPROVED_BY:
                approved_by = None
            refund_method = str(get('refund_method') or '').strip() or None
            if refund_method and refund_method not in VALID_METHODS:
                refund_method = None
            phone_number = str(get('phone_number') or '').strip() or None

            if not staff_full_name:
                staff_full_name = f"{staff_obj.first_name} {staff_obj.last_name}".strip()

            if dry_run:
                self.stdout.write(
                    f"[DRY RUN] Row {row_num}: Would create refund for {email} | "
                    f"date={expense_date} amount={requested_amount} status={status_val}"
                )
                created += 1
                continue

            try:
                with transaction.atomic():
                    ExpenseRefund.objects.create(
                        staff=staff_obj,
                        staff_full_name=staff_full_name,
                        expense_date=expense_date,
                        requested_amount=requested_amount,
                        approved_amount=approved_amount,
                        description=description,
                        volunteer_comment=volunteer_comment,
                        admin_comment=admin_comment,
                        approved_by=approved_by,
                        file_url=None,  # Historical import — no receipt files
                        status=status_val,
                        refund_method=refund_method,
                        phone_number=phone_number,
                    )
                created += 1
            except Exception as e:
                self.stderr.write(self.style.ERROR(f"Row {row_num}: DB error — {e}"))
                errors += 1

        wb.close()

        self.stdout.write(f"\n{'[DRY RUN] ' if dry_run else ''}Import complete:")
        self.stdout.write(f"  Created : {created}")
        self.stdout.write(f"  Skipped : {skipped}")
        self.stdout.write(f"  Errors  : {errors}")

        if errors == 0:
            self.stdout.write(self.style.SUCCESS('✅ Import finished successfully.'))
        else:
            self.stdout.write(self.style.WARNING(f'⚠️ Import finished with {errors} error(s).'))
