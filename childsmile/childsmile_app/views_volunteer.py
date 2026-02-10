"""
Volunteer and Tutor registration views - Handle TOTP-based registration
"""
import os
import base64
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from django.http import JsonResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django_ratelimit.decorators import ratelimit
from django.core.mail import send_mail
from django.conf import settings
from django.db import DatabaseError, transaction, IntegrityError
from django.utils.timezone import now
from django.core.files.uploadedfile import UploadedFile
from .models import (
    Staff, TOTPCode, SignedUp, Pending_Tutor, 
    General_Volunteer, Tutors, Role, Children,
    Tutorships, PrevTutorshipStatuses, Tasks, Task_Types,
    MaritalStatus, Tutor_Feedback, General_V_Feedback, PossibleMatches
)
from .utils import *
from .audit_utils import log_api_action
from .logger import api_logger
import json
import datetime
import traceback
import re
import os
import base64
import pandas as pd



@conditional_csrf
@api_view(["POST"])
@ratelimit(key='ip', rate='5/m', method='POST', block=True)
def register_send_totp(request):
    """
    Send TOTP code for user registration verification
    """
    api_logger.info("register_send_totp called")
    try:
        data = request.data
        email = data.get("email", "").strip().lower()
        first_name = data.get("first_name", "").strip()
        surname = data.get("surname", data.get("last_name", "")).strip()
        birth_date = data.get("birth_date", "").strip()  # Birth date in dd/mm/yyyy format
        age = data.get("age", 0)  # Age calculated from birth_date on frontend
        gender = data.get("gender", "").strip()
        phone_prefix = data.get("phone_prefix", "").strip()
        phone_suffix = data.get("phone_suffix", "").strip()
        city = data.get("city", "").strip()
        comment = data.get("comment", "").strip()
        want_tutor = data.get("want_tutor", False)
        user_id = data.get("id", "").strip()  # Israeli ID

        # Check which fields are missing
        missing_fields = []
        if not email:
            missing_fields.append("email")
        if not first_name:
            missing_fields.append("first_name")
        if not surname:
            missing_fields.append("surname")
        if not birth_date:
            missing_fields.append("birth_date")
        if not gender:
            missing_fields.append("gender")
        if not phone_prefix or not phone_suffix:
            missing_fields.append("phone")
        if not city:
            missing_fields.append("city")
        if not user_id:
            missing_fields.append("id")

        # Validate required fields
        if missing_fields:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message=f"Missing required fields: {', '.join(missing_fields)}",
                status_code=400,
                additional_data={'attempted_email': email, 'missing_fields': missing_fields}
            )
            return JsonResponse({
                "error": f"Missing required fields: {', '.join(missing_fields)}"
            }, status=400)

        # Check if user already registered
        if SignedUp.objects.filter(email=email).exists() or Staff.objects.filter(email=email).exists() or Tutors.objects.filter(tutor_email=email).exists():
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message=f"Email '{email}' is already registered",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": f"This email is already registered"
            }, status=400)

        # Format phone with dash
        phone = f"{phone_prefix}-{phone_suffix}"
        
        # Check if phone already exists (normalize for comparison)
        import re
        phone_normalized = re.sub(r'[-\s]', '', phone)
        existing_phones = SignedUp.objects.values_list('phone', flat=True)
        for existing_phone in existing_phones:
            existing_normalized = re.sub(r'[-\s]', '', str(existing_phone or ''))
            if existing_normalized == phone_normalized:
                log_api_action(
                    request=request,
                    action='USER_REGISTRATION_FAILED',
                    success=False,
                    error_message=f"Phone number is already registered",
                    status_code=400,
                    additional_data={'attempted_email': email, 'attempted_phone': phone}
                )
                return JsonResponse({
                    "error": "This phone number is already registered"
                }, status=400)

        # Store registration data in session
        request.session['pending_registration'] = {
            "id": user_id,
            "email": email,
            "first_name": first_name,
            "surname": surname,
            "birth_date": birth_date,  # Store birth_date in dd/mm/yyyy format
            "age": age,  # Age calculated from birth_date on frontend
            "gender": gender,
            "phone": phone,
            "city": city,
            "comment": comment,
            "want_tutor": want_tutor
        }
        request.session['registration_new_user_email'] = email

        # Mark old codes as used
        TOTPCode.objects.filter(email=email, used=False).update(used=True)
        
        # Generate new code
        code = TOTPCode.generate_code()
        TOTPCode.objects.create(email=email, code=code)

        # Send verification email
        subject = "אימות הרשמה - חיוך של ילד"
        message = f"""
        שלום {first_name},
        
        תודה שנרשמת ל-חיוך של ילד!
        
        קוד האימות שלך הוא: {code}
        
        הקוד יפוג בעוד 5 דקות.
        
        אנא הזן קוד זה כדי להשלים את ההרשמה.
        
        בברכה,
        צוות חיוך של ילד
        """
        
        try:
            send_mail(
                subject,
                message,
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )
            
            return JsonResponse({
                "message": "Verification code sent to your email",
                "email": email
            })
            
        except Exception as email_error:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message=f"Failed to send verification email: {str(email_error)}",
                status_code=500,
                additional_data={
                    'attempted_email': email,
                    'attempted_first_name': first_name,
                    'attempted_surname': surname
                }
            )
            return JsonResponse({
                "error": f"Failed to send verification email"
            }, status=500)
        
    except Exception as e:
        api_logger.error(f"Error in register_send_totp: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='USER_REGISTRATION_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'attempted_email': data.get('email', 'Unknown')}
        )
        return JsonResponse({"error": "Registration failed"}, status=500)


@conditional_csrf
@api_view(["POST"])
@ratelimit(key='ip', rate='10/m', method='POST', block=True)
def register_verify_totp(request):
    api_logger.info("register_verify_totp called")
    """
    Verify TOTP and complete user registration
    """
    try:
        api_logger.debug(f"register_verify_totp called")
        
        data = json.loads(request.body)
        code = data.get('code', '').strip()
        
        email = request.session.get('registration_new_user_email', '').strip().lower()
        api_logger.debug(f"Email from session: '{email}', code from request: '{code}'")
        
        if not email or not code:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message="Invalid session or missing code",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({"error": "Invalid session or missing code"}, status=400)
        
        # Find the TOTP record
        totp_record = TOTPCode.objects.filter(
            email=email,
            used=False
        ).order_by('-created_at').first()
        
        api_logger.debug(f"Found TOTP record: {totp_record}")
        
        if not totp_record:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message="Invalid or expired code",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({"error": "Invalid or expired code"}, status=400)
            
        if not totp_record.is_valid():
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message="Code has expired or too many attempts",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({"error": "Code has expired or too many attempts"}, status=400)
        
        totp_record.attempts += 1
        totp_record.save()
        
        if totp_record.code != code:
            if totp_record.attempts >= 3:
                totp_record.used = True
                totp_record.save()
                log_api_action(
                    request=request,
                    action='USER_REGISTRATION_FAILED',
                    success=False,
                    error_message="Too many failed attempts",
                    status_code=429,
                    additional_data={'attempted_email': email}
                )
                return JsonResponse({"error": "Too many failed attempts"}, status=429)
            
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message="Invalid code",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({"error": "Invalid code"}, status=400)
        
        totp_record.used = True
        totp_record.save()
        api_logger.debug(f"TOTP verification successful")
        
        registration_data = request.session.get('pending_registration')
        api_logger.debug(f"Registration data from session: {registration_data}")
        
        if not registration_data:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_FAILED',
                success=False,
                error_message="Registration session expired",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({"error": "Registration session expired"}, status=400)
        
        api_logger.debug(f"About to create volunteer/tutor")
        result = create_volunteer_or_tutor_internal(registration_data, request)
        
        if 'pending_registration' in request.session:
            del request.session['pending_registration']
        if 'registration_new_user_email' in request.session:
            del request.session['registration_new_user_email']
        
        api_logger.debug(f"Registration completed successfully")
        return result
        
    except Exception as e:
        api_logger.error(f"Error in register_verify_totp: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='USER_REGISTRATION_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'attempted_email': email if 'email' in locals() else 'Unknown'}
        )
        return JsonResponse({"error": "Registration failed"}, status=500)


@conditional_csrf
@api_view(["POST"])
def create_volunteer_or_tutor(request):
    api_logger.info("create_volunteer_or_tutor called")
    """
    Create a new volunteer or tutor user via direct API call (not via registration flow)
    """
    from .utils import parse_date_string, calculate_age_from_birth_date
    
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="Authentication credentials were not provided",
                status_code=403,
                additional_data={'attempted_email': request.data.get('email', 'Unknown')}
            )
            return JsonResponse({
                "detail": "Authentication credentials were not provided."
            }, status=403)

        user = Staff.objects.get(staff_id=user_id)
        
        data = request.data
        email = data.get("email", "").strip().lower()
        first_name = data.get("first_name", "").strip()
        surname = data.get("surname", data.get("last_name", "")).strip()
        
        # Parse birth_date and calculate age from it
        birth_date_str = data.get("birth_date", "").strip()
        birth_date = parse_date_string(birth_date_str) if birth_date_str else None
        
        # Use calculated age from birth_date, or fallback to provided age
        if birth_date:
            age = calculate_age_from_birth_date(birth_date)
        else:
            age = int(data.get("age", 0)) if data.get("age") else 0
        
        gender = data.get("gender", "")
        phone = data.get("phone", "").strip()
        city = data.get("city", "").strip()
        comment = data.get("comment", "").strip()
        want_tutor = data.get("want_tutor", False)

        if not email or not first_name or not surname:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="Missing required fields: email, first_name, surname",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": "Missing required fields: email, first_name, surname"
            }, status=400)

        if SignedUp.objects.filter(email=email).exists():
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message=f"Email '{email}' is already registered",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": f"Email '{email}' is already registered"
            }, status=400)

        # Create SignedUp record
        signup_user = SignedUp.objects.create(
            email=email,
            first_name=first_name,
            surname=surname,
            birth_date=birth_date,  # Store birth_date
            age=age,  # Store calculated age
            gender=gender,
            phone=phone,
            city=city,
            comment=comment,
            want_tutor=want_tutor
        )

        # Create Staff record
        staff_user = Staff.objects.create(
            username=email.split('@')[0],
            email=email,
            first_name=first_name,
            last_name=surname
        )

        # Create role-specific record
        if want_tutor:
            pending_tutor = Pending_Tutor.objects.create(
                id=signup_user
            )
            tutor_role = Role.objects.get(role_name="Tutor")
            staff_user.roles.add(tutor_role)
        else:
            general_volunteer = General_Volunteer.objects.create(
                id=signup_user,
                staff=staff_user,
                notes="Created by admin"
            )
            volunteer_role = Role.objects.get(role_name="General Volunteer")
            staff_user.roles.add(volunteer_role)

        log_api_action(
            request=request,
            action='CREATE_VOLUNTEER_SUCCESS',
            affected_tables=['childsmile_app_signedup', 'childsmile_app_staff'],
            entity_type='User',
            entity_ids=[signup_user.id],
            success=True,
            additional_data={
                'created_user_email': email,
                'user_type': "Tutor" if want_tutor else "General Volunteer",
                'created_by_admin': True,
                'first_name': first_name,
                'surname': surname,
                'staff_id': staff_user.staff_id
            }
        )

        return JsonResponse({
            "message": "User created successfully",
            "user_id": signup_user.id,
            "email": email,
            "want_tutor": want_tutor
        }, status=201)

    except Staff.DoesNotExist:
        log_api_action(
            request=request,
            action='CREATE_VOLUNTEER_FAILED',
            success=False,
            error_message="Authentication failed - staff not found",
            status_code=403,
            additional_data={'attempted_email': data.get('email', 'Unknown')}
        )
        return JsonResponse({
            "error": "Authentication failed"
        }, status=403)
    except Exception as e:
        api_logger.error(f"Error in create_volunteer_or_tutor: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='CREATE_VOLUNTEER_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'attempted_email': data.get('email', 'Unknown')}
        )
        return JsonResponse({"error": str(e)}, status=500)

@transaction.atomic
def create_volunteer_or_tutor_internal(data, request=None):
    api_logger.debug("create_volunteer_or_tutor_internal called")
    """
    Internal function to create volunteer/tutor - ORIGINAL WORKING VERSION with audit improvements
    """
    from .utils import parse_date_string, calculate_age_from_birth_date
    
    try:
        # Extract ID from data (Israeli ID from registration)
        user_id = data.get("id")
        first_name = data.get("first_name")
        surname = data.get("surname")
        
        # Parse birth_date and calculate age from it
        birth_date_str = data.get("birth_date")
        birth_date = parse_date_string(birth_date_str) if birth_date_str else None
        
        # Use calculated age from birth_date, or fallback to provided age
        if birth_date:
            age = calculate_age_from_birth_date(birth_date)
        else:
            age = int(data.get("age", 0))
        
        gender = data.get("gender") == "Female"
        phone = data.get("phone")  # Already formatted in register_send_totp
        city = data.get("city")
        comment = data.get("comment", "")
        email = data.get("email")
        want_tutor = data.get("want_tutor") == "true" or data.get("want_tutor") is True

        api_logger.debug(f"Extracted user_id={user_id}, email={email}, want_tutor={want_tutor}, birth_date={birth_date}, age={age}")

        # Create username
        username = f"{first_name}_{surname}"
        index = 1
        original_username = username
        while Staff.objects.filter(username=username).exists():
            username = f"{original_username}_{index}"
            index += 1

        # Insert into SignedUp table - USE the Israeli ID
        signedup = SignedUp.objects.create(
            id=int(user_id),
            first_name=first_name,
            surname=surname,
            birth_date=birth_date,  # Store birth_date
            age=age,  # Store calculated age
            gender=gender,
            phone=phone,
            city=city,
            comment=comment,
            email=email,
            want_tutor=want_tutor,
        )
        api_logger.debug(f"Created SignedUp with id={signedup.id}, birth_date={birth_date}")

        # Get role
        role_name = "General Volunteer"
        role = Role.objects.get(role_name=role_name)

        # Create Staff with registration_approved=False (requires admin approval)
        staff = Staff.objects.create(
            username=username,
            email=email,
            first_name=first_name,
            last_name=surname,
            created_at=now(),
            registration_approved=False  # NEW: Set to False, requires admin approval
        )
        staff.roles.add(role)
        staff.refresh_from_db()
        
        # DEBUG: Verify role was added
        staff_roles = list(staff.roles.all().values_list('role_name', flat=True))
        api_logger.debug(f"Created staff {staff.staff_id} with roles: {staff_roles}")
        
        # CREATE REGISTRATION APPROVAL TASK FOR ADMINS
        full_name = f"{first_name} {surname}"
        create_tasks_for_admins_async(staff.staff_id, full_name, email)
        api_logger.info(f"Created registration approval tasks for admins for user {staff.staff_id} ({email})")

        # Create volunteer or pending tutor
        if want_tutor:
            pending_tutor = Pending_Tutor.objects.create(
                id_id=signedup.id,
                pending_status="ממתין",  # "Pending" in Hebrew
            )
            
            # NOTE: Interview task for tutor coordinators will be created ONLY after admin approves registration
            # This prevents tasks for users that don't actually exist yet
            
            # Log pending tutor creation with proper user info
            if request:
                log_api_action(
                    request=request,
                    action='CREATE_PENDING_TUTOR_SUCCESS',
                    affected_tables=['childsmile_app_pending_tutor', 'childsmile_app_signedup', 'childsmile_app_staff'],
                    entity_type='Pending_Tutor',
                    entity_ids=[pending_tutor.pending_tutor_id],
                    success=True,
                    additional_data={
                        'tutor_email': email,
                        'first_name': first_name,
                        'surname': surname,
                        'username': username,
                        'staff_id': staff.staff_id,
                        'pending_status': 'ממתין',
                        'pending_tutor_id': pending_tutor.pending_tutor_id
                    }
                )
        else:
            General_Volunteer.objects.create(
                id_id=signedup.id,
                staff_id=staff.staff_id,
                signupdate=now().date(),
                comments="",
            )
            
            # Log volunteer creation with proper user info
            if request:
                log_api_action(
                    request=request,
                    action='CREATE_VOLUNTEER_SUCCESS',
                    affected_tables=['childsmile_app_general_volunteer', 'childsmile_app_signedup', 'childsmile_app_staff'],
                    entity_type='General_Volunteer',
                    entity_ids=[signedup.id],
                    success=True,
                    additional_data={
                        'volunteer_email': email,
                        'first_name': first_name,
                        'surname': surname,
                        'username': username,
                        'staff_id': staff.staff_id,
                        'signup_date': now().date().isoformat()
                    }
                )

        # Log successful registration with proper user info
        if request:
            log_api_action(
                request=request,
                action='USER_REGISTRATION_SUCCESS',
                affected_tables=['childsmile_app_signedup', 'childsmile_app_staff'],
                entity_type='SignedUp',
                entity_ids=[signedup.id],
                success=True,
                additional_data={
                    'registration_type': 'pending_tutor' if want_tutor else 'volunteer',
                    'email': email,
                    'username': username,
                    'staff_id': staff.staff_id,
                    'first_name': first_name,
                    'surname': surname,
                    'israeli_id': user_id
                }
            )

        return JsonResponse({
            "message": "Registration completed successfully!",
            "username": username,
        }, status=201)

    except Exception as e:
        api_logger.error(f"Error in create_volunteer_or_tutor_internal: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        
        if request:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED' if not data.get("want_tutor") else 'CREATE_PENDING_TUTOR_FAILED',
                success=False,
                error_message=str(e),
                status_code=500,
                additional_data={
                    'attempted_email': data.get('email', 'unknown'),
                    'attempted_username': f"{data.get('first_name', '')}_{data.get('surname', '')}",
                    'attempted_id': data.get('id', 'unknown')
                }
            )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["POST"])
def create_pending_tutor(request):
    api_logger.info("create_pending_tutor called")
    """
    Create a pending tutor record for an existing user
    """
    try:
        user_id = request.session.get("user_id")
        if not user_id:
            log_api_action(
                request=request,
                action='CREATE_PENDING_TUTOR_FAILED',
                success=False,
                error_message="Authentication credentials were not provided",
                status_code=403,
                additional_data={'attempted_email': request.data.get('email', 'Unknown')}
            )
            return JsonResponse({
                "detail": "Authentication credentials were not provided."
            }, status=403)

        user = Staff.objects.get(staff_id=user_id)
        
        data = request.data
        email = data.get("email", "").strip().lower()

        if not email:
            log_api_action(
                request=request,
                action='CREATE_PENDING_TUTOR_FAILED',
                success=False,
                error_message="Missing required field: email",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": "Missing required field: email"
            }, status=400)

        # Find the user
        try:
            signup_user = SignedUp.objects.get(email=email)
        except SignedUp.DoesNotExist:
            log_api_action(
                request=request,
                action='CREATE_PENDING_TUTOR_FAILED',
                success=False,
                error_message=f"User with email '{email}' not found",
                status_code=404,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": f"User with email '{email}' not found"
            }, status=404)

        # Check if already a pending tutor
        if Pending_Tutor.objects.filter(id=signup_user).exists():
            log_api_action(
                request=request,
                action='CREATE_PENDING_TUTOR_FAILED',
                success=False,
                error_message=f"User is already a pending tutor",
                status_code=400,
                additional_data={'attempted_email': email}
            )
            return JsonResponse({
                "error": "User is already a pending tutor"
            }, status=400)

        # Create pending tutor record
        pending_tutor = Pending_Tutor.objects.create(
            id=signup_user,
            created_at=datetime.datetime.n
        )

        log_api_action(
            request=request,
            action='CREATE_PENDING_TUTOR_SUCCESS',
            affected_tables=['childsmile_app_pending_tutor'],
            entity_type='PendingTutor',
            entity_ids=[pending_tutor.id],
            success=True,
            additional_data={
                'user_email': email,
                'user_name': f"{signup_user.first_name} {signup_user.last_name}"
            }
        )

        return JsonResponse({
            "message": "Pending tutor record created successfully",
            "user_id": signup_user.id,
            "email": email
        }, status=201)

    except Staff.DoesNotExist:
        log_api_action(
            request=request,
            action='CREATE_PENDING_TUTOR_FAILED',
            success=False,
            error_message="Authentication failed - staff not found",
            status_code=403,
            additional_data={'attempted_email': data.get('email', 'Unknown')}
        )
        return JsonResponse({
            "error": "Authentication failed"
        }, status=403)
    except Exception as e:
        api_logger.error(f"Error in create_pending_tutor: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='CREATE_PENDING_TUTOR_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'attempted_email': data.get('email', 'Unknown')}
        )
        return JsonResponse({"error": str(e)}, status=500)


@conditional_csrf
@api_view(["PUT"])
def update_general_volunteer(request, volunteer_id):
    api_logger.info(f"update_general_volunteer called for volunteer_id: {volunteer_id}")
    """
    Update a general volunteer's information
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_GENERAL_VOLUNTEER_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse({"detail": "Authentication credentials were not provided."}, status=403)
    if not has_permission(request, "general_volunteer", "UPDATE"):
        log_api_action(
            request=request,
            action='UPDATE_GENERAL_VOLUNTEER_FAILED',
            success=False,
            error_message="You do not have permission to access this resource",
            status_code=401,
            entity_type='General_Volunteer',
            entity_ids=[volunteer_id]
        )
        return JsonResponse({"error": "You do not have permission to access this resource."}, status=401)

    try:
        volunteer = General_Volunteer.objects.get(id_id=volunteer_id)
    except General_Volunteer.DoesNotExist:
        log_api_action(
            request=request,
            action='UPDATE_GENERAL_VOLUNTEER_FAILED',
            success=False,
            error_message="General Volunteer not found",
            status_code=404,
            entity_type='General_Volunteer',
            entity_ids=[volunteer_id],
            additional_data={'attempted_email': 'Unknown'}
        )
        return JsonResponse({"error": "General Volunteer not found."}, status=404)

    data = request.data
    old_comments = volunteer.comments
    volunteer.comments = data.get("comments", volunteer.comments)
    volunteer.save()

    # Handle city update
    if "city" in data:
        new_city = data["city"]
        # Update city in SignedUp table
        try:
            signedup = SignedUp.objects.get(id=volunteer_id)
            if new_city != signedup.city:
                signedup.city = new_city
                signedup.save()
        except SignedUp.DoesNotExist:
            pass

    # Handle birth_date update - also recalculate age
    if "birth_date" in data:
        from .utils import parse_date_string, calculate_age_from_birth_date
        new_birth_date_str = data["birth_date"]
        new_birth_date = parse_date_string(new_birth_date_str) if new_birth_date_str else None
        
        try:
            signedup = SignedUp.objects.get(id=volunteer_id)
            if new_birth_date != signedup.birth_date:
                signedup.birth_date = new_birth_date
                # Recalculate age from new birth_date
                if new_birth_date:
                    signedup.age = calculate_age_from_birth_date(new_birth_date)
                signedup.save()
        except SignedUp.DoesNotExist:
            pass

    log_api_action(
        request=request,
        action='UPDATE_GENERAL_VOLUNTEER_SUCCESS',
        affected_tables=['childsmile_app_general_volunteer'],
        entity_type='General_Volunteer',
        entity_ids=[volunteer_id],
        success=True,
        additional_data={
            'old_comments': old_comments,
            'new_comments': volunteer.comments,
            'volunteer_name': f"{volunteer.id.first_name} {volunteer.id.surname}",
            'volunteer_email': volunteer.id.email
        }
    )

    return JsonResponse({"message": "General Volunteer updated successfully."}, status=200)


@conditional_csrf
@api_view(["PUT"])
def update_volunteer_id(request, old_id):
    """
    Update a volunteer/tutor Israeli ID (ת.ז) across all related tables.
    This performs a cascading update of the ID in SignedUp and all FK references.
    """
    api_logger.info(f"update_volunteer_id called for old_id: {old_id}")
    
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403,
            entity_type='SignedUp',
            entity_ids=[old_id]
        )
        return JsonResponse({"detail": "Authentication credentials were not provided."}, status=403)
    
    # Check permission for both tutors and general_volunteer
    if not (has_permission(request, "tutors", "UPDATE") or has_permission(request, "general_volunteer", "UPDATE")):
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="You do not have permission to update IDs",
            status_code=401,
            entity_type='SignedUp',
            entity_ids=[old_id]
        )
        return JsonResponse({"error": "You do not have permission to update IDs."}, status=401)

    data = request.data
    new_id = data.get("new_id")
    
    if not new_id:
        return JsonResponse({"error": "New ID is required."}, status=400)
    
    # Validate new ID format (9 digits for Israeli ID)
    new_id_str = str(new_id).strip()
    if not new_id_str.isdigit() or len(new_id_str) != 9:
        return JsonResponse({"error": "Invalid ID format. Israeli ID must be exactly 9 digits."}, status=400)
    
    new_id = int(new_id_str)
    
    # Check if new ID already exists
    if SignedUp.objects.filter(id=new_id).exists():
        return JsonResponse({"error": "This ID already exists in the system."}, status=400)
    
    try:
        signedup = SignedUp.objects.get(id=old_id)
    except SignedUp.DoesNotExist:
        return JsonResponse({"error": "Volunteer/Tutor not found."}, status=404)
    
    try:
        with transaction.atomic():
            # Store original data for new record
            original_data = {
                'first_name': signedup.first_name,
                'surname': signedup.surname,
                'birth_date': signedup.birth_date,  # Include birth_date
                'age': signedup.age,
                'gender': signedup.gender,
                'phone': signedup.phone,
                'city': signedup.city,
                'comment': signedup.comment,
                'email': signedup.email,
                'want_tutor': signedup.want_tutor,
            }
            
            affected_tables = ['childsmile_app_signedup']
            
            # Create new SignedUp record with new ID
            new_signedup = SignedUp.objects.create(
                id=new_id,
                **original_data
            )
            
            # IMPORTANT: Update tables in correct order - child tables before parent tables
            # to avoid FK constraint violations
            
            # 1. Update PrevTutorshipStatuses FK (references Tutors) - MUST be before Tutors update
            prev_status_updated = PrevTutorshipStatuses.objects.filter(tutor_id_id=old_id).update(tutor_id_id=new_id)
            if prev_status_updated:
                affected_tables.append('childsmile_app_prevtutorshipstatuses')
            
            # 2. Update Tutorships FK (references Tutors) - MUST be before Tutors update
            tutorship_updated = Tutorships.objects.filter(tutor_id=old_id).update(tutor_id=new_id)
            if tutorship_updated:
                affected_tables.append('childsmile_app_tutorships')
            
            # 3. Update Tasks related_tutor FK (references Tutors) - MUST be before Tutors update
            tasks_updated = Tasks.objects.filter(related_tutor_id=old_id).update(related_tutor_id=new_id)
            if tasks_updated:
                affected_tables.append('childsmile_app_tasks')
            
            # 4. Update Tutor_Feedback FK (references Tutors) - MUST be before Tutors update
            tutor_feedback_updated = Tutor_Feedback.objects.filter(tutor_id=old_id).update(tutor_id=new_id)
            if tutor_feedback_updated:
                affected_tables.append('childsmile_app_tutor_feedback')
            
            # 5. Update General_V_Feedback FK (references General_Volunteer) - MUST be before General_Volunteer update
            gv_feedback_updated = General_V_Feedback.objects.filter(volunteer_id=old_id).update(volunteer_id=new_id)
            if gv_feedback_updated:
                affected_tables.append('childsmile_app_general_v_feedback')
            
            # 6. Update PossibleMatches (stores tutor_id as IntegerField, not FK)
            possible_matches_updated = PossibleMatches.objects.filter(tutor_id=old_id).update(tutor_id=new_id)
            if possible_matches_updated:
                affected_tables.append('childsmile_app_possiblematches')
            
            # 7. Now update Tutors FK (references SignedUp) - after all child tables
            tutor_updated = Tutors.objects.filter(id_id=old_id).update(id_id=new_id)
            if tutor_updated:
                affected_tables.append('childsmile_app_tutors')
            
            # 8. Update General_Volunteer FK (references SignedUp) - after all child tables
            gv_updated = General_Volunteer.objects.filter(id_id=old_id).update(id_id=new_id)
            if gv_updated:
                affected_tables.append('childsmile_app_general_volunteer')
            
            # 9. Update Pending_Tutor FK (references SignedUp)
            pending_updated = Pending_Tutor.objects.filter(id_id=old_id).update(id_id=new_id)
            if pending_updated:
                affected_tables.append('childsmile_app_pending_tutor')
            
            # 10. Finally delete old SignedUp record
            signedup.delete()
            
            # Determine if this was a tutor or volunteer for proper action name
            action_name = 'UPDATE_TUTOR_SUCCESS' if tutor_updated else 'UPDATE_GENERAL_VOLUNTEER_SUCCESS'
            
            log_api_action(
                request=request,
                action=action_name,
                affected_tables=affected_tables,
                entity_type='SignedUp',
                entity_ids=[old_id, new_id],
                success=True,
                additional_data={
                    'update_type': 'ID_CHANGE',
                    'old_id': old_id,
                    'new_id': new_id,
                    'volunteer_name': f"{original_data['first_name']} {original_data['surname']}",
                    'volunteer_email': original_data['email']
                }
            )
            
            return JsonResponse({
                "message": "ID updated successfully across all related tables.",
                "old_id": old_id,
                "new_id": new_id,
                "affected_tables": affected_tables
            }, status=200)
            
    except Exception as e:
        api_logger.error(f"Error updating volunteer ID: {str(e)}\n{traceback.format_exc()}")
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            entity_type='SignedUp',
            entity_ids=[old_id],
            additional_data={'attempted_new_id': new_id}
        )
        return JsonResponse({"error": f"Error updating ID: {str(e)}"}, status=500)


@conditional_csrf
@api_view(["PUT"])
def update_volunteer_phone(request, volunteer_id):
    """
    Update a volunteer/tutor phone number in the SignedUp table.
    Validates phone format (10 digits) and uniqueness.
    """
    api_logger.info(f"update_volunteer_phone called for volunteer_id: {volunteer_id}")
    
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403,
            entity_type='SignedUp',
            entity_ids=[volunteer_id]
        )
        return JsonResponse({"detail": "Authentication credentials were not provided."}, status=403)
    
    # Check permission for both tutors and general_volunteer
    if not (has_permission(request, "tutors", "UPDATE") or has_permission(request, "general_volunteer", "UPDATE")):
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="You do not have permission to update phone numbers",
            status_code=401,
            entity_type='SignedUp',
            entity_ids=[volunteer_id]
        )
        return JsonResponse({"error": "You do not have permission to update phone numbers."}, status=401)

    data = request.data
    new_phone = data.get("phone")
    
    if not new_phone:
        return JsonResponse({"error": "Phone number is required."}, status=400)
    
    # Normalize phone - remove dashes and spaces for validation
    import re
    new_phone_normalized = re.sub(r'[-\s]', '', str(new_phone).strip())
    
    # Validate phone format (10 digits, starts with 0)
    if not new_phone_normalized.isdigit() or len(new_phone_normalized) != 10:
        return JsonResponse({"error": "Invalid phone format. Phone must be exactly 10 digits."}, status=400)
    
    if not new_phone_normalized.startswith('0'):
        return JsonResponse({"error": "Invalid phone format. Phone must start with 0."}, status=400)
    
    # Format phone with dash for storage (XXX-XXXXXXX)
    new_phone_formatted = f"{new_phone_normalized[:3]}-{new_phone_normalized[3:]}"
    
    try:
        signedup = SignedUp.objects.get(id=volunteer_id)
    except SignedUp.DoesNotExist:
        return JsonResponse({"error": "Volunteer/Tutor not found."}, status=404)
    
    # Check if phone already exists (excluding current volunteer)
    # Normalize all existing phones for comparison
    existing_phones = SignedUp.objects.exclude(id=volunteer_id).values_list('phone', flat=True)
    for existing_phone in existing_phones:
        existing_normalized = re.sub(r'[-\s]', '', str(existing_phone or ''))
        if existing_normalized == new_phone_normalized:
            return JsonResponse({"error": "This phone number already exists in the system."}, status=400)
    
    old_phone = signedup.phone
    signedup.phone = new_phone_formatted
    signedup.save()
    
    # Also update the related volunteer/tutor updated timestamp
    is_tutor = Tutors.objects.filter(id_id=volunteer_id).exists()
    General_Volunteer.objects.filter(id_id=volunteer_id).update(updated=now())
    Tutors.objects.filter(id_id=volunteer_id).update(updated=now())
    
    # Log with appropriate action name based on user type
    action_name = 'UPDATE_TUTOR_SUCCESS' if is_tutor else 'UPDATE_GENERAL_VOLUNTEER_SUCCESS'
    log_api_action(
        request=request,
        action=action_name,
        affected_tables=['childsmile_app_signedup'],
        entity_type='SignedUp',
        entity_ids=[volunteer_id],
        success=True,
        additional_data={
            'update_type': 'PHONE_CHANGE',
            'old_phone': old_phone,
            'new_phone': new_phone_formatted,
            'volunteer_name': f"{signedup.first_name} {signedup.surname}",
            'volunteer_email': signedup.email
        }
    )
    
    return JsonResponse({
        "message": "Phone number updated successfully.",
        "old_phone": old_phone,
        "new_phone": new_phone_formatted
    }, status=200)


@conditional_csrf
@api_view(["PUT"])
def update_tutor(request, tutor_id):
    api_logger.info(f"update_tutor called for tutor_id: {tutor_id}")
    """
    Update a tutor's information including email, status, and wellness data
    """
    user_id = request.session.get("user_id")
    if not user_id:
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="Authentication credentials were not provided",
            status_code=403
        )
        return JsonResponse({"detail": "Authentication credentials were not provided."}, status=403)
    if not has_permission(request, "tutors", "UPDATE"):
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="You do not have permission to access this resource",
            status_code=401,
            entity_type='Tutor',
            entity_ids=[tutor_id]
        )
        return JsonResponse({"error": "You do not have permission to access this resource."}, status=401)

    try:
        tutor = Tutors.objects.get(id_id=tutor_id)
    except Tutors.DoesNotExist:
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_FAILED',
            success=False,
            error_message="Tutor not found",
            status_code=404,
            entity_type='Tutor',
            entity_ids=[tutor_id],
            additional_data={
                'tutor_name': 'Unknown',
                'tutor_email': 'Unknown',
                'attempted_fields': list(request.data.keys()) if hasattr(request, 'data') else []
            }
        )
        return JsonResponse({"error": "Tutor not found."}, status=404)

    data = request.data
    updated = False
    
    # Store original values for audit
    original_data = {
        'tutor_email': tutor.tutor_email,
        'relationship_status': tutor.relationship_status,
        'tutee_wellness': tutor.tutee_wellness,
        'tutorship_status': tutor.tutorship_status,
        'preferences': tutor.preferences
    }
    
    affected_tables = ['childsmile_app_tutors']

    # Update tutor_email in both Tutors and Staff tables
    if "tutor_email" in data:
        tutor.tutor_email = data["tutor_email"]
        try:
            staff = Staff.objects.get(staff_id=tutor.staff_id)
            staff.email = data["tutor_email"]
            staff.save()
            affected_tables.append('childsmile_app_staff')
        except Staff.DoesNotExist:
            pass
        updated = True

    # Try to locate existing tutorship (we'll need it later)
    tutorship = Tutorships.objects.filter(tutor_id=tutor_id).first()
    child = tutorship.child if tutorship else None

    # Only allow updating relationship_status and tutee_wellness if tutor is in tutorship
    if ("relationship_status" in data or "tutee_wellness" in data) and tutorship and child:
        if "relationship_status" in data:
            new_status = data["relationship_status"]
            # Validate relationship_status against MaritalStatus choices
            valid_statuses = [choice[0] for choice in MaritalStatus.choices]
            if new_status not in valid_statuses:
                log_api_action(
                    request=request,
                    action='UPDATE_TUTOR_FAILED',
                    success=False,
                    error_message=f"Invalid relationship_status: '{new_status}'. Must be one of: {', '.join(valid_statuses)}",
                    status_code=400,
                    entity_type='Tutor',
                    entity_ids=[tutor_id],
                    additional_data={
                        'attempted_value': new_status,
                        'valid_values': valid_statuses,
                        'tutor_name': f"{tutor.staff.username if tutor.staff else 'Unknown'}",
                        'tutor_email': tutor.tutor_email
                    }
                )
                return JsonResponse({
                    "error": f"Invalid relationship_status: '{new_status}'. Must be one of: {', '.join(valid_statuses)}"
                }, status=400)
            
            if new_status != tutor.relationship_status:
                tutor.relationship_status = new_status
                child.marital_status = new_status
                child.save()
                affected_tables.append('childsmile_app_children')
                updated = True

        if "tutee_wellness" in data:
            new_wellness = data["tutee_wellness"]
            if new_wellness != tutor.tutee_wellness:
                tutor.tutee_wellness = new_wellness
                child.current_medical_state = new_wellness
                child.save()
                if 'childsmile_app_children' not in affected_tables:
                    affected_tables.append('childsmile_app_children')
                updated = True

    # --- Tutorship status logic + PrevTutorshipStatuses ---
    if "tutorship_status" in data:
        new_tutor_status = data["tutorship_status"]
        if new_tutor_status != tutor.tutorship_status:
            tutor.tutorship_status = new_tutor_status
            updated = True

            # Find existing prev record
            prev = PrevTutorshipStatuses.objects.filter(tutor_id=tutor).order_by('-last_updated').first()

            if prev:
                prev.tutor_tut_status = new_tutor_status
                prev.save()
                affected_tables.append('childsmile_app_prevtutorshipstatuses')
            else:
                PrevTutorshipStatuses.objects.create(
                    tutor_id=tutor,
                    child_id=child if child else None,
                    tutor_tut_status=new_tutor_status,
                    child_tut_status=child.tutorship_status if child else "",
                )
                affected_tables.append('childsmile_app_prevtutorshipstatuses')

    # --- Child status logic (if provided) ---
    if "child_tut_status" in data and child:
        new_child_status = data["child_tut_status"]

        # assuming child has field tutorship_status or equivalent
        if getattr(child, "tutorship_status", "") != new_child_status:
            setattr(child, "tutorship_status", new_child_status)
            child.save()
            if 'childsmile_app_children' not in affected_tables:
                affected_tables.append('childsmile_app_children')
            updated = True

            # update PrevTutorshipStatuses
            prev = PrevTutorshipStatuses.objects.filter(child_id=child).order_by('-last_updated').first()

            if prev:
                prev.child_tut_status = new_child_status
                prev.save()
                if 'childsmile_app_prevtutorshipstatuses' not in affected_tables:
                    affected_tables.append('childsmile_app_prevtutorshipstatuses')
            else:
                PrevTutorshipStatuses.objects.create(
                    tutor_id=tutor,
                    child_id=child,
                    tutor_tut_status=tutor.tutorship_status,
                    child_tut_status=new_child_status,
                )
                if 'childsmile_app_prevtutorshipstatuses' not in affected_tables:
                    affected_tables.append('childsmile_app_prevtutorshipstatuses')

    if "preferences" in data:
        tutor.preferences = data["preferences"]
        updated = True

    # Handle city update
    if "city" in data:
        new_city = data["city"]
        # Update city in SignedUp table via the tutor's id relationship
        try:
            signedup = SignedUp.objects.get(id=tutor.id_id)
            if new_city != signedup.city:
                signedup.city = new_city
                signedup.save()
                affected_tables.append('childsmile_app_signedup')
                original_data['city'] = signedup.city
                updated = True
        except SignedUp.DoesNotExist:
            pass

    # Handle birth_date update - also recalculate age
    if "birth_date" in data:
        from .utils import parse_date_string, calculate_age_from_birth_date
        new_birth_date_str = data["birth_date"]
        new_birth_date = parse_date_string(new_birth_date_str) if new_birth_date_str else None
        
        try:
            signedup = SignedUp.objects.get(id=tutor.id_id)
            if new_birth_date != signedup.birth_date:
                signedup.birth_date = new_birth_date
                # Recalculate age from new birth_date
                if new_birth_date:
                    signedup.age = calculate_age_from_birth_date(new_birth_date)
                signedup.save()
                if 'childsmile_app_signedup' not in affected_tables:
                    affected_tables.append('childsmile_app_signedup')
                original_data['birth_date'] = signedup.birth_date
                updated = True
        except SignedUp.DoesNotExist:
            pass

    if updated:
        tutor.save()
        
        # Determine what changed for audit
        changed_fields = {}
        for field, original_value in original_data.items():
            if field in data:
                new_value = getattr(tutor, field, data[field])
                if original_value != new_value:
                    changed_fields[field] = {
                        'old': original_value,
                        'new': new_value
                    }

        log_api_action(
            request=request,
            action='UPDATE_TUTOR_SUCCESS',
            affected_tables=affected_tables,
            entity_type='Tutor',
            entity_ids=[tutor_id],
            success=True,
            additional_data={
                'changed_fields': changed_fields,
                'has_tutorship': tutorship is not None,
                'child_id': child.child_id if child else None,
                'child_name': f"{child.childfirstname} {child.childsurname}" if child else 'Unknown',
                'tutor_name': f"{tutor.staff.username if tutor.staff else 'Unknown'}",
                'tutor_email': tutor.tutor_email
            }
        )
        
        return JsonResponse({"message": "Tutor updated successfully."}, status=200)
    else:
        log_api_action(
            request=request,
            action='UPDATE_TUTOR_SUCCESS',
            entity_type='Tutor',
            entity_ids=[tutor_id],
            success=True,
            additional_data={
                'changed_fields': {},
                'no_updates_made': True,
                'tutor_name': f"{tutor.staff.username if tutor.staff else 'Unknown'}",
                'tutor_email': tutor.tutor_email
            }
        )
        return JsonResponse({"message": "No fields updated."}, status=200)

# ==================== BULK IMPORT VOLUNTEERS/TUTORS ====================



@conditional_csrf
@api_view(['POST'])
def import_volunteers_endpoint(request):
    
    try:
        # Check feature flag - inline (no helper method)
        if not (os.environ.get("BLOCK_ACCESS_AFTER_APPROVAL", "False").lower() == "true"):
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="Bulk import feature is not enabled",
                status_code=403,
                additional_data={'reason': 'Feature disabled'}
            )
            return JsonResponse(
                {'error': 'Bulk import feature is not enabled'},
                status=403
            )
        
        # Check authentication (403 - not authenticated)
        user_id = request.session.get("user_id")
        if not user_id:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="Authentication credentials were not provided",
                status_code=403,
                additional_data={'reason': 'Not authenticated'}
            )
            return JsonResponse(
                {'error': 'Authentication credentials were not provided'},
                status=403
            )
        
        try:
            user = Staff.objects.get(staff_id=user_id)
        except Staff.DoesNotExist:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="User not found",
                status_code=403,
                additional_data={'reason': 'User does not exist'}
            )
            return JsonResponse(
                {'error': 'User not found'},
                status=403
            )
        
        # Check permission (401 - authenticated but no permission)
        has_tutor_create = has_permission(request, "tutors", "CREATE")
        has_volunteer_create = has_permission(request, "general_volunteer", "CREATE")
        
        if not (has_tutor_create or has_volunteer_create):
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_FAILED',
                success=False,
                error_message="You do not have permission to import volunteers or tutors",
                status_code=401,
                additional_data={'reason': 'Permission denied', 'user_email': user.email}
            )
            api_logger.critical(f"Unauthorized import attempt by {user.email} - lacks CREATE permission")
            return JsonResponse(
                {'error': 'You do not have permission to import volunteers or tutors'},
                status=401
            )
        
        api_logger.info(f"Import permission granted for user: {user.email}")
        
        # Get uploaded file
        if 'file' not in request.FILES:
            return JsonResponse(
                {'error': 'No file provided'},
                status=400
            )
        
        file: UploadedFile = request.FILES['file']
        
        # Validate file extension
        if not file.name.endswith('.xlsx'):
            return JsonResponse(
                {'error': 'File must be .xlsx format'},
                status=400
            )
        
        dry_run = request.data.get('dry_run', 'false').lower() == 'true'
        
        # Read Excel file
        try:
            df = pd.read_excel(file, dtype=str)
        except Exception as e:
            api_logger.error(f"Failed to read Excel file: {str(e)}")
            return JsonResponse(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=400
            )
        
        if df.empty:
            return JsonResponse(
                {'error': 'Excel file is empty'},
                status=400
            )
        
        # Get required roles
        try:
            general_volunteer_role = Role.objects.get(role_name="General Volunteer")
        except Role.DoesNotExist:
            return JsonResponse(
                {'error': 'Role "General Volunteer" not found in database'},
                status=500
            )
        
        try:
            tutor_role = Role.objects.get(role_name="Tutor")
        except Role.DoesNotExist:
            return JsonResponse(
                {'error': 'Role "Tutor" not found in database'},
                status=500
            )
        
        # Process records
        total_records = len(df)
        success_count = 0
        error_count = 0
        skipped_count = 0
        general_volunteer_count = 0
        tutor_with_tutee_count = 0
        tutor_no_tutee_count = 0
        pending_tutor_count = 0
        
        results = []
        
        for idx, row in df.iterrows():
            row_num = idx + 2
            # get_clean_string inline for names
            first_name_raw = row.get('שם פרטי', '')
            first_name = '' if (first_name_raw is None or pd.isna(first_name_raw) or str(first_name_raw).lower() == 'nan') else str(first_name_raw).strip()
            
            surname_raw = row.get('שם משפחה', '')
            surname = '' if (surname_raw is None or pd.isna(surname_raw) or str(surname_raw).lower() == 'nan') else str(surname_raw).strip()
            
            result = {
                'row_num': row_num,
                'first_name': first_name,
                'surname': surname,
                'email': '',
                'status': '',
                'record_type': '',
                'details': ''
            }
            
            try:
                # Parse data - INLINE, NO HELPER METHODS
                # get_clean_string inline
                id_number_raw = row.get('תעודת זהות', '')
                id_number = '' if (id_number_raw is None or pd.isna(id_number_raw) or str(id_number_raw).lower() == 'nan') else str(id_number_raw).strip()
                
                phone_raw = row.get('מספר טלפון', '')
                phone = '' if (phone_raw is None or pd.isna(phone_raw) or str(phone_raw).lower() == 'nan') else str(phone_raw).strip()
                
                # Format phone number - ensure leading zero and remove dashes/spaces
                if phone:
                    phone_normalized = phone.replace('-', '').replace(' ', '').strip()
                    # If doesn't start with 0, add it
                    if phone_normalized and not phone_normalized.startswith('0'):
                        phone_normalized = '0' + phone_normalized
                    # Validate it's 10 digits
                    if phone_normalized.isdigit() and len(phone_normalized) == 10:
                        phone = f"{phone_normalized[:3]}-{phone_normalized[3:]}"  # Format as XXX-XXXXXXX
                    else:
                        phone = ''  # Invalid phone, set to empty
                else:
                    phone = ''
                
                # clean_email inline
                email_raw = row.get('מייל')
                if email_raw and not pd.isna(email_raw):
                    email = str(email_raw).strip().replace('\n', '').replace('\r', '')
                    if not email or email.lower() == 'nan':
                        email = None
                else:
                    email = None
                
                # clean_city inline
                city_raw = row.get('עיר מגורים', '')
                city_val = '' if (city_raw is None or pd.isna(city_raw) or str(city_raw).lower() == 'nan') else str(city_raw).strip()
                city_mapping = {
                    'תל אביב': 'תל אביב - יפו',
                    'מודיעין': 'מודיעין-מכבים-רעות',
                    'מודעין': 'מודיעין-מכבים-רעות',
                    'פתח תקוה': 'פתח תקווה',
                    'קריית אתא': 'קרית אתא',
                    'קריית נטפים': 'קרית נטפים',
                    'יהוד מונוסון': 'יהוד-מונוסון',
                    'קיבוץ חפץ חיים': 'חפץ חיים',
                    'מושב בני ראם': 'בני ראם',
                    'מושב חמד': 'חמד',
                    'מושב פורת': 'פורת',
                    'יד רמב״ם (מושב)': 'יד רמבם',
                    'יישוב נופים': 'נופים',
                    'מגד אל כרום': 'מג\'ד אל-כרום',
                    'גבעת שמואל אבל עושה שירות': 'גבעת שמואל',
                    'מושה טפחות': 'טפחות',
                    'עלי זהב לומד בשדרות': 'עלי זהב',
                    'ירושלים- תא': 'ירושלים',
                    'ראשל״צ': 'ראשון לציון',
                    'הדר גנים': 'גנות הדר',
                    'רעננה(מגדל עוז)': 'רעננה',
                }
                city = city_val.replace('\n', ' ').replace('\r', '').strip() if city_val else ''
                for separator in ['/', ',']:
                    if separator in city:
                        city = city.split(separator)[0].strip()
                if city in city_mapping:
                    city = city_mapping[city]
                else:
                    for key, value in city_mapping.items():
                        if key in city:
                            city = value
                            break
                
                # parse_birth_date inline
                date_val = row.get('תאריך לידה', '')
                if not date_val or pd.isna(date_val) or str(date_val).lower() == 'nan':
                    birth_date = None
                else:
                    try:
                        from datetime import datetime as dt
                        if isinstance(date_val, dt):
                            birth_date = date_val.date()
                        else:
                            date_str = str(date_val).strip()
                            birth_date = None
                            for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y', '%m/%d/%Y', '%m-%d-%Y', '%m.%d.%Y']:
                                try:
                                    parsed_date = dt.strptime(date_str, fmt).date()
                                    today = dt.now().date()
                                    age_check = today.year - parsed_date.year
                                    if (today.month, today.day) < (parsed_date.month, parsed_date.day):
                                        age_check -= 1
                                    if 13 <= age_check <= 120:
                                        birth_date = parsed_date
                                        break
                                except ValueError:
                                    continue
                    except Exception:
                        birth_date = None
                
                # calculate_age_from_birth_date inline
                if birth_date:
                    from datetime import datetime as dt
                    today = dt.now().date()
                    calculated_age = today.year - birth_date.year
                    if (today.month, today.day) < (birth_date.month, birth_date.day):
                        calculated_age -= 1
                    calculated_age = max(0, calculated_age)
                else:
                    calculated_age = 0
                
                # parse_gender inline
                gender_val = row.get('מין', '')
                if isinstance(gender_val, bool):
                    gender = gender_val
                elif isinstance(gender_val, str):
                    gender_lower = gender_val.lower().strip()
                    gender = gender_lower in ['true', 'נקבה', 'female', 'f', '1']
                else:
                    gender = False
                
                # parse_want_tutor inline
                want_tutor_val = row.get('סוג התנדבות', '')
                if isinstance(want_tutor_val, bool):
                    want_tutor = want_tutor_val
                elif isinstance(want_tutor_val, str):
                    val_lower = want_tutor_val.lower().strip()
                    want_tutor = val_lower in ['true', 'כן', 'yes', '1']
                else:
                    want_tutor = False
                
                # parse_status inline
                status_raw = row.get('סטטוס', '')
                if not status_raw or pd.isna(status_raw) or str(status_raw).lower() == 'nan':
                    status_val = None
                else:
                    status_val = str(status_raw).strip()
                
                # get_clean_string inline for comments
                volunteer_comment_raw = row.get('הערות המתנדב', '')
                volunteer_comment = '' if (volunteer_comment_raw is None or pd.isna(volunteer_comment_raw) or str(volunteer_comment_raw).lower() == 'nan') else str(volunteer_comment_raw).strip()
                
                coordinator_comment_raw = row.get('הערות הרכז', '')
                coordinator_comment = '' if (coordinator_comment_raw is None or pd.isna(coordinator_comment_raw) or str(coordinator_comment_raw).lower() == 'nan') else str(coordinator_comment_raw).strip()
                
                result['email'] = email or ''
                
                # Build comments
                signedup_comment_parts = []
                tutor_preferences = None
                
                if want_tutor:
                    if volunteer_comment:
                        tutor_preferences = volunteer_comment
                    if coordinator_comment:
                        signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
                else:
                    if volunteer_comment:
                        signedup_comment_parts.append(f"הערות מתנדב: {volunteer_comment}")
                    if coordinator_comment:
                        signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
                
                signedup_comment = ' | '.join(signedup_comment_parts) if signedup_comment_parts else None
                
                # Validate
                if not first_name or not surname:
                    result['status'] = 'Error'
                    result['details'] = 'חסר שם פרטי או שם משפחה'
                    error_count += 1
                    results.append(result)
                    continue
                
                if not id_number:
                    result['status'] = 'Error'
                    result['details'] = 'חסרה תעודת זהות'
                    error_count += 1
                    results.append(result)
                    continue
                
                try:
                    id_int = int(id_number)
                except ValueError:
                    result['status'] = 'Error'
                    result['details'] = f'תעודת זהות לא מספרית: {id_number}'
                    error_count += 1
                    results.append(result)
                    continue
                
                # Check duplicates
                if SignedUp.objects.filter(id=id_int).exists():
                    result['status'] = 'Skipped'
                    result['details'] = f'ת.ז. כבר קיימת במערכת'
                    skipped_count += 1
                    results.append(result)
                    continue
                
                if email and Staff.objects.filter(email=email).exists():
                    result['status'] = 'Skipped'
                    result['details'] = f'מייל כבר קיים במערכת'
                    skipped_count += 1
                    results.append(result)
                    continue
                
                # Determine record type
                if want_tutor:
                    if status_val == "יש חניך":
                        record_type = "חונך - יש חניך"
                    elif status_val == "אין חניך":
                        record_type = "חונך - אין חניך"
                    else:
                        record_type = f"חונך ממתין - {status_val or 'ללא סטטוס'}"
                else:
                    record_type = "מתנדב כללי"
                
                result['record_type'] = record_type
                
                if dry_run:
                    result['status'] = 'OK'
                    result['details'] = f'בדיקה בלבד: {record_type}'
                    success_count += 1
                    results.append(result)
                    continue
                
                # === CREATE RECORDS ===
                with transaction.atomic():
                    # 1. Create SignedUp
                    signedup = SignedUp.objects.create(
                        id=id_int,
                        first_name=first_name,
                        surname=surname,
                        age=calculated_age,
                        birth_date=birth_date,
                        gender=gender,
                        phone=phone,
                        city=city,
                        comment=signedup_comment,
                        email=email,
                        want_tutor=want_tutor,
                    )
                    
                    # 2. Create Staff with unique username
                    username = f"{first_name}_{surname}"
                    index = 1
                    original_username = username
                    while Staff.objects.filter(username=username).exists():
                        username = f"{original_username}_{index}"
                        index += 1
                    
                    staff = Staff.objects.create(
                        username=username,
                        email=email,
                        first_name=first_name,
                        last_name=surname,
                        created_at=now(),
                        registration_approved=True,
                        is_active=True,
                        deactivation_reason="suspended"
                    )
                    
                    # 3. Create role-specific record
                    if want_tutor:
                        staff.roles.add(tutor_role)
                        
                        if status_val == "יש חניך":
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="יש_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            tutor_with_tutee_count += 1
                        elif status_val == "אין חניך":
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="אין_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            tutor_no_tutee_count += 1
                        else:
                            Tutors.objects.create(
                                id_id=signedup.id,
                                staff=staff,
                                tutorship_status="אין_חניך",
                                tutor_email=email,
                                preferences=tutor_preferences,
                            )
                            Pending_Tutor.objects.create(
                                id_id=signedup.id,
                                pending_status="ממתין",
                            )
                            pending_tutor_count += 1
                    else:
                        staff.roles.add(general_volunteer_role)
                        General_Volunteer.objects.create(
                            id_id=signedup.id,
                            staff_id=staff.staff_id,
                            signupdate=now().date(),
                            comments="",
                        )
                        general_volunteer_count += 1
                
                result['status'] = 'OK'
                result['details'] = f'נוצר בהצלחה: {record_type}'
                success_count += 1
                results.append(result)
                
            except IntegrityError as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאת מסד נתונים: {str(e)}'
                error_count += 1
                results.append(result)
            except Exception as e:
                result['status'] = 'Error'
                result['details'] = f'שגיאה כללית: {str(e)}'
                error_count += 1
                results.append(result)
        
        # Create result Excel file - INLINE, NO HELPER METHODS
        wb = Workbook()
        ws = wb.active
        ws.title = "תוצאות ייבוא"
        
        headers = ['שורה', 'שם פרטי', 'שם משפחה', 'מייל', 'סטטוס ייבוא', 'סוג רשומה', 'פרטים']
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result.get('row_num', ''))
            ws.cell(row=row_idx, column=2, value=result['first_name'])
            ws.cell(row=row_idx, column=3, value=result['surname'])
            ws.cell(row=row_idx, column=4, value=result.get('email', ''))
            
            status_cell = ws.cell(row=row_idx, column=5, value=result['status'])
            if result['status'] == 'OK':
                status_cell.fill = ok_fill
            elif result['status'] == 'Error':
                status_cell.fill = error_fill
            elif result['status'] == 'Warning':
                status_cell.fill = warning_fill
            
            ws.cell(row=row_idx, column=6, value=result.get('record_type', ''))
            ws.cell(row=row_idx, column=7, value=result.get('details', ''))
        
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 60
        ws.sheet_view.rightToLeft = True
        
        result_excel = BytesIO()
        wb.save(result_excel)
        result_excel.seek(0)
        
        # Log the import action
        if success_count > 0 or error_count > 0 or skipped_count > 0:
            log_api_action(
                request=request,
                action='CREATE_VOLUNTEER_SUCCESS' if success_count > 0 else 'CREATE_VOLUNTEER_FAILED',
                affected_tables=['childsmile_app_signedup', 'childsmile_app_staff', 'childsmile_app_general_volunteer', 'childsmile_app_tutors', 'childsmile_app_pending_tutor'],
                entity_type='Bulk Import',
                success=success_count > 0,
                additional_data={
                    'total_records': total_records,
                    'success_count': success_count,
                    'error_count': error_count,
                    'skipped_count': skipped_count,
                    'dry_run': dry_run,
                    'breakdown': {
                        'general_volunteer': general_volunteer_count,
                        'tutor_with_tutee': tutor_with_tutee_count,
                        'tutor_no_tutee': tutor_no_tutee_count,
                        'pending_tutor': pending_tutor_count,
                    }
                }
            )
        
        # Return results
        response_data = {
            'total': total_records,
            'success': success_count,
            'skipped': skipped_count,
            'error': error_count,
            'dry_run': dry_run,
            'breakdown': {
                'general_volunteer': general_volunteer_count,
                'tutor_with_tutee': tutor_with_tutee_count,
                'tutor_no_tutee': tutor_no_tutee_count,
                'pending_tutor': pending_tutor_count,
            },
            'message': f'✅ ייבוא הושלם: {success_count} מתוך {total_records} הרשומות יובאו בהצלחה' if success_count > 0 else '❌ לא הצליח לייבא רשומות',
            'has_errors': error_count > 0 or skipped_count > 0,
            'result_file_available': True
        }
        
        # If dry_run, return Excel file for preview
        if dry_run:
            response = FileResponse(result_excel, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            from datetime import datetime as dt
            response['Content-Disposition'] = f'attachment; filename="import_preview_{dt.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            return response
        
        # For regular import, return JSON with Excel file bytes embedded as base64
        # This allows frontend to download the results file
        result_excel.seek(0)
        excel_base64 = base64.b64encode(result_excel.read()).decode('utf-8')
        response_data['result_file'] = excel_base64
        from datetime import datetime as dt
        response_data['result_filename'] = f"import_results_{dt.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        return JsonResponse(response_data, status=200)
    
    except Exception as e:
        api_logger.error(f"Import endpoint error: {str(e)}")
        api_logger.error(f"Full traceback: {traceback.format_exc()}")
        log_api_action(
            request=request,
            action='CREATE_VOLUNTEER_FAILED',
            success=False,
            error_message=str(e),
            status_code=500,
            additional_data={'error_type': type(e).__name__}
        )
        return JsonResponse(
            {'error': f'Unexpected error: {str(e)}'},
            status=500
        )
