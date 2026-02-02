#!/usr/bin/env python3
"""
Volunteer Import Script
=======================
Imports volunteers from the production Excel file into the system.
Creates records in: SignedUp, Staff, General_Volunteer/Pending_Tutor/Tutors tables.

IMPORTANT: 
- NO TOTP codes sent
- NO registration approval tasks created
- NO emails sent
- All imported users are set as registration_approved=True
- Uses birth_date field (system calculates age automatically)

Logic based on want_tutor and status from Excel:
- want_tutor=False → General Volunteer (General_Volunteer table, role=General Volunteer)
- want_tutor=True + status="יש חניך" → Tutor with tutee (Tutors table, tutorship_status=יש_חניך)
- want_tutor=True + status="אין חניך" → Tutor without tutee (Tutors table, tutorship_status=אין_חניך)
- want_tutor=True + status="ממתין לראיון" or None → Pending Tutor (Tutors + Pending_Tutor tables)

Duplicate Handling:
- Skips records with existing ID (in SignedUp)
- Skips records with existing email (in Staff)

Usage:
    python import_volunteers.py <excel_file_path> [--dry-run] [--limit N]

Examples:
    # Dry run first 10 records:
    python import_volunteers.py "../מתנדבים.xlsx" --dry-run --limit 10
    
    # Import all records:
    python import_volunteers.py "../מתנדבים.xlsx" --limit 99999

Output:
    Creates an Excel file with import results in the same directory as the input file.
"""

import sys
import os
from datetime import datetime

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'childsmile.settings')

import django
django.setup()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from childsmile_app.models import (
    SignedUp, Staff, Role, General_Volunteer, Pending_Tutor, Tutors
)


def clean_email(email_val):
    """Clean email: strip whitespace and newlines."""
    if not email_val or pd.isna(email_val):
        return None
    email = str(email_val).strip().replace('\n', '').replace('\r', '')
    if not email or email.lower() == 'nan':
        return None
    return email


def parse_gender(gender_val):
    """Parse gender value to boolean (True = Female, False = Male)."""
    if isinstance(gender_val, bool):
        return gender_val
    if isinstance(gender_val, str):
        gender_lower = gender_val.lower().strip()
        if gender_lower in ['true', 'נקבה', 'female', 'f', '1']:
            return True
        return False
    return False


def parse_want_tutor(val):
    """Parse want_tutor value to boolean."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        val_lower = val.lower().strip()
        if val_lower in ['true', 'כן', 'yes', '1']:
            return True
    return False


def parse_status(status_val):
    """Parse status value from Excel - keep original value."""
    if not status_val or pd.isna(status_val) or str(status_val).lower() == 'nan':
        return None
    return str(status_val).strip()


def parse_birth_date(date_val):
    """Parse birth date from Excel. Returns date object or None."""
    if not date_val or pd.isna(date_val) or str(date_val).lower() == 'nan':
        return None
    try:
        # Try various date formats
        if isinstance(date_val, datetime):
            return date_val.date()
        date_str = str(date_val).strip()
        # Try different formats including datetime with time component
        for fmt in ['%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d.%m.%Y']:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        return None
    except Exception:
        return None


def calculate_age_from_birth_date(birth_date):
    """Calculate age from birth date."""
    if not birth_date:
        return 0
    today = datetime.now().date()
    age = today.year - birth_date.year
    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1
    return max(0, age)


def get_clean_string(val):
    """Get a clean string value, or empty string if None/NaN."""
    if val is None or pd.isna(val) or str(val).lower() == 'nan':
        return ''
    return str(val).strip()


def clean_city(city_val):
    """Clean city value - take only the first city if multiple are provided."""
    city = get_clean_string(city_val)
    if not city:
        return ''
    # Handle multiple cities separated by / or ,
    for separator in ['/', ',']:
        if separator in city:
            city = city.split(separator)[0].strip()
    # Remove any newlines
    city = city.replace('\n', ' ').replace('\r', '').strip()
    return city


def find_available_id(base_id):
    """Find an available ID starting from base_id, incrementing by 1 if taken."""
    current_id = base_id
    max_attempts = 1000  # Prevent infinite loop
    
    for _ in range(max_attempts):
        # Check if ID exists in SignedUp
        if not SignedUp.objects.filter(id=current_id).exists():
            return current_id, current_id != base_id
        current_id += 1
    
    return None, False


def create_result_excel(results, input_file_path):
    """Create an Excel file with import results."""
    # Generate output filename
    input_dir = os.path.dirname(input_file_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'import_results_{timestamp}.xlsx'
    output_path = os.path.join(input_dir if input_dir else '.', output_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "תוצאות ייבוא"
    
    # Headers
    headers = ['שורה', 'שם פרטי', 'שם משפחה', 'מייל', 'סטטוס ייבוא', 'סוג רשומה', 'פרטים']
    
    # Style for headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Style for status cells
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    warning_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result.get('row_num', ''))
        ws.cell(row=row_idx, column=2, value=result['first_name'])
        ws.cell(row=row_idx, column=3, value=result['surname'])
        ws.cell(row=row_idx, column=4, value=result.get('email', ''))
        
        status_cell = ws.cell(row=row_idx, column=5, value=result['status'])
        if result['status'] == 'OK':
            status_cell.fill = ok_fill
        elif result['status'] == 'Error':
            status_cell.fill = error_fill
        elif result['status'] == 'Warning':
            status_cell.fill = warning_fill
        
        ws.cell(row=row_idx, column=6, value=result.get('record_type', ''))
        ws.cell(row=row_idx, column=7, value=result.get('details', ''))
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 60
    
    # Set RTL for the sheet
    ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)
    return output_path


def import_volunteers(excel_path, dry_run=False, limit=99999):
    """
    Main import function.
    
    Args:
        excel_path: Path to the Excel file
        dry_run: If True, don't actually create records, just validate
        limit: Maximum number of rows to process (99999 = all)
    """
    print(f"\n{'='*70}")
    print(f"{'DRY RUN - ' if dry_run else ''}ייבוא מתנדבים מקובץ: {excel_path}")
    print(f"מגבלת שורות: {limit if limit < 99999 else 'הכל'}")
    print(f"{'='*70}\n")
    
    # Read Excel file - all as strings to preserve leading zeros
    df = pd.read_excel(excel_path, dtype=str)
    
    # Apply limit
    actual_rows = min(len(df), limit)
    df = df.head(actual_rows)
    
    print(f"נמצאו {len(df)} רשומות לעיבוד (מתוך סה\"כ {actual_rows} בקובץ)")
    print(f"עמודות: {list(df.columns)}\n")
    
    total_records = len(df)
    success_count = 0
    error_count = 0
    warning_count = 0
    skipped_count = 0  # New counter for skipped duplicates
    
    # Counters by type
    general_volunteer_count = 0
    tutor_with_tutee_count = 0
    tutor_no_tutee_count = 0
    pending_tutor_count = 0
    
    results = []
    
    # Get required roles
    try:
        general_volunteer_role = Role.objects.get(role_name="General Volunteer")
        print(f"✓ Found role: General Volunteer (id={general_volunteer_role.id})")
    except Role.DoesNotExist:
        print("❌ ERROR: Role 'General Volunteer' not found!")
        return None
    
    try:
        tutor_role = Role.objects.get(role_name="Tutor")
        print(f"✓ Found role: Tutor (id={tutor_role.id})")
    except Role.DoesNotExist:
        print("❌ ERROR: Role 'Tutor' not found!")
        return None
    
    print()
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        first_name = get_clean_string(row.get('שם פרטי', ''))
        surname = get_clean_string(row.get('שם משפחה', ''))
        
        result = {
            'row_num': row_num,
            'first_name': first_name,
            'surname': surname,
            'email': '',
            'status': '',
            'record_type': '',
            'details': ''
        }
        
        try:
            # Parse data
            id_number = get_clean_string(row.get('תעודת זהות', ''))
            phone = get_clean_string(row.get('מספר טלפון', ''))
            email = clean_email(row.get('מייל'))
            city = clean_city(row.get('עיר מגורים', ''))  # Use clean_city to handle multiple cities
            birth_date = parse_birth_date(row.get('תאריך לידה', ''))
            # Calculate age ONLY for validation (must be 13+), system will calc age from birth_date
            calculated_age = calculate_age_from_birth_date(birth_date)
            gender = parse_gender(row.get('מין', ''))
            want_tutor = parse_want_tutor(row.get('סוג התנדבות', ''))
            status = parse_status(row.get('סטטוס', ''))  # Keep original value from Excel
            volunteer_comment = get_clean_string(row.get('הערות המתנדב', ''))
            coordinator_comment = get_clean_string(row.get('הערות הרכז', ''))
            
            result['email'] = email or ''
            
            # Build comment for SignedUp based on want_tutor
            # If want_tutor=True: volunteer_comment goes to Tutors.preferences
            # If want_tutor=False: volunteer_comment goes to SignedUp.comment
            signedup_comment_parts = []
            tutor_preferences = None
            
            if want_tutor:
                # For tutors, volunteer comments are preferences
                if volunteer_comment:
                    tutor_preferences = volunteer_comment
                if coordinator_comment:
                    signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
            else:
                # For general volunteers, all comments go to SignedUp
                if volunteer_comment:
                    signedup_comment_parts.append(f"הערות מתנדב: {volunteer_comment}")
                if coordinator_comment:
                    signedup_comment_parts.append(f"הערות רכז: {coordinator_comment}")
            
            signedup_comment = ' | '.join(signedup_comment_parts) if signedup_comment_parts else None
            
            # Validate required fields
            if not first_name or not surname:
                result['status'] = 'Error'
                result['details'] = 'חסר שם פרטי או שם משפחה'
                error_count += 1
                results.append(result)
                continue
            
            if not id_number:
                result['status'] = 'Error'
                result['details'] = 'חסרה תעודת זהות'
                error_count += 1
                results.append(result)
                continue
            
            # Validate ID is numeric
            try:
                id_int = int(id_number)
            except ValueError:
                result['status'] = 'Error'
                result['details'] = f'תעודת זהות לא מספרית: {id_number}'
                error_count += 1
                results.append(result)
                continue
            
            # Check for duplicate ID - SKIP (don't error)
            if SignedUp.objects.filter(id=id_int).exists():
                result['status'] = 'Skipped'
                result['details'] = f'ת.ז. כבר קיימת במערכת - דילוג: {id_number}'
                skipped_count += 1
                results.append(result)
                continue
            
            # Check for duplicate email in Staff - SKIP (don't error)
            if email and Staff.objects.filter(email=email).exists():
                result['status'] = 'Skipped'
                result['details'] = f'מייל כבר קיים במערכת - דילוג: {email}'
                skipped_count += 1
                results.append(result)
                continue
            
            # ID is available - use as-is
            available_id = id_int
            
            # Determine record type based on want_tutor and original status from Excel
            # Keep status values as-is for display purposes
            if want_tutor:
                if status == "יש חניך":
                    record_type = "חונך - יש חניך"
                    tutorship_status_db = "יש_חניך"  # DB value with underscore
                elif status == "אין חניך":
                    record_type = "חונך - אין חניך"
                    tutorship_status_db = "אין_חניך"
                else:
                    # "ממתין לראיון" or None - needs interview/matching
                    record_type = f"חונך ממתין - {status or 'ללא סטטוס'}"
                    tutorship_status_db = "אין_חניך"  # Default for pending
            else:
                record_type = "מתנדב כללי"
                tutorship_status_db = None
            
            result['record_type'] = record_type
            
            if dry_run:
                # In dry run, just validate and report
                birth_info = f"תאריך לידה: {birth_date}" if birth_date else "אין תאריך לידה"
                result['status'] = 'OK'
                result['details'] = f'{birth_info} | גיל מחושב: {calculated_age}'
                success_count += 1
                results.append(result)
                continue
            
            # === START TRANSACTION ===
            with transaction.atomic():
                # 1. Create SignedUp record
                signedup = SignedUp.objects.create(
                    id=available_id,
                    first_name=first_name,
                    surname=surname,
                    age=calculated_age,  # Required field - calculated from birth_date
                    birth_date=birth_date,  # Also store birth_date for future recalculation
                    gender=gender,
                    phone=phone,
                    city=city,
                    comment=signedup_comment,
                    email=email,
                    want_tutor=want_tutor,
                )
                
                # 2. Create unique username
                username = f"{first_name}_{surname}"
                index = 1
                original_username = username
                while Staff.objects.filter(username=username).exists():
                    username = f"{original_username}_{index}"
                    index += 1
                
                # 3. Create Staff record (registration_approved=True)
                staff = Staff.objects.create(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=surname,
                    created_at=now(),
                    registration_approved=True,  # Already approved - imported by admin
                    is_active=True
                )
                
                # 4. Create role-specific record based on want_tutor and status
                if want_tutor:
                    # All tutors get Tutor role
                    staff.roles.add(tutor_role)
                    
                    if status == "יש חניך":
                        # Approved tutor with tutee
                        Tutors.objects.create(
                            id_id=signedup.id,
                            staff=staff,
                            tutorship_status="יש_חניך",
                            tutor_email=email,
                            preferences=tutor_preferences,
                        )
                        tutor_with_tutee_count += 1
                        
                    elif status == "אין חניך":
                        # Approved tutor without tutee
                        Tutors.objects.create(
                            id_id=signedup.id,
                            staff=staff,
                            tutorship_status="אין_חניך",
                            tutor_email=email,
                            preferences=tutor_preferences,
                        )
                        tutor_no_tutee_count += 1
                        
                    else:
                        # Pending tutor (status is "ממתין לראיון" or None)
                        # Create both Tutors and Pending_Tutor records
                        Tutors.objects.create(
                            id_id=signedup.id,
                            staff=staff,
                            tutorship_status="אין_חניך",  # No tutee yet
                            tutor_email=email,
                            preferences=tutor_preferences,
                        )
                        Pending_Tutor.objects.create(
                            id_id=signedup.id,
                            pending_status="ממתין",  # Waiting for tutee match
                        )
                        pending_tutor_count += 1
                else:
                    # General volunteer
                    staff.roles.add(general_volunteer_role)
                    General_Volunteer.objects.create(
                        id_id=signedup.id,
                        staff_id=staff.staff_id,
                        signupdate=now().date(),
                        comments="",
                    )
                    general_volunteer_count += 1
            
            # === END TRANSACTION ===
            
            result['status'] = 'OK'
            result['details'] = f'נוצר בהצלחה: {record_type}'
            success_count += 1
            results.append(result)
            
        except IntegrityError as e:
            result['status'] = 'Error'
            result['details'] = f'שגיאת מסד נתונים: {str(e)}'
            error_count += 1
            results.append(result)
            
        except Exception as e:
            result['status'] = 'Error'
            result['details'] = f'שגיאה כללית: {str(e)}'
            error_count += 1
            results.append(result)
    
    # Create result Excel file
    output_path = create_result_excel(results, excel_path)
    
    # Print summary
    print(f"\n{'='*70}")
    print(f"סיכום ייבוא {'(DRY RUN)' if dry_run else ''}")
    print(f"{'='*70}")
    print(f"סה\"כ רשומות בקובץ: {total_records}")
    print(f"הצלחה (OK): {success_count}")
    print(f"הצלחה עם אזהרות (Warning): {warning_count}")
    print(f"כשלון (Error): {error_count}")
    
    if not dry_run:
        print(f"\nפירוט לפי סוג:")
        print(f"  • מתנדבים כלליים: {general_volunteer_count}")
        print(f"  • חונכים עם חניך: {tutor_with_tutee_count}")
        print(f"  • חונכים ללא חניך: {tutor_no_tutee_count}")
        print(f"  • חונכים ממתינים להתאמה: {pending_tutor_count}")
    
    print(f"\nקובץ תוצאות נשמר ב: {output_path}")
    print(f"{'='*70}\n")
    
    return {
        'total': total_records,
        'success': success_count,
        'warning': warning_count,
        'error': error_count,
        'output_file': output_path,
        'breakdown': {
            'general_volunteer': general_volunteer_count,
            'tutor_with_tutee': tutor_with_tutee_count,
            'tutor_no_tutee': tutor_no_tutee_count,
            'pending_tutor': pending_tutor_count,
        }
    }


def main():
    import argparse
    
    parser = argparse.ArgumentParser(
        description='Import volunteers from Excel file',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  Dry run (validate only):
    python import_volunteers_v2.py --dry-run "../מתנדבים עלייה למערכת ללא עזבו מורחב.xlsx"
  
  Actual import:
    python import_volunteers_v2.py "../מתנדבים עלייה למערכת ללא עזבו מורחב.xlsx"
        """
    )
    parser.add_argument('excel_file', help='Path to the Excel file to import')
    parser.add_argument('--dry-run', action='store_true', 
                        help='Validate data without actually importing')
    parser.add_argument('--limit', type=int, default=99999,
                        help='Maximum number of records to import (default: all)')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel_file):
        print(f"Error: File not found: {args.excel_file}")
        sys.exit(1)
    
    import_volunteers(args.excel_file, dry_run=args.dry_run, limit=args.limit)


if __name__ == '__main__':
    main()
