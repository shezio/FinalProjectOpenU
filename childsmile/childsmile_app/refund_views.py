"""
Expense Refund Views (החזרי הוצאות)

Endpoints:
  GET    /api/refunds/                        – list (admin: all | volunteer: own only)
  POST   /api/refunds/create/                 – create new request (any authenticated user)
  PUT    /api/refunds/update/<refund_id>/     – update (admin: all fields | volunteer: own pending only)
  DELETE /api/refunds/delete/<refund_id>/     – hard delete from DB (admin only — use sparingly, history is preserved via status)
  GET    /api/refunds/upload-url/             – get Azure Blob SAS pre-signed upload URL
  POST   /api/refunds/import/                 – bulk import from Excel (admin only)
  GET    /api/refunds/phone-hint/             – return last successful payout phone for current user

Security rules:
  - All endpoints require an authenticated session (user_id in session).
  - Volunteers can only VIEW / EDIT their own records.
  - Any attempt to access another user's record returns 403 immediately.
  - Admins (System Administrator) bypass ownership checks.
  - Rejection is done via status update (בוטל/נדחה), NOT via delete — keeping history.
  - Hard delete is available to admins for genuine garbage/test data removal only.
"""

import json
import os
import uuid
import datetime
from decimal import Decimal, InvalidOperation

from django.conf import settings
from django.http import JsonResponse, FileResponse
from django.db import transaction
from django.db.models import Sum, Q
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from rest_framework.decorators import api_view

from .models import (
    Staff,
    Tasks,
    Task_Types,
    ExpenseRefund,
)
from .utils import (
    has_permission,
    is_admin,
    conditional_csrf,
    create_task_internal,
)
from .audit_utils import log_api_action
from .logger import api_logger
from .whatsapp_utils import (
    send_refund_new_request_to_admin_whatsapp,
    send_refund_status_update_to_volunteer_whatsapp,
    send_refund_payment_required_whatsapp,
)

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ──────────────────────────────────────────────────────────────────────────────

REFUND_STATUS_CHOICES = ['ממתין', 'אושר', 'אושר חלקית', 'שולם', 'בוטל/נדחה']
PAYMENT_METHOD_CHOICES = ['ביט', 'פייבוקס', 'העברה בנקאית', 'אשראי', 'מזומן']
COORDINATOR_CHOICES = ['נעם', 'טל', 'נבו', 'אורי', 'ליאם']
ISRAELI_PHONE_IMPORT = r'^0[2-9]\d{7,8}$'  # used in validation helper

# Admin phone for WhatsApp notifications — resolved from DB at runtime (same as utils.py)
LIAM_ADMIN_PHONE = os.getenv('LIAM_ADMIN_PHONE', '')  # kept as fallback, prefer DB lookup


def _get_liam_admin_phone():
    """Return Liam's phone from the Staff table (same lookup used elsewhere in the codebase)."""
    try:
        liam = Staff.objects.filter(first_name="ליאם", last_name="אביבי").first()
        return liam.staff_phone if liam and liam.staff_phone else LIAM_ADMIN_PHONE
    except Exception:
        return LIAM_ADMIN_PHONE

MAX_RECEIPT_SIZE_MB = 10


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────────────────────────

def _get_authenticated_user(request):
    """Return (staff_obj, None) or (None, error_JsonResponse)."""
    user_id = request.session.get("user_id")
    if not user_id:
        return None, JsonResponse(
            {"detail": "Authentication credentials were not provided."}, status=403
        )
    try:
        staff = Staff.objects.get(staff_id=user_id)
        return staff, None
    except Staff.DoesNotExist:
        return None, JsonResponse({"detail": "User not found."}, status=403)


def _ownership_guard(refund, staff, user_is_admin):
    """
    Returns an error JsonResponse if a non-admin tries to access a record
    that does not belong to them.  Returns None if access is allowed.
    """
    if user_is_admin:
        return None
    if refund.staff_id != staff.staff_id:
        return JsonResponse({"detail": "Forbidden."}, status=403)
    return None


def _validate_israeli_phone(phone):
    """Basic Israeli phone validation (10-digit, starts with 05/02/03/04/08/09)."""
    import re
    return bool(re.match(r'^0[2-9]\d{7,8}$', phone))


# ──────────────────────────────────────────────────────────────────────────────
# GET REFUNDS  (list)
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["GET"])
def get_refunds(request):
    api_logger.info("get_refunds called")

    staff, err = _get_authenticated_user(request)
    if err:
        log_api_action(request=request, action='VIEW_REFUNDS_FAILED',
                       success=False, error_message="Not authenticated", status_code=403)
        return err

    user_is_admin = is_admin(staff)

    # SECURITY: non-admins are hard-limited to their own records only
    if user_is_admin:
        qs = ExpenseRefund.objects.all().select_related('staff', 'related_task')
    else:
        qs = ExpenseRefund.objects.filter(staff=staff).select_related('staff', 'related_task')

    refunds_data = []
    for r in qs.order_by('-created_at'):
        refunds_data.append({
            "id": r.refund_id,
            "staff_id": r.staff_id,
            "staff_full_name": r.staff_full_name,
            "created_at": r.created_at.strftime("%d/%m/%Y %H:%M"),
            "updated_at": r.updated_at.strftime("%d/%m/%Y %H:%M"),
            "expense_date": r.expense_date.strftime("%Y-%m-%d"),
            "requested_amount": str(r.requested_amount),
            "approved_amount": str(r.approved_amount) if r.approved_amount is not None else None,
            "description": r.description,
            "volunteer_comment": r.volunteer_comment,
            "admin_comment": r.admin_comment,
            "approved_by": r.approved_by,
            "file_url": r.file_url,
            "status": r.status,
            "refund_method": r.refund_method,
            "phone_number": r.phone_number,
            "updated_by": r.updated_by,
            "related_task_id": r.related_task_id,
        })

    log_api_action(request=request, action='VIEW_REFUNDS', success=True,
                   status_code=200, entity_type='ExpenseRefund',
                   affected_tables=['childsmile_app_expenserefund'])
    return JsonResponse({"refunds": refunds_data}, status=200)


# ──────────────────────────────────────────────────────────────────────────────
# CREATE REFUND
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["POST"])
def create_refund(request):
    api_logger.info("create_refund called")

    staff, err = _get_authenticated_user(request)
    if err:
        log_api_action(request=request, action='CREATE_REFUND_FAILED',
                       success=False, error_message="Not authenticated", status_code=403)
        return err

    data = request.data

    # ── Validate required fields ──────────────────────────────────────────────
    required_fields = ['expense_date', 'requested_amount', 'description']
    missing = [f for f in required_fields if not data.get(f)]
    if missing:
        return JsonResponse({"error": f"Missing required fields: {', '.join(missing)}"}, status=400)

    # ── Validate requested_amount ─────────────────────────────────────────────
    try:
        requested_amount = Decimal(str(data['requested_amount']))
        if requested_amount <= 0:
            raise ValueError()
    except (InvalidOperation, ValueError):
        return JsonResponse({"error": "requested_amount must be a positive number."}, status=400)

    # ── Validate status (default ממתין) ───────────────────────────────────────
    status_val = data.get('status', 'ממתין')
    if status_val not in REFUND_STATUS_CHOICES:
        return JsonResponse({"error": f"Invalid status: {status_val}"}, status=400)

    # ── Validate refund_method: mandatory only when status is שולם ────────────
    refund_method = data.get('refund_method') or None
    if status_val == 'שולם' and not refund_method:
        return JsonResponse({"error": "refund_method is required when status is 'שולם'."}, status=400)
    if refund_method and refund_method not in PAYMENT_METHOD_CHOICES:
        return JsonResponse({"error": f"Invalid refund_method: {refund_method}"}, status=400)

    # ── Validate approved_amount: mandatory only for אושר חלקית ──────────────
    approved_amount = None
    if data.get('approved_amount'):
        try:
            approved_amount = Decimal(str(data['approved_amount']))
        except (InvalidOperation, ValueError):
            return JsonResponse({"error": "approved_amount must be a valid number."}, status=400)
    if status_val in ('אושר', 'אושר חלקית') and approved_amount is None:
        return JsonResponse({"error": "approved_amount is required when status is 'אושר' or 'אושר חלקית'."}, status=400)

    # ── Validate approved_by ──────────────────────────────────────────────────
    approved_by = data.get('approved_by') or None
    if approved_by and approved_by not in COORDINATOR_CHOICES:
        return JsonResponse({"error": f"Invalid approved_by value: {approved_by}"}, status=400)

    # ── Validate phone_number ─────────────────────────────────────────────────
    phone_number = data.get('phone_number') or None
    if phone_number and not _validate_israeli_phone(phone_number):
        return JsonResponse({"error": "phone_number must be a valid Israeli phone number (e.g. 0541234567)."}, status=400)

    # ── Resolve target staff (allow submitting on behalf of another user) ────────
    on_behalf_id = data.get('on_behalf_of_staff_id')
    target_staff = staff
    if on_behalf_id:
        try:
            target_staff = Staff.objects.get(staff_id=int(on_behalf_id))
        except (Staff.DoesNotExist, ValueError, TypeError):
            return JsonResponse({"error": f"Staff with id {on_behalf_id} not found."}, status=400)

    # ── Build staff full name ──────────────────────────────────────────────────
    staff_full_name = f"{target_staff.first_name} {target_staff.last_name}".strip()

    try:
        with transaction.atomic():
            refund = ExpenseRefund.objects.create(
                staff=target_staff,
                staff_full_name=staff_full_name,
                expense_date=data['expense_date'],
                requested_amount=requested_amount,
                approved_amount=approved_amount,
                description=data['description'],
                volunteer_comment=data.get('volunteer_comment') or None,
                admin_comment=None,
                approved_by=approved_by,
                file_url=data.get('file_url') or None,
                status=status_val,
                refund_method=refund_method,
                phone_number=phone_number,
                updated_by=staff.username,
            )

            # ── Save phone preference if volunteer requested it ───────────────
            save_phone = data.get('save_phone_for_future', False)
            if save_phone and phone_number:
                target_staff.staff_phone = phone_number
                target_staff.save(update_fields=['staff_phone'])

            # ── Auto-create task for Liam (System Administrator) ─────────────
            try:
                task_type = Task_Types.objects.filter(task_type="החזר הוצאות").first()
                if task_type:
                    liam_admins = Staff.objects.filter(roles__role_name='System Administrator').distinct()
                    for admin_staff in liam_admins:
                        due = datetime.date.today() + datetime.timedelta(days=7)
                        task_data = {
                            "type": task_type.id,
                            "description": f"בקשת החזר הוצאות - {staff_full_name} - {requested_amount}₪",
                            "due_date": due.strftime("%Y-%m-%d"),
                            "assigned_to": admin_staff.staff_id,
                            # Deep link stored in explanation field so Tasks board can render it
                            "explanation": f"/refunds?highlight={refund.refund_id}",
                            "user_info": {
                                "refund_id": refund.refund_id,
                                "staff_full_name": staff_full_name,
                                "requested_amount": str(requested_amount),
                                "expense_date": str(data['expense_date']),
                            },
                        }
                        task = create_task_internal(task_data)
                        refund.related_task = task
                        refund.save(update_fields=['related_task'])
                        break  # assign to first admin only
            except Exception as task_err:
                # Non-fatal: task creation failure should NOT abort the refund
                api_logger.error(f"Failed to create task for refund #{refund.refund_id}: {task_err}")

            # ── WhatsApp: notify admin of new request ─────────────────────────
            liam_phone = _get_liam_admin_phone()
            if liam_phone:
                try:
                    send_refund_new_request_to_admin_whatsapp(
                        admin_phone=liam_phone,
                        volunteer_full_name=staff_full_name,
                        requested_amount=requested_amount,
                    )
                except Exception as wa_err:
                    api_logger.error(f"WhatsApp admin notify failed for refund #{refund.refund_id}: {wa_err}")
            else:
                api_logger.warning("Liam's phone not found — skipping WhatsApp notify for new refund")

        log_api_action(request=request, action='CREATE_REFUND', success=True,
                       status_code=201, entity_type='ExpenseRefund',
                       entity_ids=[refund.refund_id],
                       affected_tables=['childsmile_app_expenserefund'])
        return JsonResponse({"message": "בקשת החזר נוצרה בהצלחה.", "id": refund.refund_id}, status=201)

    except Exception as e:
        api_logger.error(f"create_refund error: {e}")
        log_api_action(request=request, action='CREATE_REFUND_FAILED',
                       success=False, error_message=str(e), status_code=500)
        return JsonResponse({"error": "שגיאת שרת. אנא נסה שוב."}, status=500)


# ──────────────────────────────────────────────────────────────────────────────
# UPDATE REFUND
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["PUT"])
def update_refund(request, refund_id):
    api_logger.info(f"update_refund called for refund_id={refund_id}")

    staff, err = _get_authenticated_user(request)
    if err:
        log_api_action(request=request, action='UPDATE_REFUND_FAILED',
                       success=False, error_message="Not authenticated", status_code=403)
        return err

    try:
        refund = ExpenseRefund.objects.get(refund_id=refund_id)
    except ExpenseRefund.DoesNotExist:
        return JsonResponse({"error": "בקשת ההחזר לא נמצאה."}, status=404)

    user_is_admin = is_admin(staff)

    # SECURITY: ownership guard
    guard_err = _ownership_guard(refund, staff, user_is_admin)
    if guard_err:
        log_api_action(request=request, action='UPDATE_REFUND_FORBIDDEN',
                       success=False, error_message="Forbidden – not owner", status_code=403,
                       entity_ids=[refund_id])
        return guard_err

    data = request.data
    old_status = refund.status

    try:
        with transaction.atomic():
            # Fields volunteers can edit (only while status is ממתין)
            if not user_is_admin:
                if refund.status != 'ממתין':
                    return JsonResponse(
                        {"error": "ניתן לערוך בקשה רק כאשר הסטטוס הוא 'ממתין'."}, status=400
                    )
                allowed_fields = ['expense_date', 'requested_amount', 'description',
                                  'volunteer_comment', 'file_url', 'phone_number',
                                  'approved_by']
                for field in allowed_fields:
                    if field in data:
                        setattr(refund, field, data[field] or None)

                # Phone save preference
                if data.get('save_phone_for_future') and data.get('phone_number'):
                    phone = data['phone_number']
                    if not _validate_israeli_phone(phone):
                        return JsonResponse({"error": "מספר טלפון ישראלי לא תקין."}, status=400)
                    staff.staff_phone = phone
                    staff.save(update_fields=['staff_phone'])

            else:
                # Admin can update everything
                if 'expense_date' in data:
                    refund.expense_date = data['expense_date']
                if 'requested_amount' in data:
                    try:
                        refund.requested_amount = Decimal(str(data['requested_amount']))
                    except (InvalidOperation, ValueError):
                        return JsonResponse({"error": "requested_amount לא תקין."}, status=400)
                if 'approved_amount' in data:
                    try:
                        refund.approved_amount = Decimal(str(data['approved_amount'])) if data['approved_amount'] else None
                    except (InvalidOperation, ValueError):
                        return JsonResponse({"error": "approved_amount לא תקין."}, status=400)
                if 'description' in data:
                    refund.description = data['description']
                if 'volunteer_comment' in data:
                    refund.volunteer_comment = data['volunteer_comment'] or None
                if 'admin_comment' in data:
                    refund.admin_comment = data['admin_comment'] or None
                if 'approved_by' in data:
                    approved_by = data['approved_by']
                    if approved_by and approved_by not in COORDINATOR_CHOICES:
                        return JsonResponse({"error": f"approved_by לא תקין: {approved_by}"}, status=400)
                    refund.approved_by = approved_by or None
                if 'file_url' in data:
                    refund.file_url = data['file_url'] or None
                if 'phone_number' in data:
                    refund.phone_number = data['phone_number'] or None
                if 'refund_method' in data:
                    rm = data['refund_method']
                    if rm and rm not in PAYMENT_METHOD_CHOICES:
                        return JsonResponse({"error": f"refund_method לא תקין: {rm}"}, status=400)
                    refund.refund_method = rm or None

                # Status change with validations
                if 'status' in data:
                    new_status = data['status']
                    if new_status not in REFUND_STATUS_CHOICES:
                        return JsonResponse({"error": f"סטטוס לא תקין: {new_status}"}, status=400)
                    if new_status == 'שולם' and not refund.refund_method:
                        return JsonResponse(
                            {"error": "refund_method נדרש כאשר הסטטוס הוא 'שולם'."}, status=400
                        )
                    if new_status in ('אושר', 'אושר חלקית') and refund.approved_amount is None:
                        return JsonResponse(
                            {"error": "approved_amount נדרש כאשר הסטטוס הוא 'אושר' או 'אושר חלקית'."}, status=400
                        )
                    refund.status = new_status

            # auto_now=True on updated_at means Django sets it on every .save()
            refund.updated_by = staff.username
            refund.save()

            new_status = refund.status
            status_changed = old_status != new_status

            # ── WhatsApp: notify volunteer of status change ───────────────────
            if status_changed and user_is_admin:
                volunteer_phone = refund.phone_number or refund.staff.staff_phone
                if volunteer_phone:
                    try:
                        send_refund_status_update_to_volunteer_whatsapp(
                            volunteer_phone=volunteer_phone,
                            volunteer_full_name=refund.staff_full_name,
                            new_status=new_status,
                            approved_amount=refund.approved_amount if new_status != 'בוטל/נדחה' else 0,
                            admin_comment=refund.admin_comment,
                        )
                    except Exception as wa_err:
                        api_logger.error(f"WhatsApp volunteer notify failed for refund #{refund_id}: {wa_err}")

            # ── WhatsApp: notify אורי פלזנר to process payment ───────────────
            if status_changed and user_is_admin and new_status in ('אושר', 'אושר חלקית'):
                try:
                    from .models import SignedUp
                    uri_staff = Staff.objects.filter(email='oriplezner1@gmail.com').first()
                    uri_phone = uri_staff.staff_phone if uri_staff else None
                    # Fallback: look up phone in SignedUp table if Staff phone is empty
                    if not uri_phone:
                        uri_signedup = SignedUp.objects.filter(email='oriplezner1@gmail.com').first()
                        uri_phone = uri_signedup.phone if uri_signedup and uri_signedup.phone else None
                    if uri_phone:
                        payment_phone = refund.phone_number or refund.staff.staff_phone
                        send_refund_payment_required_whatsapp(
                            uri_phone=uri_phone,
                            volunteer_full_name=refund.staff_full_name,
                            approved_amount=refund.approved_amount,
                            payment_phone=payment_phone,
                            refund_method=refund.refund_method,
                            approved_by=refund.approved_by,
                            status=new_status,
                        )
                    else:
                        api_logger.warning("אורי פלזנר (oriplezner1@gmail.com) not found in Staff or SignedUp, or has no phone — skipping payment WhatsApp")
                except Exception as wa_err:
                    api_logger.error(f"WhatsApp Uri payment notify failed for refund #{refund_id}: {wa_err}")

        log_api_action(request=request, action='UPDATE_REFUND', success=True,
                       status_code=200, entity_type='ExpenseRefund',
                       entity_ids=[refund_id],
                       affected_tables=['childsmile_app_expenserefund'])
        return JsonResponse({"message": "בקשת ההחזר עודכנה בהצלחה."}, status=200)

    except Exception as e:
        api_logger.error(f"update_refund error: {e}")
        log_api_action(request=request, action='UPDATE_REFUND_FAILED',
                       success=False, error_message=str(e), status_code=500,
                       entity_ids=[refund_id])
        return JsonResponse({"error": "שגיאת שרת. אנא נסה שוב."}, status=500)


# ──────────────────────────────────────────────────────────────────────────────
# DELETE REFUND  (hard delete — admin only, use sparingly)
# Rejection should be done via status = 'בוטל/נדחה' to preserve history.
# This endpoint is for genuine garbage/test data removal only.
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["DELETE"])
def delete_refund(request, refund_id):
    api_logger.info(f"delete_refund called for refund_id={refund_id}")

    staff, err = _get_authenticated_user(request)
    if err:
        log_api_action(request=request, action='DELETE_REFUND_FAILED',
                       success=False, error_message="Not authenticated", status_code=403)
        return err

    user_is_admin = is_admin(staff)
    if not user_is_admin:
        log_api_action(request=request, action='DELETE_REFUND_FORBIDDEN',
                       success=False, error_message="Forbidden – admins only", status_code=403,
                       entity_ids=[refund_id])
        return JsonResponse({"detail": "Forbidden."}, status=403)

    try:
        refund = ExpenseRefund.objects.get(refund_id=refund_id)
    except ExpenseRefund.DoesNotExist:
        return JsonResponse({"error": "בקשת ההחזר לא נמצאה."}, status=404)

    # ── Delete blob from Azure before removing the DB row ─────────────────────
    if refund.file_url and settings.IS_PROD:
        try:
            from azure.storage.blob import BlobServiceClient
            conn_str = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
            container = os.getenv('AZURE_REFUNDS_CONTAINER', 'refund-receipts')
            if conn_str:
                # Extract blob name from URL: everything after the container segment
                # URL format: https://<account>.blob.core.windows.net/<container>/<blob_name>
                url_path = refund.file_url.split(f"/{container}/", 1)
                if len(url_path) == 2:
                    blob_name = url_path[1].split("?")[0]  # strip any SAS query string
                    service = BlobServiceClient.from_connection_string(conn_str)
                    service.get_blob_client(container=container, blob=blob_name).delete_blob()
                    api_logger.info(f"delete_refund: blob deleted — {blob_name}")
        except Exception as blob_err:
            # Log but don't block DB deletion — blob has 7-day lifecycle anyway
            api_logger.warning(f"delete_refund: blob deletion failed for refund {refund_id}: {blob_err}")

    refund.delete()

    log_api_action(request=request, action='DELETE_REFUND', success=True,
                   status_code=200, entity_type='ExpenseRefund', entity_ids=[refund_id],
                   affected_tables=['childsmile_app_expenserefund'],
                   additional_data={"deleted_by": staff.username})
    return JsonResponse({"message": "בקשת ההחזר נמחקה בהצלחה."}, status=200)


# ──────────────────────────────────────────────────────────────────────────────
# PHONE HINT  (for form pre-fill)
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["GET"])
def get_refund_phone_hint(request):
    """
    Returns the best phone number to pre-fill for Bit/Paybox payout.
    Priority:
      1. Last phone used in a successfully PAID refund for this user
      2. Staff profile phone (staff_phone field)
    """
    api_logger.info("get_refund_phone_hint called")

    staff, err = _get_authenticated_user(request)
    if err:
        return err

    # Check historical paid refunds for a previous phone
    last_paid = (
        ExpenseRefund.objects
        .filter(staff=staff, status='שולם', phone_number__isnull=False)
        .exclude(phone_number='')
        .order_by('-updated_at')
        .values_list('phone_number', flat=True)
        .first()
    )

    phone = last_paid or staff.staff_phone or ''
    return JsonResponse({"phone_hint": phone}, status=200)


# ──────────────────────────────────────────────────────────────────────────────
# AZURE BLOB STORAGE — Pre-signed upload URL
# ──────────────────────────────────────────────────────────────────────────────

@conditional_csrf
@api_view(["GET"])
def get_receipt_upload_url(request):
    """
    PROD: Returns a short-lived Azure Blob SAS URL for direct browser upload.
    LOCAL (IS_PROD=False): Saves the file via a POST to /api/refunds/upload-local/ instead —
      but this endpoint still returns a fake upload_url pointing to the local upload endpoint
      so the frontend flow stays identical.

    File size enforcement (10 MB) is done on the frontend before requesting the URL.
    """
    api_logger.info("get_receipt_upload_url called")

    staff, err = _get_authenticated_user(request)
    if err:
        return err

    filename = request.GET.get('filename', 'receipt')
    safe_filename = os.path.basename(filename)

    if not settings.IS_PROD:
        # ── Local dev: return a fake upload_url pointing at our own upload endpoint
        unique_id = uuid.uuid4().hex[:12]
        blob_name = f"refunds/{staff.staff_id}/{unique_id}_{safe_filename}"
        # The frontend will PUT to this URL — we'll intercept it in the local upload view
        local_upload_url = request.build_absolute_uri(f"/api/refunds/upload-local/?blob={blob_name}")
        local_blob_url = request.build_absolute_uri(f"/api/refunds/file/{blob_name}")
        return JsonResponse({"upload_url": local_upload_url, "blob_url": local_blob_url}, status=200)

    # ── PROD: Azure Blob Storage
    try:
        from azure.storage.blob import generate_blob_sas, BlobSasPermissions
        from datetime import timezone as dt_timezone

        connection_string = os.getenv('AZURE_STORAGE_CONNECTION_STRING')
        account_name = os.getenv('AZURE_STORAGE_ACCOUNT_NAME')
        account_key = os.getenv('AZURE_STORAGE_ACCOUNT_KEY')
        container = os.getenv('AZURE_REFUNDS_CONTAINER', 'refund-receipts')

        if not all([connection_string, account_name, account_key]):
            api_logger.error("Azure Blob Storage credentials not configured.")
            return JsonResponse({"error": "Receipt upload is not configured."}, status=503)

        ts = datetime.datetime.now(dt_timezone.utc).strftime("%Y%m%d%H%M%S")
        blob_name = f"refunds/{staff.staff_id}/{ts}_{safe_filename}"

        expiry = datetime.datetime.now(dt_timezone.utc) + datetime.timedelta(minutes=15)
        sas_token = generate_blob_sas(
            account_name=account_name,
            container_name=container,
            blob_name=blob_name,
            account_key=account_key,
            permission=BlobSasPermissions(write=True, create=True),
            expiry=expiry,
        )

        upload_url = f"https://{account_name}.blob.core.windows.net/{container}/{blob_name}?{sas_token}"
        blob_url = f"https://{account_name}.blob.core.windows.net/{container}/{blob_name}"
        return JsonResponse({"upload_url": upload_url, "blob_url": blob_url}, status=200)

    except ImportError:
        api_logger.error("azure-storage-blob package not installed.")
        return JsonResponse({"error": "Azure SDK not installed."}, status=503)
    except Exception as e:
        api_logger.error(f"get_receipt_upload_url error: {e}")
        return JsonResponse({"error": "שגיאה ביצירת URL להעלאה."}, status=500)


@conditional_csrf
@api_view(["PUT"])
def local_upload_receipt(request):
    """
    LOCAL DEV ONLY — receives the raw file body PUT by the frontend and saves it
    to Django's default file storage (MEDIA_ROOT/refunds/...).
    """
    if settings.IS_PROD:
        return JsonResponse({"error": "Not available in production."}, status=403)

    blob_name = request.GET.get('blob', '')
    if not blob_name:
        return JsonResponse({"error": "Missing blob param."}, status=400)

    try:
        content = request.body
        path = default_storage.save(blob_name, ContentFile(content))
        api_logger.info(f"local_upload_receipt saved: {path}")
        return JsonResponse({"saved": path}, status=200)
    except Exception as e:
        api_logger.error(f"local_upload_receipt error: {e}")
        return JsonResponse({"error": str(e)}, status=500)


@api_view(["GET"])
def serve_local_receipt(request, blob_path):
    """
    LOCAL DEV ONLY — serves a previously uploaded receipt file.
    """
    if settings.IS_PROD:
        return JsonResponse({"error": "Not available in production."}, status=403)
    try:
        f = default_storage.open(blob_path)
        return FileResponse(f)
    except Exception:
        return JsonResponse({"error": "File not found."}, status=404)


# ──────────────────────────────────────────────────────────────────────────────
# IMPORT  (admin only)
# ──────────────────────────────────────────────────────────────────────────────

REFUND_COLUMN_MAP = {
    # Hebrew headers
    'שם מלא': 'staff_full_name',
    'אימייל': 'email',
    'תאריך הוצאה': 'expense_date',
    'סכום מבוקש': 'requested_amount',
    'סכום שאושר': 'approved_amount',
    'תיאור': 'description',
    'הערת מתנדב': 'volunteer_comment',
    'הערת מנהל': 'admin_comment',
    'אושר על ידי': 'approved_by',
    'סטטוס': 'status',
    'אמצעי תשלום': 'refund_method',
    'טלפון': 'phone_number',
    # English headers
    'Full Name': 'staff_full_name',
    'Email': 'email',
    'Expense Date': 'expense_date',
    'Requested Amount': 'requested_amount',
    'Approved Amount': 'approved_amount',
    'Description': 'description',
    'Volunteer Comment': 'volunteer_comment',
    'Admin Comment': 'admin_comment',
    'Approved By': 'approved_by',
    'Status': 'status',
    'Payment Method': 'refund_method',
    'Phone': 'phone_number',
}

REFUND_VALID_STATUSES = ['ממתין', 'אושר', 'אושר חלקית', 'שולם', 'בוטל/נדחה']


def _parse_date_refund(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.date()
    if isinstance(val, datetime.date):
        return val
    val = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y'):
        try:
            return datetime.datetime.strptime(val, fmt).date()
        except ValueError:
            continue
    return None


def _parse_decimal_refund(val):
    if val is None or str(val).strip() == '':
        return None
    try:
        return Decimal(str(val).replace(',', '.').strip())
    except InvalidOperation:
        return None


@conditional_csrf
@api_view(["POST"])
def import_refunds(request):
    if 1 == 2:
        pass
    return JsonResponse({"error": "Feature disabled."}, status=403)
    """
    Import historical expense refund data from an uploaded Excel file.
    Admin only.

    Expected Excel columns (Hebrew headers):
        שם מלא / אימייל / תאריך הוצאה / סכום מבוקש / סכום שאושר /
        תיאור / הערת מתנדב / הערת מנהל / אושר על ידי / סטטוס / אמצעי תשלום / טלפון
    """
    api_logger.info("import_refunds called")

    staff, err = _get_authenticated_user(request)
    if err:
        return err

    if not is_admin(staff):
        return JsonResponse({"detail": "Forbidden – admins only."}, status=403)

    if 'file' not in request.FILES:
        return JsonResponse({"error": "לא נמצא קובץ. יש לצרף קובץ Excel."}, status=400)

    uploaded = request.FILES['file']
    if not uploaded.name.endswith('.xlsx'):
        return JsonResponse({"error": "קובץ חייב להיות בפורמט .xlsx"}, status=400)

    try:
        import openpyxl
    except ImportError:
        return JsonResponse({"error": "openpyxl לא מותקן בשרת."}, status=503)

    try:
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(uploaded.read()), read_only=True, data_only=True)
        ws = wb.active
    except Exception as e:
        return JsonResponse({"error": f"שגיאה בפתיחת קובץ: {e}"}, status=400)

    # Parse header row
    headers = [str(cell.value).strip() if cell.value else '' for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col_index = {REFUND_COLUMN_MAP[h]: idx for idx, h in enumerate(headers) if h in REFUND_COLUMN_MAP}

    required_cols = {'email', 'expense_date', 'requested_amount', 'description'}
    missing_cols = required_cols - set(col_index.keys())
    if missing_cols:
        return JsonResponse({"error": f"עמודות חסרות: {missing_cols}"}, status=400)

    created = skipped = 0
    errors = []

    with transaction.atomic():
        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            def get(col_name):
                idx = col_index.get(col_name)
                return row[idx] if idx is not None and idx < len(row) else None

            email = str(get('email') or '').strip().lower()
            expense_date = _parse_date_refund(get('expense_date'))
            requested_amount = _parse_decimal_refund(get('requested_amount'))
            description = str(get('description') or '').strip()
            staff_full_name = str(get('staff_full_name') or '').strip()

            # Skip blank rows
            if not any([email, expense_date, requested_amount]):
                continue

            # Validate mandatory
            if not email or not expense_date or not requested_amount or not description:
                errors.append(f"שורה {row_num}: שדות חובה חסרים — דולג")
                skipped += 1
                continue

            # Look up staff
            try:
                row_staff = Staff.objects.get(email=email)
            except Staff.DoesNotExist:
                errors.append(f"שורה {row_num}: לא נמצא איש צוות עם אימייל {email!r} — דולג")
                skipped += 1
                continue
            except Staff.MultipleObjectsReturned:
                row_staff = Staff.objects.filter(email=email).first()

            # Duplicate detection
            if ExpenseRefund.objects.filter(
                staff=row_staff,
                expense_date=expense_date,
                requested_amount=requested_amount,
            ).exists():
                skipped += 1
                continue

            approved_amount = _parse_decimal_refund(get('approved_amount'))
            status = str(get('status') or '').strip() or 'ממתין'
            if status not in REFUND_VALID_STATUSES:
                status = 'ממתין'

            ExpenseRefund.objects.create(
                staff=row_staff,
                staff_full_name=staff_full_name or f"{row_staff.first_name} {row_staff.last_name}",
                expense_date=expense_date,
                requested_amount=requested_amount,
                approved_amount=approved_amount,
                description=description,
                volunteer_comment=str(get('volunteer_comment') or '').strip() or None,
                admin_comment=str(get('admin_comment') or '').strip() or None,
                approved_by=str(get('approved_by') or '').strip() or None,
                status=status,
                refund_method=str(get('refund_method') or '').strip() or None,
                phone_number=str(get('phone_number') or '').strip() or None,
                updated_by=staff.username,
            )
            created += 1

    log_api_action(
        request=request,
        action='IMPORT_REFUNDS_SUCCESS',
        affected_tables=['childsmile_app_expenserefund'],
        entity_type='ExpenseRefund',
        success=True,
        additional_data={'created': created, 'skipped': skipped, 'error_count': len(errors)},
    )

    return JsonResponse({
        "created": created,
        "skipped": skipped,
        "errors": errors,
        "message": f"✅ יובאו {created} רשומות, דולגו {skipped}.",
    }, status=200)


# placeholder so old code doesn't break – reports are frontend-only
def get_refund_report(request):
    """
    Generates an expense summary report for the given period.

    Query params:
      period  = 'monthly' | 'quarterly' | 'annual'  (default: monthly)
      year    = YYYY   (default: current year)
      month   = 1-12  (required for monthly)
      quarter = 1-4   (required for quarterly)
      format  = 'json' | 'excel' | 'pdf'  (default: json)

    Admin only.

    PDF export: uses reportlab (official donor layout).
    Excel export: uses openpyxl (internal audit format).
    Both formats are outlined below with clear PLACEHOLDER comments
    so they can be implemented when the libraries are confirmed available.
    """
    api_logger.info("get_refund_report called")

    staff, err = _get_authenticated_user(request)
    if err:
        return err

    if not is_admin(staff):
        return JsonResponse({"detail": "Forbidden – admins only."}, status=403)

    period = request.GET.get('period', 'monthly')
    year = int(request.GET.get('year', datetime.date.today().year))
    fmt = request.GET.get('format', 'json')

    # Build date range
    if period == 'monthly':
        month = int(request.GET.get('month', datetime.date.today().month))
        date_from = datetime.date(year, month, 1)
        last_day = (date_from.replace(day=28) + datetime.timedelta(days=4))
        date_to = last_day - datetime.timedelta(days=last_day.day)
        period_label = f"{month:02d}/{year}"
    elif period == 'quarterly':
        quarter = int(request.GET.get('quarter', ((datetime.date.today().month - 1) // 3) + 1))
        month_start = (quarter - 1) * 3 + 1
        date_from = datetime.date(year, month_start, 1)
        month_end = month_start + 2
        last_day = (datetime.date(year, month_end, 28) + datetime.timedelta(days=4))
        date_to = last_day - datetime.timedelta(days=last_day.day)
        period_label = f"Q{quarter}/{year}"
    elif period == 'annual':
        date_from = datetime.date(year, 1, 1)
        date_to = datetime.date(year, 12, 31)
        period_label = str(year)
    else:
        return JsonResponse({"error": "period must be monthly, quarterly, or annual."}, status=400)

    qs = ExpenseRefund.objects.filter(
        expense_date__gte=date_from,
        expense_date__lte=date_to,
    ).select_related('staff').order_by('expense_date')

    totals = qs.aggregate(
        total_requested=Sum('requested_amount'),
        total_approved=Sum('approved_amount'),
    )

    rows = []
    for r in qs:
        rows.append({
            "id": r.refund_id,
            "staff_full_name": r.staff_full_name,
            "expense_date": r.expense_date.strftime("%Y-%m-%d"),
            "requested_amount": str(r.requested_amount),
            "approved_amount": str(r.approved_amount) if r.approved_amount else None,
            "status": r.status,
            "refund_method": r.refund_method,
            "description": r.description,
            "approved_by": r.approved_by,
        })

    report_data = {
        "period": period_label,
        "date_from": str(date_from),
        "date_to": str(date_to),
        "total_requested": str(totals['total_requested'] or 0),
        "total_approved": str(totals['total_approved'] or 0),
        "record_count": len(rows),
        "rows": rows,
    }

    if fmt == 'json':
        log_api_action(request=request, action='EXPORT_REPORT_REFUNDS_JSON', success=True,
                       status_code=200, report_name=f"expense_report_{period_label}")
        return JsonResponse(report_data, status=200)

    if fmt == 'excel':
        # ── EXCEL EXPORT (openpyxl) ──────────────────────────────────────────
        # PLACEHOLDER: Build xlsx with openpyxl matching the internal audit format.
        # Columns: מזהה | שם מלא | תאריך הוצאה | סכום מבוקש | סכום שאושר | סטטוס | אמצעי תשלום | תיאור | אושר על ידי
        # Sheet title: f"דוח החזרי הוצאות – {period_label}"
        # Return as HttpResponse with content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        try:
            import openpyxl
            from django.http import HttpResponse
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = f"דוח {period_label}"
            headers = ['מזהה', 'שם מלא', 'תאריך הוצאה', 'סכום מבוקש', 'סכום שאושר',
                       'סטטוס', 'אמצעי תשלום', 'תיאור', 'אושר על ידי']
            ws.append(headers)
            for row in rows:
                ws.append([
                    row['id'], row['staff_full_name'], row['expense_date'],
                    row['requested_amount'], row['approved_amount'] or '',
                    row['status'], row['refund_method'] or '',
                    row['description'], row['approved_by'] or '',
                ])
            from io import BytesIO
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            log_api_action(request=request, action='EXPORT_REPORT_REFUNDS_EXCEL', success=True,
                           status_code=200, report_name=f"expense_report_{period_label}")
            response = HttpResponse(
                buf.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="expense_report_{period_label}.xlsx"'
            return response
        except ImportError:
            return JsonResponse({"error": "openpyxl not installed."}, status=503)

    if fmt == 'pdf':
        # ── PDF EXPORT (reportlab) ───────────────────────────────────────────
        # PLACEHOLDER: Build PDF with official donor layout using reportlab.
        # Layout: Header with org logo/name, period title, summary box (totals),
        #         table of records with RTL text support via python-bidi,
        #         footer with generation timestamp.
        # Return as HttpResponse with content_type='application/pdf'
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.lib import colors
            from io import BytesIO
            from django.http import HttpResponse

            buf = BytesIO()
            doc = SimpleDocTemplate(buf, pagesize=A4)
            styles = getSampleStyleSheet()
            elements = []

            elements.append(Paragraph(f"דוח החזרי הוצאות – {period_label}", styles['Title']))
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(
                f"סה\"כ מבוקש: {report_data['total_requested']} ₪ | "
                f"סה\"כ אושר: {report_data['total_approved']} ₪ | "
                f"מספר רשומות: {report_data['record_count']}",
                styles['Normal']
            ))
            elements.append(Spacer(1, 12))

            table_data = [['מזהה', 'שם', 'תאריך', 'מבוקש', 'אושר', 'סטטוס']]
            for row in rows:
                table_data.append([
                    str(row['id']), row['staff_full_name'], row['expense_date'],
                    row['requested_amount'], row['approved_amount'] or '-', row['status'],
                ])
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6366f1')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
            ]))
            elements.append(t)
            doc.build(elements)
            buf.seek(0)
            log_api_action(request=request, action='EXPORT_REPORT_REFUNDS_PDF', success=True,
                           status_code=200, report_name=f"expense_report_{period_label}")
            response = HttpResponse(buf.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="expense_report_{period_label}.pdf"'
            return response
        except ImportError:
            return JsonResponse({"error": "reportlab not installed."}, status=503)

    return JsonResponse({"error": "format must be json, excel, or pdf."}, status=400)
