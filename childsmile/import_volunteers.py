#!/usr/bin/env python3
"""
Volunteer Import Script
=======================
Imports volunteers from an Excel file mimicking the registration flow.
Creates records in: SignedUp, Staff, General_Volunteer/Pending_Tutor tables.
For users with status "ממתין לראיון" - creates interview tasks.

IMPORTANT: NO TOTP codes sent, NO registration approval tasks created.
All imported users are set as registration_approved=True.

Usage:
    python import_volunteers.py <excel_file_path>

Example:
    python import_volunteers.py "../מתנדבים עלייה למערכת ללא עזבו מורחב.xlsx"

Output:
    Creates an Excel file with import results in the same directory as the input file.
"""

import sys
import os
from datetime import datetime, timedelta

# Setup Django environment
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'childsmile.settings')

import django
django.setup()

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.db import transaction, IntegrityError
from django.utils.timezone import now
from childsmile_app.models import (
    SignedUp, Staff, Role, General_Volunteer, Pending_Tutor, 
    Tutors, Task, Task_Types, TutorshipStatus
)


# Excel column mapping (Hebrew to English)
COLUMN_MAP = {
    'שם פרטי': 'first_name',
    'שם משפחה': 'surname',
    'מין': 'gender',
    'תאריך לידה': 'birth_date',
    'גיל': 'age',
    'תעודת זהות': 'id_number',
    'מספר טלפון': 'phone',
    'מייל': 'email',
    'עיר מגורים': 'city',
    'סוג התנדבות': 'want_tutor',
    'סטטוס': 'status',
    'הערות המתנדב': 'comment',
    'הערות הרכז': 'coordinator_comment',
}


def fix_phone(phone_str):
    """Add leading zero to phone if missing."""
    if phone_str and not str(phone_str).startswith('0'):
        return '0' + str(phone_str)
    return str(phone_str) if phone_str else ''


def parse_gender(gender_val):
    """Parse gender value to boolean (True = Female, False = Male)."""
    if isinstance(gender_val, bool):
        return gender_val
    if isinstance(gender_val, str):
        gender_lower = gender_val.lower().strip()
        if gender_lower in ['true', 'נקבה', 'female', 'f', '1']:
            return True
        return False
    return False


def parse_want_tutor(val):
    """Parse want_tutor value to boolean."""
    if isinstance(val, bool):
        return val
    if isinstance(val, str):
        val_lower = val.lower().strip()
        if val_lower in ['true', 'כן', 'yes', '1']:
            return True
    return False


def parse_status(status_val):
    """Parse status value from Excel."""
    if not status_val or str(status_val).lower() == 'nan':
        return None
    return str(status_val).strip()


def find_available_id(base_id):
    """Find an available ID starting from base_id, incrementing by 1 if taken."""
    current_id = base_id
    max_attempts = 1000  # Prevent infinite loop
    
    for _ in range(max_attempts):
        # Check if ID exists in SignedUp
        if not SignedUp.objects.filter(id=current_id).exists():
            return current_id, current_id != base_id
        current_id += 1
    
    return None, False


def create_result_excel(results, input_file_path):
    """Create an Excel file with import results."""
    # Generate output filename
    input_dir = os.path.dirname(input_file_path)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'import_results_{timestamp}.xlsx'
    output_path = os.path.join(input_dir if input_dir else '.', output_filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "תוצאות ייבוא"
    
    # Headers
    headers = ['שם פרטי', 'שם משפחה', 'סטטוס ייבוא', 'פרטים']
    
    # Style for headers
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    # Style for status cells
    ok_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    error_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    diff_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Write data
    for row_idx, result in enumerate(results, 2):
        ws.cell(row=row_idx, column=1, value=result['first_name'])
        ws.cell(row=row_idx, column=2, value=result['surname'])
        
        status_cell = ws.cell(row=row_idx, column=3, value=result['status'])
        if result['status'] == 'OK':
            status_cell.fill = ok_fill
        elif result['status'] == 'Error':
            status_cell.fill = error_fill
        elif result['status'] == 'Diff':
            status_cell.fill = diff_fill
        
        ws.cell(row=row_idx, column=4, value=result['details'])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 60
    
    # Set RTL for the sheet
    ws.sheet_view.rightToLeft = True
    
    wb.save(output_path)
    return output_path


def create_interview_task(pending_tutor, tutor_coordinators):
    """Create interview tasks for tutor coordinators."""
    try:
        task_type = Task_Types.objects.filter(task_type="ראיון מועמד לחונכות").first()
        if not task_type:
            return False, "Task type 'ראיון מועמד לחונכות' not found"
        
        tasks_created = 0
        for coordinator in tutor_coordinators:
            Task.objects.create(
                description="ראיון מועמד לחונכות",
                due_date=now().date() + timedelta(days=7),
                completed=False,
                task_type=task_type,
                assigned_to=coordinator,
                pending_tutor=pending_tutor
            )
            tasks_created += 1
        
        return True, f"Created {tasks_created} interview tasks"
    except Exception as e:
        return False, str(e)


def import_volunteers(excel_path):
    """Main import function - mimics registration flow without TOTP."""
    print(f"\n{'='*60}")
    print(f"מתחיל ייבוא מתנדבים מקובץ: {excel_path}")
    print(f"{'='*60}\n")
    
    # Read Excel file - all as strings to preserve leading zeros
    df = pd.read_excel(excel_path, dtype=str)
    
    total_records = len(df)
    success_count = 0
    error_count = 0
    diff_count = 0
    
    results = []
    
    # Get required objects
    try:
        general_volunteer_role = Role.objects.get(role_name="General Volunteer")
    except Role.DoesNotExist:
        print("ERROR: Role 'General Volunteer' not found!")
        return
    
    try:
        tutor_role = Role.objects.get(role_name="Tutor")
    except Role.DoesNotExist:
        print("ERROR: Role 'Tutor' not found!")
        return
    
    # Get tutor coordinators for interview tasks
    tutor_coordinator_role = Role.objects.filter(role_name="Tutors Coordinator").first()
    tutor_coordinators = Staff.objects.filter(roles=tutor_coordinator_role) if tutor_coordinator_role else []
    
    for idx, row in df.iterrows():
        row_num = idx + 2  # Excel row number (1-indexed + header)
        first_name = str(row.get('שם פרטי', '')).strip()
        surname = str(row.get('שם משפחה', '')).strip()
        
        result = {
            'first_name': first_name,
            'surname': surname,
            'status': '',
            'details': ''
        }
        
        try:
            # Parse data
            id_number = str(row.get('תעודת זהות', '')).strip()
            phone = fix_phone(row.get('מספר טלפון', ''))
            email = str(row.get('מייל', '')).strip() if pd.notna(row.get('מייל')) else None
            city = str(row.get('עיר מגורים', '')).strip()
            age_str = row.get('גיל', '')
            gender = parse_gender(row.get('מין', ''))
            want_tutor = parse_want_tutor(row.get('סוג התנדבות', ''))
            status = parse_status(row.get('סטטוס', ''))
            
            # Combine comments
            comment_parts = []
            if pd.notna(row.get('הערות המתנדב')) and str(row.get('הערות המתנדב')).strip():
                comment_parts.append(f"הערות מתנדב: {row.get('הערות המתנדב')}")
            if pd.notna(row.get('הערות הרכז')) and str(row.get('הערות הרכז')).strip():
                comment_parts.append(f"הערות רכז: {row.get('הערות הרכז')}")
            comment = ' | '.join(comment_parts) if comment_parts else None
            
            # Clean email
            if email:
                email = email.strip().replace('\n', '').replace('\r', '')
                if not email or email.lower() == 'nan':
                    email = None
            
            # Parse age
            try:
                age = int(float(age_str)) if age_str and age_str.lower() != 'nan' else 0
            except (ValueError, TypeError):
                age = 0
            
            # Validate required fields
            if not first_name or not surname:
                result['status'] = 'Error'
                result['details'] = 'חסר שם פרטי או שם משפחה'
                error_count += 1
                results.append(result)
                continue
            
            if not id_number or len(id_number) != 9:
                result['status'] = 'Error'
                result['details'] = f'תעודת זהות לא תקינה: {id_number}'
                error_count += 1
                results.append(result)
                continue
            
            # Convert ID to integer
            try:
                id_int = int(id_number)
            except ValueError:
                result['status'] = 'Error'
                result['details'] = f'תעודת זהות לא מספרית: {id_number}'
                error_count += 1
                results.append(result)
                continue
            
            # Track changes
            diffs = []
            original_phone = str(row.get('מספר טלפון', '')).strip()
            if phone != original_phone:
                diffs.append(f'טלפון תוקן: {original_phone} → {phone}')
            
            # Find available ID
            available_id, id_changed = find_available_id(id_int)
            
            if available_id is None:
                result['status'] = 'Error'
                result['details'] = f'לא נמצא ID פנוי (ניסיון מ-{id_int})'
                error_count += 1
                results.append(result)
                continue
            
            if id_changed:
                diffs.append(f'ID שונה: {id_number} → {available_id}')
            
            # Check for duplicate email in Staff
            if email and Staff.objects.filter(email=email).exists():
                result['status'] = 'Error'
                result['details'] = f'כתובת מייל כבר קיימת במערכת: {email}'
                error_count += 1
                results.append(result)
                continue
            
            # === START TRANSACTION - Mimic registration flow ===
            with transaction.atomic():
                # 1. Create SignedUp record
                signedup = SignedUp.objects.create(
                    id=available_id,
                    first_name=first_name,
                    surname=surname,
                    age=age,
                    gender=gender,
                    phone=phone,
                    city=city,
                    comment=comment,
                    email=email,
                    want_tutor=want_tutor,
                )
                
                # 2. Create unique username
                username = f"{first_name}_{surname}"
                index = 1
                original_username = username
                while Staff.objects.filter(username=username).exists():
                    username = f"{original_username}_{index}"
                    index += 1
                
                # 3. Create Staff record (registration_approved=True - no need for approval)
                staff = Staff.objects.create(
                    username=username,
                    email=email,
                    first_name=first_name,
                    last_name=surname,
                    created_at=now(),
                    registration_approved=True,  # Already approved - imported by admin
                    is_active=True
                )
                
                # 4. Determine what to create based on want_tutor and status
                # Status meanings:
                # - "ממתין לראיון" = Pending interview -> Pending_Tutor + interview task
                # - "יש חניך" = Approved tutor WITH tutee -> Tutors table with יש_חניך
                # - "אין חניך" = Approved tutor WITHOUT tutee -> Tutors table with אין_חניך
                # - None/other = General volunteer or Pending_Tutor based on want_tutor
                
                interview_task_created = False
                
                if status == "יש חניך":
                    # Approved tutor with tutee - create Tutors record
                    staff.roles.add(tutor_role)
                    Tutors.objects.create(
                        id_id=signedup.id,
                        staff=staff,
                        tutorship_status=TutorshipStatus.HAS_TUTEE,  # יש_חניך
                        tutor_email=email,
                    )
                    diffs.append('חונך מאושר עם חניך')
                    
                elif status == "אין חניך":
                    # Approved tutor without tutee - create Tutors record
                    staff.roles.add(tutor_role)
                    Tutors.objects.create(
                        id_id=signedup.id,
                        staff=staff,
                        tutorship_status=TutorshipStatus.NO_TUTEE,  # אין_חניך
                        tutor_email=email,
                    )
                    diffs.append('חונך מאושר ללא חניך')
                    
                elif want_tutor:
                    # Wants to be tutor but not yet approved
                    staff.roles.add(general_volunteer_role)
                    pending_status = "ממתין לראיון" if status == "ממתין לראיון" else "ממתין"
                    pending_tutor = Pending_Tutor.objects.create(
                        id_id=signedup.id,
                        pending_status=pending_status,
                    )
                    
                    # Create interview task for "ממתין לראיון" status
                    if status == "ממתין לראיון" and tutor_coordinators:
                        success, msg = create_interview_task(pending_tutor, tutor_coordinators)
                        if success:
                            interview_task_created = True
                            diffs.append('נוצרה משימת ראיון')
                        else:
                            diffs.append(f'שגיאה ביצירת משימת ראיון: {msg}')
                else:
                    # General volunteer (doesn't want to be tutor)
                    staff.roles.add(general_volunteer_role)
                    General_Volunteer.objects.create(
                        id_id=signedup.id,
                        staff_id=staff.staff_id,
                        signupdate=now().date(),
                        comments="",
                    )
            
            # === END TRANSACTION ===
            
            # Determine final status
            if diffs:
                result['status'] = 'Diff'
                result['details'] = ' | '.join(diffs)
                diff_count += 1
            else:
                result['status'] = 'OK'
                result['details'] = ''
                success_count += 1
            
            results.append(result)
            
        except IntegrityError as e:
            result['status'] = 'Error'
            result['details'] = f'שגיאת מסד נתונים: {str(e)}'
            error_count += 1
            results.append(result)
            
        except Exception as e:
            result['status'] = 'Error'
            result['details'] = f'שגיאה כללית: {str(e)}'
            error_count += 1
            results.append(result)
    
    # Create result Excel file
    output_path = create_result_excel(results, excel_path)
    
    # Print summary
    print(f"\n{'='*60}")
    print(f"סיכום ייבוא")
    print(f"{'='*60}")
    print(f"סה\"כ רשומות: {total_records}")
    print(f"הצלחה (OK): {success_count}")
    print(f"הצלחה עם שינויים (Diff): {diff_count}")
    print(f"כשלון (Error): {error_count}")
    print(f"\nקובץ תוצאות נשמר ב: {output_path}")
    print(f"{'='*60}\n")
    
    return {
        'total': total_records,
        'success': success_count,
        'diff': diff_count,
        'error': error_count,
        'output_file': output_path
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python import_volunteers.py <excel_file_path>")
        print("Example: python import_volunteers.py '../מתנדבים עלייה למערכת ללא עזבו מורחב.xlsx'")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    if not os.path.exists(excel_path):
        print(f"Error: File not found: {excel_path}")
        sys.exit(1)
    
    import_volunteers(excel_path)


if __name__ == '__main__':
    main()
